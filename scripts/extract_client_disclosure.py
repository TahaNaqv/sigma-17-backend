"""Extract the client's IFRS 17 disclosure sheets into client_source_extract.json.

Provenance tool for the movement mapping. ``client_source_extract.json`` is the
authoritative per-(line, bucket) source map lifted verbatim from the client's
``Module2_Final_Output.xlsx`` (sheets ``Gross`` / ``RI``); ``mapping_source.json`` is
generated from it, and the whole movement projection rests on both. Until now the
extract was committed without the script that produced it, so it could be trusted but
not re-derived — this closes that gap (plan §5 E6).

Default mode is **verify**: re-extract from the workbook and diff against the committed
JSON, exiting non-zero on any difference. That is the mode CI should run, so a silent
edit to the extract — or a client file that no longer says what we recorded — fails loudly.

    python scripts/extract_client_disclosure.py [path/to/Module2_Final_Output.xlsx]
    python scripts/extract_client_disclosure.py --write [path]

Per bucket cell the extraction records one of:
    computed  — references ``'IFRS Summary'!<COL>1``; source_expr resolves each reference
                to that column's header name (the client's own cells point at row 1, the
                header row, which is why they cache as text in their workbook)
    subtotal  — a same-sheet formula (SUM/arithmetic over other disclosure rows)
    override  — an INDEX/MATCH into the external ``Template Info`` workbook; detail is the
                external column letter, mapped to a stable override key in _meta
    const     — a numeric literal (a flattened cell)
    empty     — no content, but the row carries a direction sign for this bucket

NB: several client cells hold formula *text* with the leading ``=`` missing (e.g. Gross
C28). Those are classified on content, not on cell type — dropping them would silently
lose real source mappings.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

BACKEND = Path(__file__).resolve().parent.parent
PKG = BACKEND / "module2_engine" / "movement"
OUT = PKG / "client_source_extract.json"
DEFAULT_SRC = (
    BACKEND.parent / "sigma-17-desktop-app" / "Output Module 2" / "Module2_Final_Output.xlsx"
)

SUMMARY_SHEET = "IFRS Summary"
SIGNS = {"+", "-", "+/-", "-/+"}

#: sheet -> (label column, {bucket column: bucket name}, {sign column: value column}, total column)
SHEETS: dict[str, tuple] = {
    "Gross": (
        "A",
        {"C": "LRC_excl_LC", "E": "Loss_Component", "G": "LIC_excl_RA", "I": "Risk_Adjustment"},
        {"B": "C", "D": "E", "F": "G", "H": "I"},
        "J",
    ),
    "RI": (
        "B",
        {
            "D": "Assets_Remaining_Coverage",
            "F": "Loss_Recovery_Component",
            "H": "Amounts_Recoverable_IC",
            "J": "Risk_Adjustment",
        },
        {"C": "D", "E": "F", "G": "H", "I": "J"},
        "K",
    ),
}

#: External Template Info column letter -> stable override input key. These are the client's
#: manual judgement inputs; the engine fills them from the movement-override dataset.
OVERRIDE_INPUTS_RI = {
    "AD": "ri_provision_nonperformance_change",
    "BI": "ri_loss_recovery_new_onerous",
    "BJ": "ri_loss_recovery_reversal_amortization",
    "BK": "ri_loss_recovery_assumption_change",
    "BL": "ri_finance_pnl",
    "DV": "ri_pdr_accrual_reserve_bop",
    "DX": "ri_methodology_diff_loss_recovery_bop",
    "EZ": "ri_accrual_reserve_specify",
}

_SUMMARY_REF = re.compile(r"'?IFRS Summary'!\$?([A-Z]{1,3})\$?1\b")
_TEMPLATE_INFO = re.compile(r"Template Info'!\$([A-Z]{1,3})\$?:")

# ── the four note sheets (plan §2) ───────────────────────────────────────────
# Pure re-presentation of Gross/RI: every populated cell is a literal, a reference into
# the movement sheets, a reference into another note, or a sum within the note itself.

NOTES_OUT = PKG / "notes_source.json"

#: note sheet -> (label column, first label row, {column letter: column key}, source movement sheet)
NOTE_SHEETS: dict[str, tuple] = {
    "Gross_Note": (
        "A", 8,
        {"B": "LRC_excl_LC", "C": "Loss_Component", "D": "LIC_excl_RA",
         "E": "Risk_Adjustment", "F": "Total"},
        "Gross",
    ),
    "RI_Note": (
        "A", 8,
        {"B": "Assets_Remaining_Coverage", "C": "Loss_Recovery_Component",
         "D": "Amounts_Recoverable_IC", "E": "Risk_Adjustment", "F": "Total"},
        "RI",
    ),
    "IS": ("B", 5, {"C": "Total"}, None),
    "BS": ("B", 4, {"C": "Total"}, None),
}

#: Movement-sheet column letter -> bucket, per sheet. Lets a note reference such as
#: ``Gross!G32`` resolve to (line id, bucket) rather than to a cell address.
MOVEMENT_COLUMNS: dict[str, dict[str, str]] = {
    "Gross": {"C": "LRC_excl_LC", "E": "Loss_Component", "G": "LIC_excl_RA",
              "I": "Risk_Adjustment", "J": "Total"},
    "RI": {"D": "Assets_Remaining_Coverage", "F": "Loss_Recovery_Component",
           "H": "Amounts_Recoverable_IC", "J": "Risk_Adjustment", "K": "Total"},
}

_MOVEMENT_REF = re.compile(r"\b(Gross|RI)!\$?([A-Z]{1,3})\$?(\d+)\b")
_NOTE_REF = re.compile(r"\b(Gross_Note|RI_Note|IS|BS)!\$?([A-Z]{1,3})\$?(\d+)\b")
_CELL_RANGE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)")
_CELL = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")


def _cell_text(value) -> str:
    """Cell content as a string. Array formulas expose their source under ``.text``."""
    if value is None:
        return ""
    if hasattr(value, "text"):  # openpyxl ArrayFormula
        return str(value.text)
    return str(value)


def _summary_headers(wb) -> dict[str, str]:
    """IFRS Summary column letter -> header name (row 1)."""
    ws = wb[SUMMARY_SHEET]
    return {
        c.column_letter: str(c.value).strip()
        for c in next(ws.iter_rows(min_row=1, max_row=1))
        if c.value is not None
    }


def _classify(text: str, headers: dict[str, str]) -> dict:
    """One bucket cell -> its recorded source entry (without the sign)."""
    if not text:
        return {"type": "empty", "detail": None}

    if _SUMMARY_REF.search(text):
        refs = [[col, headers.get(col, col)] for col in _SUMMARY_REF.findall(text)]
        expr = _SUMMARY_REF.sub(lambda m: headers.get(m.group(1), m.group(1)), text).lstrip("=")
        return {"type": "computed", "source_expr": expr, "ref_cols": refs}

    m = _TEMPLATE_INFO.search(text)
    if m:
        return {"type": "override", "detail": m.group(1)}

    if text.startswith("="):
        return {"type": "subtotal", "detail": text}

    try:
        return {"type": "const", "detail": float(text)}
    except ValueError:
        # Formula text with the leading "=" lost and no recognised reference — record it
        # verbatim rather than silently dropping a real source mapping.
        return {"type": "subtotal", "detail": text}


def extract(src: Path) -> dict:
    wb = load_workbook(src, data_only=False)
    headers = _summary_headers(wb)
    sheets: dict[str, list] = {}

    for name, (label_col, bucket_cols, sign_cols, total_col) in SHEETS.items():
        ws = wb[name]
        lines: list[dict] = []
        for row in range(1, ws.max_row + 1):
            raw = ws[f"{label_col}{row}"].value
            if not isinstance(raw, str) or not raw.strip():
                continue
            label = raw.replace("\n", " ").strip()
            level = len(raw) - len(raw.lstrip(" "))

            buckets: dict[str, dict] = {}
            for col, bucket in bucket_cols.items():
                sign_col = next((s for s, target in sign_cols.items() if target == col), None)
                sign = _cell_text(ws[f"{sign_col}{row}"].value).strip() if sign_col else ""
                sign = sign if sign in SIGNS else None
                entry = _classify(_cell_text(ws[f"{col}{row}"].value).strip(), headers)
                if sign is None and entry["type"] == "empty":
                    continue  # nothing recorded for this bucket on this row
                buckets[bucket] = {"sign": sign, **entry}

            total = _cell_text(ws[f"{total_col}{row}"].value).strip()
            total_formula = total if total.startswith("=") else None

            if not buckets and total_formula is None:
                kind = "section"
            elif total_formula is not None:
                kind = "subtotal"
            else:
                kind = "input"

            lines.append(
                {
                    "row": row,
                    "id": re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_") or "line",
                    "label": label,
                    "level": level,
                    "kind": kind,
                    "buckets": buckets,
                    "total_formula": total_formula,
                }
            )
        sheets[name] = lines

    return {
        "_meta": {
            "purpose": (
                "Verbatim per-(line,bucket) source extraction from the CLIENT disclosure file. "
                "Authoritative sign-off input for the movement mapping. Do not hand-edit; "
                "regenerate via scripts/extract_client_disclosure.py."
            ),
            "source_file": f"sigma-17-desktop-app/Output Module 2/{src.name} (sheets Gross, RI)",
            "source_sha256": hashlib.sha256(src.read_bytes()).hexdigest(),
            "extracted_bucket_columns": {
                name: list(spec[1].values()) for name, spec in SHEETS.items()
            },
            "override_inputs_RI": OVERRIDE_INPUTS_RI,
        },
        "sheets": sheets,
    }


def _classify_note_cell(text: str, *, row: int, columns: dict[str, str]) -> dict | None:
    """One note cell -> its recorded source, or None when the cell carries nothing.

    Several client cells hold formula *text* with the leading ``=`` missing (Gross_Note
    D26/E26, RI_Note B25). They are classified on content and flagged ``literal_text`` so
    the defect stays visible in the artifact instead of being silently normalised away.
    """
    if not text:
        return None
    if text == "-":
        return {"kind": "dash"}

    literal_text = not text.startswith("=")
    body = text.lstrip("=")

    movement = _MOVEMENT_REF.findall(body)
    if movement:
        terms = [
            {"sheet": sheet, "bucket": MOVEMENT_COLUMNS[sheet].get(col, col), "row": int(r)}
            for sheet, col, r in movement
        ]
        entry = {"kind": "movement", "terms": terms}
        if literal_text:
            entry["literal_text"] = text
        return entry

    note = _NOTE_REF.findall(body)
    if note:
        target, col, r = note[0]
        entry = {
            "kind": "note",
            "ref": {"note": target, "column": NOTE_SHEETS[target][2].get(col, col), "row": int(r)},
        }
        if literal_text:
            entry["literal_text"] = text
        return entry

    try:
        return {"kind": "const", "value": float(body)}
    except ValueError:
        pass

    # Same-sheet arithmetic: horizontal (this row, across buckets) is the row total;
    # vertical (this column, other rows) aggregates other note lines.
    cells: list[tuple[str, int]] = []
    expanded = _CELL_RANGE.sub(
        lambda m: " ".join(
            f"{c}{r}"
            for c in _column_span(m.group(1), m.group(3))
            for r in range(int(m.group(2)), int(m.group(4)) + 1)
        ),
        body,
    )
    for col, r in _CELL.findall(expanded):
        cells.append((col, int(r)))
    if not cells:
        return {"kind": "unparsed", "detail": text}

    if all(r == row for _, r in cells):
        return {"kind": "row_total",
                "columns": [columns.get(c, c) for c, _ in cells if columns.get(c, c) != "Total"]}

    seen: list[int] = []
    for _, r in cells:
        if r not in seen:
            seen.append(r)
    entry = {"kind": "sum", "rows": seen}
    if literal_text:
        entry["literal_text"] = text
    return entry


def _column_span(start: str, end: str) -> list[str]:
    """Column letters from start to end inclusive (single-letter columns suffice here)."""
    if len(start) != 1 or len(end) != 1:
        return [start] if start == end else [start, end]
    return [chr(c) for c in range(ord(start), ord(end) + 1)]


def extract_notes(src: Path) -> dict:
    """Extract the four note sheets verbatim. Corrections live in the curated deviation
    layer of ``notes_schema.py``, never here — this artifact stays faithful to the client."""
    wb = load_workbook(src, data_only=False)
    sheets: dict[str, dict] = {}

    for name, (label_col, first_row, columns, source_sheet) in NOTE_SHEETS.items():
        ws = wb[name]
        lines: list[dict] = []
        used: set[str] = set()
        for row in range(first_row, ws.max_row + 1):
            raw = ws[f"{label_col}{row}"].value
            if not isinstance(raw, str) or not raw.strip():
                continue
            label = raw.replace("\n", " ").strip()

            sources: dict[str, dict] = {}
            for col, key in columns.items():
                entry = _classify_note_cell(
                    _cell_text(ws[f"{col}{row}"].value).strip(), row=row, columns=columns
                )
                if entry is not None:
                    sources[key] = entry

            if not sources:
                kind = "section"
            elif any(s["kind"] == "sum" for s in sources.values()):
                kind = "subtotal"
            else:
                kind = "input"

            base = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_") or "line"
            lid, i = base, 2
            while lid in used:
                lid, i = f"{base}_{i}", i + 1
            used.add(lid)

            lines.append({"row": row, "id": lid, "label": label, "kind": kind, "sources": sources})

        sheets[name] = {
            "title": _cell_text(ws["A1"].value).strip() or _cell_text(ws["A2"].value).strip(),
            "columns": list(columns.values()),
            "source_sheet": source_sheet,
            "lines": lines,
        }

    return {
        "_meta": {
            "purpose": (
                "Verbatim extraction of the client's four note sheets (Gross_Note, RI_Note, "
                "IS, BS) — a pure re-presentation of the Gross/RI movement sheets. Faithful "
                "to the client file INCLUDING its defects; corrections live in the curated "
                "deviation layer of notes_schema.py. Do not hand-edit; regenerate via "
                "scripts/extract_client_disclosure.py --write."
            ),
            "source_file": f"sigma-17-desktop-app/Output Module 2/{src.name} "
            f"(sheets {', '.join(NOTE_SHEETS)})",
            "source_sha256": hashlib.sha256(src.read_bytes()).hexdigest(),
            "movement_columns": MOVEMENT_COLUMNS,
        },
        "sheets": sheets,
    }


def _diff(actual: dict, committed: dict) -> list[str]:
    """Content differences, ignoring _meta.source_sha256 (the workbook legitimately gains
    sheets over time; what must not drift is what we recorded about Gross/RI)."""
    problems: list[str] = []
    a_meta = {k: v for k, v in actual["_meta"].items() if k != "source_sha256"}
    c_meta = {k: v for k, v in committed["_meta"].items() if k != "source_sha256"}
    if a_meta != c_meta:
        problems.append("_meta differs (excluding source_sha256)")

    for name in sorted(set(actual["sheets"]) | set(committed["sheets"])):
        a = {ln["row"]: ln for ln in actual["sheets"].get(name, [])}
        c = {ln["row"]: ln for ln in committed["sheets"].get(name, [])}
        for row in sorted(set(a) - set(c)):
            problems.append(f"{name} r{row}: extracted but not committed")
        for row in sorted(set(c) - set(a)):
            problems.append(f"{name} r{row}: committed but no longer in the workbook")
        for row in sorted(set(a) & set(c)):
            if a[row] != c[row]:
                for key in sorted(set(a[row]) | set(c[row])):
                    if a[row].get(key) != c[row].get(key):
                        problems.append(
                            f"{name} r{row}.{key}: workbook={a[row].get(key)!r} "
                            f"committed={c[row].get(key)!r}"
                        )
    return problems


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("source", nargs="?", default=str(DEFAULT_SRC), help="client .xlsx")
    ap.add_argument("--write", action="store_true", help="rewrite the committed extract")
    args = ap.parse_args()

    src = Path(args.source)
    if not src.exists():
        print(f"client workbook not found: {src}", file=sys.stderr)
        return 2

    data = extract(src)
    cells = sum(len(ln["buckets"]) for sheet in data["sheets"].values() for ln in sheet)
    lines = sum(len(sheet) for sheet in data["sheets"].values())
    notes = extract_notes(src)
    note_lines = sum(len(sh["lines"]) for sh in notes["sheets"].values())
    note_cells = sum(len(ln["sources"]) for sh in notes["sheets"].values() for ln in sh["lines"])

    if args.write:
        # No trailing newline: byte-identical to the committed artifact, so re-running
        # --write on an unchanged workbook produces a zero-line diff.
        OUT.write_text(json.dumps(data, indent=1, ensure_ascii=False), encoding="utf-8")
        NOTES_OUT.write_text(json.dumps(notes, indent=1, ensure_ascii=False), encoding="utf-8")
        print(f"wrote {OUT}  ({lines} lines, {cells} bucket cells)")
        print(f"wrote {NOTES_OUT}  ({note_lines} lines, {note_cells} source cells)")
        print(f"source_sha256 = {data['_meta']['source_sha256']}")
        return 0

    problems = _diff(data, json.loads(OUT.read_text(encoding="utf-8")))
    if NOTES_OUT.exists():
        committed_notes = json.loads(NOTES_OUT.read_text(encoding="utf-8"))
        if {k: v for k, v in notes.items() if k != "_meta"} != {
            k: v for k, v in committed_notes.items() if k != "_meta"
        }:
            problems.append("notes_source.json differs from the workbook")
    else:
        problems.append(f"{NOTES_OUT.name} missing — run with --write")

    if problems:
        print(f"EXTRACT DRIFT — {len(problems)} difference(s):", file=sys.stderr)
        for p in problems[:40]:
            print(f"  {p}", file=sys.stderr)
        if len(problems) > 40:
            print(f"  … {len(problems) - 40} more", file=sys.stderr)
        return 1

    print(
        f"OK — {lines} lines / {cells} bucket cells (Gross+RI) and "
        f"{note_lines} lines / {note_cells} source cells (notes) reproduce the committed extracts"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
