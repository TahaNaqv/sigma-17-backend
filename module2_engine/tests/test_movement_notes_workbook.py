"""Rendering of the note disclosure into the workbook and the JSON companion.

The load-bearing assertion here is the **preview contract**: `processing/output_preview.py`
opens output workbooks with ``data_only=True``, and openpyxl does not evaluate formulas —
so a formula-bearing cell previews as blank. Every note value must be written as a static
number. A test that only checked "the sheet exists" would pass on a workbook that renders
entirely empty in the app.
"""

import io
import types

import pandas as pd
import pytest
from openpyxl import load_workbook

from module2_engine.movement.compute import aggregated_views, build_sama_movement
from module2_engine.movement.notes import build_notes
from module2_engine.movement.workbook import (
    NOTE_TABS,
    build_json_companion,
    render_sama_workbook,
)


def _frames(rows=None):
    rows = rows or [
        {"RESERVINGCLASS": "MOTOR", "UWY": 2023,
         "Gross UPR_prev": 100.0, "Gross UPR_curr": 130.0, "GWP": 90.0,
         "Commission Expense": 7.0, "Premium Received": 40.0, "Claims Paid": 30.0,
         "GROSS - Insurance Finance (Income)/Expense": 5.0,
         "GROSS - Outstanding_prev": 200.0},
        {"RESERVINGCLASS": "PROPERTY", "UWY": 2022,
         "Gross UPR_prev": 50.0, "Gross UPR_curr": 55.0, "GWP": 20.0},
    ]
    ifrs = pd.DataFrame(rows)
    lc = pd.DataFrame([
        {"RESERVINGCLASS": r["RESERVINGCLASS"], "UWY": r["UWY"],
         "LC Discounted_PY": 0.0, "LC Discounted_CY": 0.0, "Loss Recovery Component": 0.0}
        for r in rows
    ])
    return types.SimpleNamespace(ifrs_summary_df=ifrs, allocate_sheets={"LC": lc})


def _workbook(**kw):
    res = build_sama_movement(_frames())
    return load_workbook(io.BytesIO(render_sama_workbook(res, **kw)), data_only=True)


def test_note_tabs_exist_in_the_clients_order():
    names = _workbook(reporting_date="31/12/2024").sheetnames
    assert names[:5] == ["Entity Total", *NOTE_TABS]


def test_note_cells_are_static_values_not_formulas():
    """The preview contract — data_only=True must return numbers, not None."""
    wb = _workbook(reporting_date="31/12/2024")
    for tab in NOTE_TABS:
        numeric = [
            c.value for row in wb[tab].iter_rows() for c in row
            if isinstance(c.value, (int, float))
        ]
        assert numeric, f"{tab} rendered no numeric cells — it would preview blank"
        for row in wb[tab].iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.startswith("=")), (
                    f"{tab}!{cell.coordinate} is a formula; the preview would show it empty"
                )


def test_rendered_values_match_the_computed_notes():
    """The sheet must not drift from build_notes — same numbers, same order."""
    res = build_sama_movement(_frames())
    entity = aggregated_views(res, levels=("entity",))[0]
    expected = build_notes(entity)["BS"]
    wb = _workbook(reporting_date="31/12/2024")
    rendered = {
        row[0]: row[1]
        for row in wb["BS"].iter_rows(values_only=True)
        if row and isinstance(row[0], str) and isinstance(row[1], (int, float))
    }
    for line in expected.lines:
        assert rendered[line.label] == pytest.approx(line.total, abs=0.01)


def test_reporting_date_replaces_the_val_date_placeholder():
    wb = _workbook(reporting_date="31/12/2024")
    text = " ".join(
        str(c.value) for row in wb["Gross_Note"].iter_rows() for c in row if c.value is not None
    )
    assert "As at 31/12/2024" in text
    assert "Val Date" not in text

    fallback = " ".join(
        str(c.value) for row in _workbook()["Gross_Note"].iter_rows() for c in row
        if c.value is not None
    )
    assert "As at Val Date" in fallback


def test_dash_cells_render_as_a_dash_not_a_zero():
    """RI has no contract-liability rows; the client shows '-' and so must we."""
    wb = _workbook(reporting_date="31/12/2024")
    row = next(
        r for r in wb["RI_Note"].iter_rows(values_only=True)
        if r and isinstance(r[0], str) and "liabilities – opening" in r[0]
    )
    assert set(row[1:6]) == {"-"}


def test_class_sheets_stack_the_two_note_tables_but_not_is_or_bs():
    """A per-class income statement would imply allocating GL items with no class
    dimension, so IS/BS stay entity-level (plan §11 Q5)."""
    wb = _workbook(reporting_date="31/12/2024")
    text = " ".join(
        str(c.value) for row in wb["MOTOR"].iter_rows() for c in row if c.value is not None
    )
    assert "12.2.1.1 Insurance contracts" in text  # Gross_Note title
    assert "12.2.2.1 Reinsurance contracts" in text  # RI_Note title
    assert "Income Statement" not in text
    assert "Balance Sheet" not in text


# ── JSON companion ───────────────────────────────────────────────────────────

def test_json_companion_carries_notes_and_the_deviation_ledger():
    comp = build_json_companion(build_sama_movement(_frames()), reporting_date="31/12/2024")
    assert comp["notes_schema_version"]
    assert comp["deviations"], "assumed deviations must be surfaced for the sign-off banner"
    for dev in comp["deviations"]:
        assert set(dev) >= {"id", "note", "client_cell", "resolution", "evidence", "status"}
        assert dev["status"] == "assumed"

    entity = next(v for v in comp["views"] if v["level"] == "entity")
    assert set(entity["notes"]) == set(NOTE_TABS)
    klass = next(v for v in comp["views"] if v["level"] == "class")
    assert set(klass["notes"]) == {"Gross_Note", "RI_Note"}


def test_json_note_values_match_the_workbook():
    res = build_sama_movement(_frames())
    comp = build_json_companion(res, levels=("entity",))
    entity = aggregated_views(res, levels=("entity",))[0]
    tables = build_notes(entity)
    for name, block in comp["views"][0]["notes"].items():
        for line in block["lines"]:
            expected = tables[name].line(line["id"]).values
            for col, value in line["values"].items():
                if value is None:
                    assert expected[col] is None
                else:
                    assert value == pytest.approx(expected[col], abs=0.01)


# ── degenerate inputs (plan §6.4) ────────────────────────────────────────────

def test_levels_without_entity_emit_no_note_tabs():
    res = build_sama_movement(_frames())
    wb = load_workbook(io.BytesIO(render_sama_workbook(res, levels=("class",))), data_only=True)
    assert not set(wb.sheetnames) & set(NOTE_TABS)


def test_empty_result_still_renders():
    empty = build_sama_movement(
        types.SimpleNamespace(
            ifrs_summary_df=pd.DataFrame(columns=["RESERVINGCLASS", "UWY"]),
            allocate_sheets={},
        )
    )
    wb = load_workbook(io.BytesIO(render_sama_workbook(empty)), data_only=True)
    assert wb.sheetnames == ["Movement Analysis"]
