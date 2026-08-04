"""Render the IFRS 17 movement disclosure to an .xlsx workbook (openpyxl) plus a
machine-readable JSON companion.

Faithful to the client's SAMA template: the exact Gross/RI line structure with the four
measurement buckets + Total, the direction sign column, and the real nested subtotals
(resolved by ``compute.row_values``; rows the client flattened to statics come from
``schema.RECONSTRUCTED_FORMULAS``, which carries the verification of each against the
client's own Total column).

Grains (plan Q1): an entity Total worksheet, then one worksheet per reserving class
carrying the class total followed by its per-cohort (UWY) detail — all by summation, so
every aggregate ties to the sum of its parts. The closing line is rendered as the
roll-forward — opening + ΔP&L ± cash flows, the sign taken per sheet from
``schema.CLOSING_CASHFLOW_SIGN`` (Gross adds, RI subtracts: a liability and an asset
sign their cash flows in mirror image).

Uses openpyxl (a guaranteed dependency) — NOT xlsxwriter (optional, absent in prod) and
NOT the pandas WRITE_ENGINE (whose engine choice must stay free for the bit-identical
process pipeline).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .compute import MovementResult, SheetResult, aggregated_views, row_values
from .notes import NoteTable, build_notes
from .notes_schema import NOTES_SCHEMA_VERSION, STATUS_ASSUMED, deviations
from .schema import SCHEMA, SCHEMA_VERSION, Sheet

# Subtotal resolution lives in compute (the note layer needs the same numbers, and the
# two must agree by construction). Kept under the old private name for call sites here.
_row_values = row_values

_NUMFMT = "#,##0;(#,##0)"
#: The client's accounting format, used on the note tables so they read like their template.
_NOTE_NUMFMT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'

#: Notes rendered as their own tabs, in the client's order, at the entity grain only.
#: ``IS``/``BS`` are entity-level financial statements (plan §11 Q5); the two note tables
#: are additive, so they are also stacked under each class sheet.
NOTE_TABS: tuple[str, ...] = ("Gross_Note", "RI_Note", "IS", "BS")
STACKED_NOTES: tuple[str, ...] = ("Gross_Note", "RI_Note")

#: Merged column-group headers, mirroring the client's two-level header block. The RI
#: wording is taken from the client's own ``RI`` movement sheet rather than from
#: ``RI_Note``, where the Gross captions were copy-pasted.
_NOTE_COLUMN_GROUPS: dict[str, tuple[tuple[str, int], ...]] = {
    "Gross_Note": (("Liability for remaining coverage", 2), ("Liability for incurred claims", 2)),
    "RI_Note": (("Assets for remaining coverage", 2),
                ("Amounts recoverable on incurred claims", 2)),
}
_THIN = Side(style="thin", color="DDDDDD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")

_F = {
    "title": Font(bold=True, size=13),
    "meta": Font(italic=True, color="888888"),
    "sheet_tag": Font(bold=True, color="1F4E78"),
    "bucket_font": Font(bold=True, color="FFFFFF"),
    "bucket_fill": PatternFill("solid", fgColor="1F4E78"),
    "bold": Font(bold=True),
    "sign": Font(color="888888"),
    "open_fill": PatternFill("solid", fgColor="FCE4D6"),
    "close_fill": PatternFill("solid", fgColor="E2EFDA"),
    "sub_fill": PatternFill("solid", fgColor="F2F2F2"),
}


def _num(v) -> float:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return 0.0
    return f if f == f and f not in (float("inf"), float("-inf")) else 0.0


# ── workbook rendering ───────────────────────────────────────────────────────

def _safe_sheet_name(name: str, used: set[str]) -> str:
    clean = "".join(c for c in name if c not in r"[]:*?/\\")[:31].strip() or "Sheet"
    base, candidate, i = clean[:28], clean, 2
    while candidate in used:
        candidate = f"{base}~{i}"
        i += 1
    used.add(candidate)
    return candidate


def _closing_label(label: str, reporting_date: str | None) -> str:
    if reporting_date and "as at" in label.lower():
        return label.rsplit("as at", 1)[0] + f"as at {reporting_date}"
    return label


def render_sama_workbook(
    result: MovementResult, *, reporting_date: str | None = None,
    levels: tuple[str, ...] = ("entity", "class", "cohort"),
) -> bytes:
    """Return .xlsx bytes: an Entity Total sheet, then one sheet per class (class total
    followed by its cohort detail)."""
    wb = Workbook()
    wb.remove(wb.active)
    views = aggregated_views(result, levels=levels)
    used: set[str] = set()

    # Entity Total, then the four note tabs in the client's order — the notes come before
    # the per-class detail, matching how their workbook is laid out.
    for v in [x for x in views if x["level"] == "entity"]:
        _render_view(wb.create_sheet(_safe_sheet_name("Entity Total", used)), v, reporting_date)
        for name, table in _notes_for_view(v).items():
            _render_note(wb.create_sheet(_safe_sheet_name(name, used)), 1, table, reporting_date)

    cohorts_by_class: dict[str, list] = {}
    for v in [x for x in views if x["level"] == "cohort"]:
        cohorts_by_class.setdefault(v["reserving_class"], []).append(v)
    class_views = [x for x in views if x["level"] == "class"]

    if class_views:
        for cv in class_views:
            ws = wb.create_sheet(_safe_sheet_name(cv["reserving_class"], used))
            r = _render_view(ws, cv, reporting_date)
            for coh in sorted(cohorts_by_class.get(cv["reserving_class"], []), key=lambda x: x["uwy"]):
                r = _render_view(ws, coh, reporting_date, start=r + 1)
    else:  # cohort-only: group cohorts onto one sheet per class
        for rc in sorted(cohorts_by_class):
            ws = wb.create_sheet(_safe_sheet_name(rc, used))
            r = 1
            for coh in sorted(cohorts_by_class[rc], key=lambda x: x["uwy"]):
                r = _render_view(ws, coh, reporting_date, start=r)

    if not wb.worksheets:  # nothing produced (empty result)
        wb.create_sheet("Movement Analysis")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_view(ws, view: dict, reporting_date: str | None, start: int = 1) -> int:
    ws.column_dimensions["A"].width = 54
    ws.column_dimensions["B"].width = 6
    for col in ("C", "D", "E", "F", "G"):
        ws.column_dimensions[col].width = 18
    r = start
    c = ws.cell(r, 1, f"IFRS 17 Movement Analysis — {view['label']}")
    c.font = _F["title"]
    r += 1
    c = ws.cell(r, 1, f"schema v{SCHEMA_VERSION} · authoritative mapping (client-signed)")
    c.font = _F["meta"]
    r += 2
    for sname in ("Gross", "RI"):
        sres = view["sheets"].get(sname)
        if sres is None:
            continue
        c = ws.cell(r, 1, sname)
        c.font = _F["sheet_tag"]
        r += 1
        r = _render_table(ws, r, SCHEMA.sheets[sname], sres, reporting_date) + 2
    # The note tables restate this view's movement — stacked here for class/cohort grains
    # (the entity grain gets them as dedicated tabs instead).
    if view.get("level") != "entity":
        for table in _notes_for_view(view).values():
            r = _render_note(ws, r, table, reporting_date) + 2
    return r


def _render_table(ws, r0: int, sheet: Sheet, sres: SheetResult, reporting_date) -> int:
    vbuckets = list(sheet.value_buckets)
    resolved = row_values(sheet, sres)
    r = r0

    # bucket header row
    for j, label in enumerate(["", ""] + [b.replace("_", " ") for b in vbuckets] + ["Total"]):
        cell = ws.cell(r, 1 + j, label)
        cell.fill = _F["bucket_fill"]
        cell.border = _BORDER
        if j >= 2:
            cell.font = _F["bucket_font"]
            cell.alignment = _CENTER
    r += 1

    for ln in sheet.lines:
        raw = _closing_label(ln.label, reporting_date) if ln.kind == "closing" else ln.label
        label = "    " * ln.level + raw
        if ln.kind == "section":
            cell = ws.cell(r, 1, label)
            cell.font = _F["bold"]
            cell.fill = _F["sub_fill"]
            cell.border = _BORDER
            r += 1
            continue
        vals = resolved.get(ln.row, {})
        sign = next((s for b, s in ln.signs.items() if b != "Total"), "") if ln.kind == "input" else ""
        fill = {"opening": _F["open_fill"], "closing": _F["close_fill"]}.get(ln.kind)
        if ln.kind in ("opening", "subtotal", "closing"):
            fill = fill or _F["sub_fill"]
        bold = ln.kind in ("opening", "subtotal", "closing")
        _value_row(ws, r, label, vals, vbuckets, sign=sign, fill=fill, bold=bold)
        r += 1
    return r


def _value_row(ws, r, label, vals, vbuckets, *, sign="", fill=None, bold=False):
    lc = ws.cell(r, 1, label)
    lc.border = _BORDER
    if bold:
        lc.font = _F["bold"]
    if fill:
        lc.fill = fill
    sc = ws.cell(r, 2, sign or "")
    sc.border = _BORDER
    sc.alignment = _CENTER
    sc.font = _F["sign"]
    if fill:
        sc.fill = fill
    total = 0.0
    for j, b in enumerate(vbuckets):
        v = _num(vals.get(b, 0.0))
        total += v
        cell = ws.cell(r, 3 + j, v)
        cell.number_format = _NUMFMT
        cell.border = _BORDER
        if bold:
            cell.font = _F["bold"]
        if fill:
            cell.fill = fill
    tcell = ws.cell(r, 3 + len(vbuckets), total)
    tcell.number_format = _NUMFMT
    tcell.border = _BORDER
    if bold:
        tcell.font = _F["bold"]
    if fill:
        tcell.fill = fill


# ── note disclosure rendering (Gross_Note / RI_Note / IS / BS) ───────────────

def _render_note(ws, r0: int, table: NoteTable, reporting_date: str | None) -> int:
    """Render one note table as **static values**. Never formulas: the output preview
    loads with ``data_only=True`` and openpyxl does not evaluate, so a formula-bearing
    cell would preview blank."""
    ws.column_dimensions["A"].width = 62
    columns = list(table.columns)
    for i in range(len(columns)):
        ws.column_dimensions[chr(ord("B") + i)].width = 20

    r = r0
    ws.cell(r, 1, table.title).font = _F["title"]
    r += 1
    stamp = ws.cell(
        r, 1,
        f"schema v{SCHEMA_VERSION} · notes v{NOTES_SCHEMA_VERSION}"
        + (f" · {len(deviations(STATUS_ASSUMED))} presentation assumptions pending client confirmation"
           if deviations(STATUS_ASSUMED) else ""),
    )
    stamp.font = _F["meta"]
    r += 1

    span = len(columns)
    header = ws.cell(r, 2, f"As at {reporting_date}" if reporting_date else "As at Val Date")
    header.font = _F["bold"]
    header.alignment = _CENTER
    if span > 1:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=1 + span)
    r += 1

    groups = _NOTE_COLUMN_GROUPS.get(table.name)
    if groups:
        col = 2
        for title, width in groups:
            cell = ws.cell(r, col, title)
            cell.font = _F["bucket_font"]
            cell.fill = _F["bucket_fill"]
            cell.alignment = _CENTER
            cell.border = _BORDER
            if width > 1:
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + width - 1)
            for offset in range(width):
                ws.cell(r, col + offset).fill = _F["bucket_fill"]
                ws.cell(r, col + offset).border = _BORDER
            col += width
        total = ws.cell(r, col, "Total")
        total.font = _F["bucket_font"]
        total.fill = _F["bucket_fill"]
        total.alignment = _CENTER
        total.border = _BORDER
        r += 1

    for i, column in enumerate(columns):
        cell = ws.cell(r, 2 + i, column.replace("_", " "))
        cell.font = _F["bucket_font"]
        cell.fill = _F["bucket_fill"]
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.cell(r, 1).fill = _F["bucket_fill"]
    ws.cell(r, 1).border = _BORDER
    r += 1

    for line in table.lines:
        bold = line.kind in ("subtotal", "section")
        label = ws.cell(r, 1, line.label)
        label.border = _BORDER
        if bold:
            label.font = _F["bold"]
        if line.kind == "section":
            label.fill = _F["sub_fill"]
            for i in range(len(columns)):
                ws.cell(r, 2 + i).fill = _F["sub_fill"]
                ws.cell(r, 2 + i).border = _BORDER
            r += 1
            continue
        if line.kind == "subtotal":
            label.fill = _F["sub_fill"]
        for i, column in enumerate(columns):
            value = line.values.get(column)
            cell = ws.cell(r, 2 + i, "-" if value is None else _num(value))
            cell.number_format = _NOTE_NUMFMT
            cell.border = _BORDER
            if value is None:
                cell.alignment = _CENTER
            if bold:
                cell.font = _F["bold"]
            if line.kind == "subtotal":
                cell.fill = _F["sub_fill"]
        r += 1
    return r


def _notes_for_view(view: dict) -> dict[str, NoteTable]:
    """The note tables a view carries. ``IS``/``BS`` are entity-level statements — a
    per-class income statement would imply allocating general-ledger items that have no
    class dimension (plan §11 Q5)."""
    tables = build_notes(view)
    if view.get("level") == "entity":
        return {name: tables[name] for name in NOTE_TABS if name in tables}
    return {name: tables[name] for name in STACKED_NOTES if name in tables}


# ── machine-readable JSON companion (plan Q4: structured feed for preview/API) ──

def build_json_companion(
    result: MovementResult, *, reporting_date: str | None = None,
    levels: tuple[str, ...] = ("entity", "class", "cohort"),
) -> dict:
    """A structured mirror of the workbook: per grain-view, per sheet, the ordered lines
    with label/level/kind/sign and per-bucket values (subtotals resolved). Downstream/API
    consumers use this instead of cracking the xlsx."""
    out_views = []
    for v in aggregated_views(result, levels=levels):
        sheets = {}
        for sname, sres in v["sheets"].items():
            sh = SCHEMA.sheets[sname]
            row_values = _row_values(sh, sres)
            lines = []
            for ln in sh.lines:
                if ln.kind == "section":
                    lines.append({"id": ln.id, "label": ln.label, "level": ln.level, "kind": "section"})
                    continue
                vals = {b: round(row_values.get(ln.row, {}).get(b, 0.0), 2) for b in sh.value_buckets}
                lines.append({
                    "id": ln.id, "label": ln.label, "level": ln.level, "kind": ln.kind,
                    "sign": next((s for b, s in ln.signs.items() if b != "Total"), None) if ln.kind == "input" else None,
                    "buckets": vals, "total": round(sum(vals.values()), 2),
                })
            sheets[sname] = {"buckets": list(sh.value_buckets), "lines": lines}
        notes = {
            name: {
                "title": table.title,
                "columns": list(table.columns),
                "lines": [
                    {
                        "id": ln.id, "row": ln.row, "label": ln.label, "kind": ln.kind,
                        "values": {
                            col: (None if val is None else round(val, 2))
                            for col, val in ln.values.items()
                        },
                    }
                    for ln in table.lines
                ],
            }
            for name, table in _notes_for_view(v).items()
        }
        out_views.append({
            "level": v["level"], "label": v["label"],
            "reserving_class": v["reserving_class"], "uwy": v["uwy"],
            "sheets": sheets, "notes": notes,
        })
    return {
        "schema_version": SCHEMA_VERSION,
        "notes_schema_version": NOTES_SCHEMA_VERSION,
        "reporting_date": reporting_date,
        # Presentation assumptions still awaiting the client's confirmation (plan §8). The
        # UI surfaces the count on the sign-off banner; nothing diverges silently.
        "deviations": [
            {"id": d.id, "note": d.note, "row": d.row, "client_cell": d.client_cell,
             "resolution": d.resolution, "evidence": d.evidence, "status": d.status}
            for d in deviations(STATUS_ASSUMED)
        ],
        "views": out_views,
    }
