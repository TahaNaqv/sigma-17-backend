"""Render a SensitivityResult to ``Sensitivity_Analysis.xlsx``.

openpyxl rather than ``WRITE_ENGINE``: this is a presentation artefact that needs
per-cell fills, number formats and multi-table sheets, and it is small enough that
the write-speed argument for XlsxWriter does not apply.

Design notes that are requirements, not preferences:

* **Absolute and percent are both rendered, always.** A threshold residual like the
  Loss Component can move +145% on a base three orders of magnitude below GMM LRC;
  a percent-only view would rank it as the book's dominant risk, which is false.
* **The tornado ranks by absolute**, for the same reason.
* **Structural zeros are shown as "-", not 0.00%**, so a measure that genuinely does
  not respond to a lever is visibly distinct from one that responds by ~nothing.
* **Scenario Definitions echoes ``base -> shocked`` in absolute terms**, so the reader
  never has to infer whether "+10%" meant relative or percentage points.
"""

from __future__ import annotations

import io
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from module2_engine.scenarios import (
    ALL_MEASURES,
    LEVER_DISCOUNT,
    LEVER_RA,
    LEVER_ULR,
    MEASURES_BY_KEY,
    RATIO,
    TOTAL,
    SensitivityResult,
)

HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill("solid", fgColor="008080")
TITLE_FONT = Font(bold=True, size=12)
MUTED_FONT = Font(color="808080", italic=True)

POS_FILL = PatternFill("solid", fgColor="C6EFCE")   # green
NEG_FILL = PatternFill("solid", fgColor="FFC7CE")   # red
NEUTRAL_FILL = PatternFill("solid", fgColor="F2F2F2")

MONEY_FMT = '_-* #,##0_-;[Red]-* #,##0_-;_-* "-"_-;_-@_-'
RATIO_FMT = '0.00%'
PCT_DELTA_FMT = '+0.000%;-0.000%;"-"'
THIN = Side(style="thin", color="D9D9D9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def _autosize(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _unit_label(lever: str) -> str:
    return {LEVER_RA: "relative %", LEVER_DISCOUNT: "basis points",
            LEVER_ULR: "percentage points"}.get(lever, "")


def _write_definitions(wb: Workbook, result: SensitivityResult) -> None:
    ws = wb.create_sheet("Scenario Definitions")
    ws["A1"] = "Sensitivity scenario definitions"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = (
        "Units differ per lever and are NOT interchangeable. Risk adjustment is shocked "
        "RELATIVELY (a +10% shock scales an existing 4.63% loading to 5.09%). Discounting "
        "is shocked in ABSOLUTE basis points on the CY annual spot curve; the PY curve is "
        "deliberately left untouched because it is the prior period's locked-in basis. "
        "Loss ratio is shocked in ABSOLUTE percentage points on Selected ULR."
    )
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=3, start_column=1, end_row=5, end_column=7)

    hdr = ["Scenario", "Lever", "Magnitude", "Unit", "Applies to", "Base", "Shocked"]
    r0 = 7
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(hdr))

    for i, echo in enumerate(result.resolved, start=1):
        r = r0 + i
        ws.cell(row=r, column=1, value=echo["label"])
        ws.cell(row=r, column=2, value=echo["lever"])
        mag = echo["magnitude"]
        ws.cell(row=r, column=3, value=mag * 100 if echo["lever"] != LEVER_DISCOUNT else mag)
        ws.cell(row=r, column=4, value=_unit_label(echo["lever"]))
        scope = echo.get("scope_classes") or []
        ws.cell(row=r, column=5, value=", ".join(scope) if scope else "All classes")
        if "base_min" in echo:
            ws.cell(row=r, column=6,
                    value=f"{echo['base_min']:.4%} - {echo['base_max']:.4%}")
            ws.cell(row=r, column=7,
                    value=f"{echo['shocked_min']:.4%} - {echo['shocked_max']:.4%}")
        elif echo["lever"] == LEVER_DISCOUNT:
            ws.cell(row=r, column=6, value="CY annual spot curve")
            ws.cell(row=r, column=7, value=f"parallel {echo['shift']:+.4%}; PY untouched")
        for c in range(1, len(hdr) + 1):
            ws.cell(row=r, column=c).border = BOX

    if result.warnings:
        wr = r0 + len(result.resolved) + 2
        ws.cell(row=wr, column=1, value="Warnings").font = TITLE_FONT
        for j, w in enumerate(result.warnings, start=1):
            ws.cell(row=wr + j, column=1, value=w).font = MUTED_FONT

    _autosize(ws, {1: 22, 2: 12, 3: 12, 4: 20, 5: 30, 6: 26, 7: 32})


def _measure_rows(result: SensitivityResult) -> list[Any]:
    return [m for m in ALL_MEASURES if m.key in result.base]


def _write_levels(wb: Workbook, result: SensitivityResult, reserving_class: str) -> None:
    """Base + every scenario as absolute levels, one column per scenario."""
    ws = wb.create_sheet("Levels")
    ws["A1"] = "Measure levels — base and each scenario"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Reserving class: {'All classes (total)' if reserving_class == TOTAL else reserving_class}"
    ws["A2"].font = MUTED_FONT

    r0 = 4
    ws.cell(row=r0, column=1, value="Measure")
    ws.cell(row=r0, column=2, value="Base")
    for j, (shock, _) in enumerate(result.per_scenario, start=3):
        ws.cell(row=r0, column=j, value=shock.label)
    _style_header(ws, r0, 2 + len(result.per_scenario))

    for i, m in enumerate(_measure_rows(result), start=1):
        r = r0 + i
        ws.cell(row=r, column=1, value=m.label)
        fmt = RATIO_FMT if m.kind == RATIO else MONEY_FMT
        c = ws.cell(row=r, column=2, value=result.base[m.key].get(reserving_class, 0.0))
        c.number_format = fmt
        c.border = BOX
        for j, (_, vals) in enumerate(result.per_scenario, start=3):
            c = ws.cell(row=r, column=j, value=vals.get(m.key, {}).get(reserving_class, 0.0))
            c.number_format = fmt
            c.border = BOX
        ws.cell(row=r, column=1).border = BOX
    ws.freeze_panes = ws.cell(row=r0 + 1, column=2)
    _autosize(ws, {1: 34, **{c: 18 for c in range(2, 3 + len(result.per_scenario))}})


def _write_comparison(
    wb: Workbook, result: SensitivityResult, reserving_class: str, *, percent: bool
) -> None:
    title = "Comparison — Percent" if percent else "Comparison — Absolute"
    ws = wb.create_sheet(title)
    ws["A1"] = f"{title} (vs base)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Percent deltas exaggerate movements on small bases — read alongside the "
        "absolute sheet." if percent else
        "Absolute deltas. This is the ranking basis used by the Tornado sheet."
    )
    ws["A2"].font = MUTED_FONT
    ws["A3"] = f"Reserving class: {'All classes (total)' if reserving_class == TOTAL else reserving_class}"
    ws["A3"].font = MUTED_FONT

    r0 = 5
    ws.cell(row=r0, column=1, value="Measure")
    for j, (shock, _) in enumerate(result.per_scenario, start=2):
        ws.cell(row=r0, column=j, value=shock.label)
    _style_header(ws, r0, 1 + len(result.per_scenario))

    for i, m in enumerate(_measure_rows(result), start=1):
        r = r0 + i
        ws.cell(row=r, column=1, value=m.label).border = BOX
        base = result.base[m.key].get(reserving_class, 0.0)
        for j, (_, vals) in enumerate(result.per_scenario, start=2):
            v = vals.get(m.key, {}).get(reserving_class, 0.0)
            delta = v - base
            cell = ws.cell(row=r, column=j)
            if abs(delta) <= 1e-9:
                # Structural zero: this lever does not reach this measure.
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")
                cell.font = MUTED_FONT
                cell.fill = NEUTRAL_FILL
            elif percent:
                if abs(base) <= 1e-9:
                    cell.value = "n/a"
                    cell.font = MUTED_FONT
                else:
                    cell.value = delta / abs(base)
                    cell.number_format = PCT_DELTA_FMT
                    cell.fill = POS_FILL if delta > 0 else NEG_FILL
            else:
                cell.value = delta
                cell.number_format = RATIO_FMT if m.kind == RATIO else MONEY_FMT
                cell.fill = POS_FILL if delta > 0 else NEG_FILL
            cell.border = BOX
    ws.freeze_panes = ws.cell(row=r0 + 1, column=2)
    _autosize(ws, {1: 34, **{c: 16 for c in range(2, 2 + len(result.per_scenario))}})


def _write_tornado(wb: Workbook, result: SensitivityResult, reserving_class: str) -> None:
    ws = wb.create_sheet("Tornado")
    ws["A1"] = "Measures ranked by peak absolute sensitivity"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (
        "Ranked on ABSOLUTE movement. Ranking on percent would promote threshold "
        "residuals (e.g. Loss Component) above measures that dominate the balance sheet."
    )
    ws["A2"].font = MUTED_FONT

    t = result.tornado(reserving_class)
    r0 = 4
    hdr = ["Measure", "Peak absolute move", "Most negative", "Most positive"]
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=r0, column=i, value=h)
    _style_header(ws, r0, len(hdr))
    for i, row in enumerate(t.itertuples(index=False), start=1):
        m = MEASURES_BY_KEY.get(row.measure_key)
        fmt = RATIO_FMT if (m and m.kind == RATIO) else MONEY_FMT
        r = r0 + i
        ws.cell(row=r, column=1, value=row.measure).border = BOX
        for col, val in ((2, row.max_abs_delta), (3, row.min_delta), (4, row.max_delta)):
            c = ws.cell(row=r, column=col, value=float(val))
            c.number_format = fmt
            c.border = BOX
    _autosize(ws, {1: 34, 2: 22, 3: 20, 4: 20})


def _write_by_class(wb: Workbook, result: SensitivityResult) -> None:
    """One row per (class, measure, scenario) — the drill-down behind the totals."""
    ws = wb.create_sheet("By Class")
    hdr = ["Reserving class", "Measure", "Scenario", "Lever", "Base", "Shocked",
           "Absolute delta", "Percent delta"]
    for i, h in enumerate(hdr, start=1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(hdr))
    r = 1
    for rc in result.reserving_classes:
        frame = result.comparison(rc)
        for row in frame.itertuples(index=False):
            r += 1
            m = MEASURES_BY_KEY.get(row.measure_key)
            fmt = RATIO_FMT if (m and m.kind == RATIO) else MONEY_FMT
            ws.cell(row=r, column=1, value=rc)
            ws.cell(row=r, column=2, value=row.measure)
            ws.cell(row=r, column=3, value=row.scenario)
            ws.cell(row=r, column=4, value=row.lever)
            for col, val in ((5, row.base), (6, row.value), (7, row.abs_delta)):
                c = ws.cell(row=r, column=col, value=float(val))
                c.number_format = fmt
            pc = ws.cell(row=r, column=8)
            if row.pct_delta is None or (isinstance(row.pct_delta, float) and np.isnan(row.pct_delta)):
                pc.value = None
            else:
                pc.value = float(row.pct_delta)
                pc.number_format = PCT_DELTA_FMT
    ws.freeze_panes = "A2"
    _autosize(ws, {1: 30, 2: 32, 3: 18, 4: 12, 5: 18, 6: 18, 7: 18, 8: 14})


def render_sensitivity_workbook(
    result: SensitivityResult, *, reserving_class: str = TOTAL
) -> bytes:
    """Serialize a SensitivityResult to xlsx bytes."""
    wb = Workbook()
    wb.remove(wb.active)
    _write_definitions(wb, result)
    _write_levels(wb, result, reserving_class)
    _write_comparison(wb, result, reserving_class, percent=False)
    _write_comparison(wb, result, reserving_class, percent=True)
    _write_tornado(wb, result, reserving_class)
    _write_by_class(wb, result)
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()
