from __future__ import annotations

import io
import json
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from typing import BinaryIO
from zipfile import ZipFile

from openpyxl import load_workbook

from .output_column_kinds import classify_columns, classify_rows


@dataclass(frozen=True)
class OutputFileItem:
    path: str
    size: int
    kind: str


@dataclass(frozen=True)
class SheetMetaItem:
    name: str
    rows: int
    columns: int


def _normalize_zip_path(path: str) -> str:
    p = (path or "").replace("\\", "/").strip()
    if not p or p.startswith("/") or ".." in p.split("/"):
        raise ValueError("Invalid file path.")
    return p


def _normalize_sheet_name(sheet_name: str) -> str:
    name = (sheet_name or "").strip()
    if not name:
        raise ValueError("Sheet name is required.")
    return name


def _as_int(value: str | None, *, field: str, default: int) -> int:
    if value is None or str(value).strip() == "":
        return default
    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ValueError(f"{field} must be an integer.") from exc
    if parsed < 1:
        raise ValueError(f"{field} must be >= 1.")
    return parsed


def parse_page_args(
    page_raw: str | None,
    page_size_raw: str | None,
    *,
    default_page_size: int,
    max_page_size: int,
) -> tuple[int, int]:
    page = _as_int(page_raw, field="page", default=1)
    page_size = _as_int(page_size_raw, field="page_size", default=default_page_size)
    if page_size > max_page_size:
        raise ValueError(f"page_size exceeds max {max_page_size}.")
    return page, page_size


def _infer_kind(path: str) -> str:
    if path.lower().endswith(".xlsx"):
        return "xlsx"
    return "other"


def _to_json_value(v):
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, (date, time)):
        return v.isoformat()
    return str(v)


def _open_zip(zip_source: bytes | BinaryIO) -> ZipFile:
    if isinstance(zip_source, bytes):
        return ZipFile(io.BytesIO(zip_source), "r")
    return ZipFile(zip_source, "r")


def list_preview_files(zip_source: bytes | BinaryIO) -> list[OutputFileItem]:
    with _open_zip(zip_source) as zf:
        out: list[OutputFileItem] = []
        for info in zf.infolist():
            if info.is_dir():
                continue
            path = _normalize_zip_path(info.filename)
            kind = _infer_kind(path)
            if kind == "xlsx":
                out.append(OutputFileItem(path=path, size=info.file_size, kind=kind))
    out.sort(key=lambda x: x.path.lower())
    return out


#: The movement job's machine-readable companion, written beside the workbook.
MOVEMENT_NOTES_JSON = "IFRS17_Movement_Analysis.json"

#: Bound on the companion we will parse. It is ~1 MB for the full 96-view grid; anything
#: far beyond that is not a companion we wrote, and json.loads on an unbounded blob from
#: storage is not something to do casually.
_MAX_COMPANION_BYTES = 32 * 1024 * 1024


def read_json_companion(zip_source: bytes | BinaryIO, file_path: str = MOVEMENT_NOTES_JSON) -> dict:
    """Parse a JSON companion out of an output ZIP.

    The three preview endpoints read workbooks; this reads the structured feed the movement
    job writes alongside one, so callers get the disclosure as data instead of re-deriving
    it from rendered cells.
    """
    target = _normalize_zip_path(file_path)
    if not target.lower().endswith(".json"):
        raise ValueError("Requested file is not a JSON companion.")
    with _open_zip(zip_source) as zf:
        try:
            info = zf.getinfo(target)
        except KeyError as exc:
            raise ValueError("Requested file was not found in output ZIP.") from exc
        if info.file_size > _MAX_COMPANION_BYTES:
            raise ValueError("JSON companion is too large to parse.")
        raw = zf.read(target)
    try:
        payload = json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("JSON companion is not readable.") from exc
    if not isinstance(payload, dict):
        raise ValueError("JSON companion has an unexpected shape.")
    return payload


def _read_workbook_from_zip(zip_source: bytes | BinaryIO, file_path: str):
    target = _normalize_zip_path(file_path)
    with _open_zip(zip_source) as zf:
        try:
            raw = zf.read(target)
        except KeyError as exc:
            raise ValueError("Requested file was not found in output ZIP.") from exc
    return load_workbook(io.BytesIO(raw), read_only=True, data_only=True)


def list_workbook_sheets(zip_source: bytes | BinaryIO, file_path: str) -> list[SheetMetaItem]:
    wb = _read_workbook_from_zip(zip_source, file_path)
    try:
        out: list[SheetMetaItem] = []
        for ws in wb.worksheets:
            rows = int(ws.max_row or 0)
            cols = int(ws.max_column or 0)
            out.append(SheetMetaItem(name=ws.title, rows=rows, columns=cols))
        return out
    finally:
        wb.close()


def read_sheet_page(
    zip_source: bytes | BinaryIO,
    *,
    file_path: str,
    sheet_name: str,
    page: int,
    page_size: int,
    max_cells: int,
) -> dict:
    if page < 1:
        raise ValueError("page must be >= 1.")
    if page_size < 1:
        raise ValueError("page_size must be >= 1.")
    wb = _read_workbook_from_zip(zip_source, file_path)
    try:
        sheet_key = _normalize_sheet_name(sheet_name)
        if sheet_key not in wb.sheetnames:
            raise ValueError("Requested sheet was not found in workbook.")
        ws = wb[sheet_key]
        max_row = int(ws.max_row or 0)
        max_col = int(ws.max_column or 0)
        if max_col > max_cells:
            raise ValueError("Sheet is too wide to preview within cell limit.")
        if max_row == 0 or max_col == 0:
            return {
                "file": _normalize_zip_path(file_path),
                "sheet": sheet_key,
                "page": page,
                "page_size": page_size,
                "total_rows": 0,
                "columns": [],
                "column_kinds": [],
                "row_kinds": None,
                "rows": [],
            }

        row_iter = ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True)
        header_row = next(row_iter, None)
        if header_row is None:
            return {
                "file": _normalize_zip_path(file_path),
                "sheet": sheet_key,
                "page": page,
                "page_size": page_size,
                "total_rows": 0,
                "columns": [],
                "column_kinds": [],
                "row_kinds": None,
                "rows": [],
            }

        columns = [str(v).strip() if v is not None and str(v).strip() else f"Column {i + 1}" for i, v in enumerate(header_row)]
        # Per-column display hint (ratio -> %, factor -> 3dp, number -> default)
        # so the generic preview grid can format actuarial ratios correctly.
        column_kinds = classify_columns(sheet_key, columns)
        total_rows = max(0, max_row - 1)

        start_data_row_idx = (page - 1) * page_size + 1  # 1-indexed over data rows
        end_data_row_idx = min(total_rows, start_data_row_idx + page_size - 1)
        if start_data_row_idx > total_rows:
            return {
                "file": _normalize_zip_path(file_path),
                "sheet": sheet_key,
                "page": page,
                "page_size": page_size,
                "total_rows": total_rows,
                "columns": columns,
                "column_kinds": column_kinds,
                "row_kinds": None,
                "rows": [],
            }

        rows: list[list] = []
        cell_count = 0
        data_idx = 0
        for row in row_iter:
            data_idx += 1
            if data_idx < start_data_row_idx:
                continue
            if data_idx > end_data_row_idx:
                break
            normalized = [_to_json_value(v) for v in row]
            rows.append(normalized)
            cell_count += len(normalized)
            if cell_count > max_cells:
                raise ValueError("Requested page exceeds output preview cell limit.")

        # Per-ROW display hint, for sheets whose kind varies down the sheet rather than
        # across it. A triangle holds cumulative money, age-to-age factors and a factor count
        # in the SAME columns, so column kinds alone render a 1.015748 factor as "1.01".
        # `None` for every other sheet, which keeps column classification authoritative there.
        #
        # Classified over the WHOLE sheet, then sliced to the page: a row's kind is inherited
        # from the block label above it, so classifying a page in isolation would restart that
        # inheritance at the page boundary and mislabel every row on page 2.
        row_kinds = None
        if classify_rows(sheet_key, []) is not None:
            labels = [
                cell[0]
                for cell in ws.iter_rows(
                    min_row=2, max_row=max_row, max_col=1, values_only=True
                )
            ]
            all_kinds = classify_rows(sheet_key, labels) or []
            row_kinds = all_kinds[start_data_row_idx - 1 : end_data_row_idx]

        return {
            "file": _normalize_zip_path(file_path),
            "sheet": sheet_key,
            "page": page,
            "page_size": page_size,
            "total_rows": total_rows,
            "columns": columns,
            "column_kinds": column_kinds,
            "row_kinds": row_kinds,
            "rows": rows,
        }
    finally:
        wb.close()
