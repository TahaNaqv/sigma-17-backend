import io
import shutil
import tempfile
import zipfile
from unittest.mock import patch

from django.contrib.auth.models import User
from django.core.files.base import ContentFile
from django.test import TestCase, override_settings
from openpyxl import Workbook
from rest_framework.test import APIClient

from accounts.models import Permission, Role
from processing.models import Module1Job
from tenants.models import Membership, Organization

TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="sigma17-test-media-")


def _xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Summary"
    ws.append(["RESERVINGCLASS", "UWY", "Amount"])
    ws.append(["Motor GROSS", "2023", 123.45])
    ws.append(["Property RI", "2024", 678.9])
    ws2 = wb.create_sheet("UW Summary")
    ws2.append(["RESERVINGCLASS", "UWY", "Exp Ratio", "RI %"])
    ws2.append(["Motor GROSS", "2023", 0.08, 0.25])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _zip_with_xlsx(xlsx_name="Combined_Summary.xlsx") -> bytes:
    xlsx_raw = _xlsx_bytes()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(xlsx_name, xlsx_raw)
        zf.writestr("notes.txt", "not previewable")
    return buf.getvalue()


def _zip_with_wide_xlsx(xlsx_name="Wide.xlsx") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Wide"
    ws.append(["c1", "c2", "c3"])
    ws.append([1, 2, 3])
    xbuf = io.BytesIO()
    wb.save(xbuf)
    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(xlsx_name, xbuf.getvalue())
    return zbuf.getvalue()


def _give_role_with_permissions(
    user: User, role_name: str, perm_keys: list[str], org: Organization
) -> None:
    """Assign a permission role to `user` within `org`. Roles are scoped to an
    organization via tenants.Membership in the multi-tenant model."""
    user.profile.active_organization = org
    user.profile.save(update_fields=["active_organization"])
    role = Role.objects.create(name=role_name)
    for key in perm_keys:
        perm, _ = Permission.objects.get_or_create(
            key=key,
            defaults={"name": key, "module": "Processing", "description": ""},
        )
        role.permissions.add(perm)
    membership = Membership.objects.create(user=user, organization=org, status="active")
    membership.roles.add(role)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class Module1OutputPreviewApiTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        # owner and other live in separate orgs so the non-owner read is
        # blocked via org isolation.
        self.org = Organization.objects.create(name="owner-org")
        self.other_org = Organization.objects.create(name="other-org")
        self.owner = User.objects.create_user(
            username="owner",
            email="owner@example.com",
            password="testpass123",
        )
        _give_role_with_permissions(
            self.owner, "Owner Role", ["runhistory.view"], self.org
        )
        self.other = User.objects.create_user(
            username="other",
            email="other@example.com",
            password="testpass123",
        )
        _give_role_with_permissions(
            self.other, "Other Role", ["runhistory.view"], self.other_org
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.owner)

        self.job = Module1Job.objects.create(
            user=self.owner,
            organization=self.org,
            job_type=Module1Job.JobType.SUMMARY,
            status=Module1Job.Status.SUCCESS,
            input_meta={},
            work_dir="module1_jobs/test-preview",
        )
        self.job.output_zip.save(
            f"{self.job.id}.zip",
            ContentFile(_zip_with_xlsx()),
            save=True,
        )

    def test_output_files_happy_path(self):
        res = self.client.get(f"/api/module1/jobs/{self.job.id}/output/files/")
        self.assertEqual(res.status_code, 200, res.content)
        body = res.json()
        self.assertEqual(str(self.job.id), body["job_id"])
        self.assertEqual(1, len(body["files"]))
        self.assertEqual("Combined_Summary.xlsx", body["files"][0]["path"])

    def test_output_sheets_happy_path(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/sheets/?file=Combined_Summary.xlsx"
        )
        self.assertEqual(res.status_code, 200, res.content)
        body = res.json()
        names = [s["name"] for s in body["sheets"]]
        self.assertIn("Combined Summary", names)
        self.assertIn("UW Summary", names)

    def test_output_rows_happy_path(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Combined_Summary.xlsx&sheet=Combined Summary&page=1&page_size=1"
        )
        self.assertEqual(res.status_code, 200, res.content)
        body = res.json()
        self.assertEqual(2, body["total_rows"])
        self.assertEqual(["RESERVINGCLASS", "UWY", "Amount"], body["columns"])
        self.assertEqual(1, len(body["rows"]))

    def test_output_rows_invalid_page_size(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Combined_Summary.xlsx&sheet=Combined Summary&page=1&page_size=9999"
        )
        self.assertEqual(res.status_code, 400)

    def test_output_rows_page_size_boundary_at_configured_max(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Combined_Summary.xlsx&sheet=Combined Summary&page=1&page_size=200"
        )
        self.assertEqual(res.status_code, 200, res.content)

    def test_output_rows_invalid_file(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Missing.xlsx&sheet=Combined Summary&page=1&page_size=25"
        )
        self.assertEqual(res.status_code, 400)

    def test_output_rows_invalid_sheet(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Combined_Summary.xlsx&sheet=MissingSheet&page=1&page_size=25"
        )
        self.assertEqual(res.status_code, 400)

    def test_output_rows_out_of_range_page_returns_empty_rows(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Combined_Summary.xlsx&sheet=Combined Summary&page=99&page_size=25"
        )
        self.assertEqual(res.status_code, 200, res.content)
        body = res.json()
        self.assertEqual(0, len(body["rows"]))

    @patch("processing.views.settings.MODULE1_OUTPUT_PREVIEW_MAX_CELLS", 2)
    def test_output_rows_enforces_max_cells_limit(self):
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Combined_Summary.xlsx&sheet=Combined Summary&page=1&page_size=1"
        )
        self.assertEqual(res.status_code, 400)
        self.assertIn("cell limit", str(res.json()).lower())

    @patch("processing.views.settings.MODULE1_OUTPUT_PREVIEW_MAX_CELLS", 2)
    def test_output_rows_wide_sheet_guardrail(self):
        self.job.output_zip.save(
            f"{self.job.id}.zip",
            ContentFile(_zip_with_wide_xlsx()),
            save=True,
        )
        res = self.client.get(
            f"/api/module1/jobs/{self.job.id}/output/rows/?file=Wide.xlsx&sheet=Wide&page=1&page_size=1"
        )
        self.assertEqual(res.status_code, 400)
        self.assertIn("too wide", str(res.json()).lower())

    def test_output_preview_forbidden_without_read_permission(self):
        no_perm_user = User.objects.create_user(
            username="noperm",
            email="noperm@example.com",
            password="testpass123",
        )
        client = APIClient()
        client.force_authenticate(user=no_perm_user)
        res = client.get(f"/api/module1/jobs/{self.job.id}/output/files/")
        self.assertEqual(res.status_code, 403)

    def test_output_preview_hidden_for_non_owner(self):
        client = APIClient()
        client.force_authenticate(user=self.other)
        res = client.get(f"/api/module1/jobs/{self.job.id}/output/files/")
        self.assertEqual(res.status_code, 404)

    def test_output_preview_404_when_output_not_ready(self):
        job2 = Module1Job.objects.create(
            user=self.owner,
            organization=self.org,
            job_type=Module1Job.JobType.SUMMARY,
            status=Module1Job.Status.RUNNING,
            input_meta={},
            work_dir="module1_jobs/test-preview-2",
        )
        res = self.client.get(f"/api/module1/jobs/{job2.id}/output/files/")
        self.assertEqual(res.status_code, 404)
