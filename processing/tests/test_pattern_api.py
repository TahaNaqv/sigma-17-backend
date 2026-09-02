"""WP3a — payment pattern dataset, preview endpoint and job wiring."""

import io
import shutil
import tempfile
from unittest.mock import patch

import pandas as pd
from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from openpyxl import Workbook
from rest_framework.test import APIClient

from accounts.models import Permission, Role
from datasets.models import Dataset, PaymentPatternRow
from processing.models import Module1Job
from tenants.models import Membership, Organization

TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="sigma17-test-media-pattern-")


def _give_role(user, role_name, perm_keys, org):
    user.profile.active_organization = org
    user.profile.save(update_fields=["active_organization"])
    role = Role.objects.create(name=role_name)
    for key in perm_keys:
        perm, _ = Permission.objects.get_or_create(
            key=key, defaults={"name": key, "module": "Processing", "description": ""}
        )
        role.permissions.add(perm)
    membership = Membership.objects.create(user=user, organization=org, status="active")
    membership.roles.add(role)


def _wide_pattern_xlsx(rows: list[tuple[str, list[float]]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Pattern"
    width = max(len(v) for _, v in rows)
    ws.append(["RESERVINGCLASS"] + list(range(width)))
    for name, vec in rows:
        ws.append([name] + list(vec))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class PaymentPatternDatasetTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        self.org = Organization.objects.create(name="Pat", slug="pat")
        self.user = User.objects.create_user("pat", "pat@example.com", "pw")
        _give_role(self.user, "ActuaryPat",
                   ["module2.run", "datasets.view", "datasets.edit"], self.org)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _import(self, content: bytes, name="Pattern A"):
        f = io.BytesIO(content)
        f.name = "pattern.xlsx"
        return self.client.post("/api/datasets/import-excel/", {
            "kind": Dataset.Kind.PAYMENT_PATTERN,
            "name": name,
            "file": f,
        }, format="multipart")

    def test_wide_excel_is_unpivoted_into_long_rows(self):
        resp = self._import(_wide_pattern_xlsx([
            ("ENGINEERING", [0.4, 0.3, 0.2, 0.1]),
            ("MARINE", [0.5, 0.5, 0.0, 0.0]),
        ]))
        self.assertEqual(resp.status_code, 201, resp.data)
        ds = Dataset.objects.get(kind=Dataset.Kind.PAYMENT_PATTERN)
        rows = PaymentPatternRow.objects.filter(dataset=ds)
        self.assertEqual(rows.count(), 8)
        eng = {r.dev_period: float(r.weight) for r in rows if r.reserving_class == "ENGINEERING"}
        self.assertAlmostEqual(eng[0], 0.4)
        self.assertAlmostEqual(eng[3], 0.1)

    def test_blank_cells_are_skipped_not_stored_as_zero(self):
        """'Not supplied' and 'supplied as zero' are different statements."""
        wb = Workbook()
        ws = wb.active
        ws.append(["RESERVINGCLASS", 0, 1, 2])
        ws.append(["ENGINEERING", 0.6, None, 0.4])
        buf = io.BytesIO()
        wb.save(buf)
        resp = self._import(buf.getvalue())
        self.assertEqual(resp.status_code, 201, resp.data)
        ds = Dataset.objects.get(kind=Dataset.Kind.PAYMENT_PATTERN)
        periods = sorted(
            r.dev_period for r in PaymentPatternRow.objects.filter(dataset=ds)
        )
        self.assertEqual(periods, [0, 2])

    def test_sheet_without_a_class_column_is_rejected_with_guidance(self):
        wb = Workbook()
        wb.active.append(["CLASS", 0, 1])
        wb.active.append(["ENGINEERING", 0.5, 0.5])
        buf = io.BytesIO()
        wb.save(buf)
        resp = self._import(buf.getvalue())
        self.assertEqual(resp.status_code, 400)
        self.assertIn("RESERVINGCLASS", str(resp.data))

    def test_sheet_without_period_columns_is_rejected(self):
        wb = Workbook()
        wb.active.append(["RESERVINGCLASS", "notes"])
        wb.active.append(["ENGINEERING", "hello"])
        buf = io.BytesIO()
        wb.save(buf)
        resp = self._import(buf.getvalue())
        self.assertEqual(resp.status_code, 400)
        self.assertIn("development-period", str(resp.data))

    def test_template_is_wide_and_downloadable(self):
        resp = self.client.get(
            f"/api/datasets/templates/{Dataset.Kind.PAYMENT_PATTERN}/"
        )
        self.assertEqual(resp.status_code, 200)
        frame = pd.read_excel(io.BytesIO(resp.content), sheet_name="Payment Pattern")
        self.assertEqual(list(frame.columns)[0], "RESERVINGCLASS")
        self.assertGreater(len(frame.columns), 20)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class PatternJobWiringTests(TestCase):
    def setUp(self):
        self.org = Organization.objects.create(name="Wire", slug="wire")
        self.user = User.objects.create_user("wire", "wire@example.com", "pw")
        _give_role(self.user, "ActuaryWire",
                   ["module2.run", "datasets.view", "datasets.edit"], self.org)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _pattern_dataset(self):
        ds = Dataset.objects.create(
            organization=self.org, kind=Dataset.Kind.PAYMENT_PATTERN, name="P"
        )
        PaymentPatternRow.objects.bulk_create([
            PaymentPatternRow(dataset=ds, reserving_class="ENGINEERING",
                              dev_period=p, weight=w, row_index=p)
            for p, w in enumerate([0.4, 0.3, 0.2, 0.1])
        ])
        ds.refresh_row_count()
        return ds

    def _combined(self) -> io.BytesIO:
        wb = Workbook()
        wb.active.title = "Combined Summary"
        wb.active.append(["RESERVINGCLASS", "UWY"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        buf.name = "Combined_Summary.xlsx"
        return buf

    @patch("processing.views.run_module2_allocate_task")
    def test_allocate_snapshots_the_pattern_dataset(self, task):
        ds = self._pattern_dataset()
        resp = self.client.post("/api/module2/jobs/allocate/", {
            "combined_summary": self._combined(),
            "payment_pattern_dataset_id": str(ds.id),
        }, format="multipart")
        self.assertEqual(resp.status_code, 202, resp.data)
        job = Module1Job.objects.get(pk=resp.data["id"])
        snaps = job.input_meta["dataset_snapshots"]["payment_pattern"]
        self.assertEqual(len(snaps), 1)
        self.assertEqual(job.input_meta["pattern_mode"], "shape_only")

        # Editing the dataset afterwards must not change the stored run.
        PaymentPatternRow.objects.filter(dataset=ds).delete()
        job.refresh_from_db()
        self.assertEqual(
            len(job.input_meta["dataset_snapshots"]["payment_pattern"]), 1
        )

    @patch("processing.views.run_module2_allocate_task")
    def test_invalid_mode_is_rejected(self, task):
        ds = self._pattern_dataset()
        resp = self.client.post("/api/module2/jobs/allocate/", {
            "combined_summary": self._combined(),
            "payment_pattern_dataset_id": str(ds.id),
            "pattern_mode": "whatever",
        }, format="multipart")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("pattern_mode", resp.data["fieldErrors"])

    @patch("processing.views.run_module2_allocate_task")
    def test_a_dataset_of_the_wrong_kind_is_rejected(self, task):
        other = Dataset.objects.create(
            organization=self.org, kind=Dataset.Kind.PREMIUM, name="Not a pattern"
        )
        resp = self.client.post("/api/module2/jobs/allocate/", {
            "combined_summary": self._combined(),
            "payment_pattern_dataset_id": str(other.id),
        }, format="multipart")
        self.assertEqual(resp.status_code, 400)

    @patch("processing.views.run_module2_allocate_task")
    def test_no_pattern_supplied_leaves_meta_clean(self, task):
        resp = self.client.post("/api/module2/jobs/allocate/", {
            "combined_summary": self._combined(),
        }, format="multipart")
        self.assertEqual(resp.status_code, 202, resp.data)
        job = Module1Job.objects.get(pk=resp.data["id"])
        self.assertNotIn("pattern_mode", job.input_meta)
        self.assertNotIn(
            "payment_pattern", (job.input_meta.get("dataset_snapshots") or {})
        )


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class PatternPreviewAndTaskTests(TestCase):
    """Preview endpoint and the real task path, against the client reference book."""

    FIXTURE = (
        __import__("pathlib").Path(__file__).resolve().parents[2]
        / "benchmarks" / "fixtures" / "m2_allocate_ref" / "Combined_Summary.xlsx"
    )

    def setUp(self):
        if not self.FIXTURE.is_file():
            self.skipTest("reference fixture not available")
        self.org = Organization.objects.create(name="Prev", slug="prev")
        self.user = User.objects.create_user("prev", "prev@example.com", "pw")
        _give_role(self.user, "ActuaryPrev",
                   ["module2.run", "datasets.view", "datasets.edit"], self.org)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _allocate_job(self, pattern_ds=None):
        from processing.tasks import run_module2_allocate_task
        from processing.utils import init_module2_allocate_job_dirs

        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_ALLOCATE,
        )
        job.work_dir = f"module1_jobs/{job.id}"
        job.save(update_fields=["work_dir"])
        combined_dir = init_module2_allocate_job_dirs(job)
        (combined_dir / "Combined_Summary.xlsx").write_bytes(self.FIXTURE.read_bytes())
        meta = {"files": {}}
        if pattern_ds is not None:
            from datasets.services.snapshots import create_snapshot
            snap = create_snapshot(dataset=pattern_ds, consumer_job=job)
            meta["dataset_snapshots"] = {"payment_pattern": [str(snap.id)]}
            meta["pattern_mode"] = "shape_only"
        job.input_meta = meta
        job.save(update_fields=["input_meta"])
        run_module2_allocate_task(str(job.id))
        job.refresh_from_db()
        return job

    def test_preview_returns_both_the_engine_and_derived_curves(self):
        job = self._allocate_job()
        self.assertEqual(job.status, Module1Job.Status.SUCCESS, job.error_message)
        resp = self.client.get(f"/api/module2/jobs/{job.id}/payment-pattern/")
        self.assertEqual(resp.status_code, 200, resp.data)

        periods = resp.data["periods"]
        self.assertGreater(len(periods), 10)
        rows = {r["reserving_class"]: r for r in resp.data["rows"]}
        self.assertIn("ENGINEERING", rows)

        engine = rows["ENGINEERING"]["engine"]
        derived = rows["ENGINEERING"]["derived"]
        self.assertAlmostEqual(sum(engine), 1.0, places=6)
        self.assertAlmostEqual(sum(derived), 1.0, places=6)

        # The whole reason this endpoint returns two curves: they differ sharply.
        # The engine sheet front-loads (~48% in period 0) where the from-inception
        # pattern does not (~6%).
        self.assertGreater(engine[0], 0.4)
        self.assertLess(derived[0], 0.15)

    def test_preview_requires_a_successful_job(self):
        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_ALLOCATE,
        )
        resp = self.client.get(f"/api/module2/jobs/{job.id}/payment-pattern/")
        self.assertEqual(resp.status_code, 400)

    def test_task_applies_the_pattern_and_persists_the_report(self):
        ds = Dataset.objects.create(
            organization=self.org, kind=Dataset.Kind.PAYMENT_PATTERN, name="Long tail"
        )
        # A deliberately long-tailed pattern, so the effect is unmistakable.
        weights = [0.85 ** k for k in range(26)]
        total = sum(weights)
        PaymentPatternRow.objects.bulk_create([
            PaymentPatternRow(dataset=ds, reserving_class="ENGINEERING",
                              dev_period=p, weight=w / total, row_index=p)
            for p, w in enumerate(weights)
        ])
        ds.refresh_row_count()

        plain = self._allocate_job()
        shocked = self._allocate_job(pattern_ds=ds)
        self.assertEqual(shocked.status, Module1Job.Status.SUCCESS, shocked.error_message)

        report = shocked.input_meta["override_report"]
        self.assertEqual(report["applied_classes"], ["ENGINEERING"])
        self.assertEqual(report["unmatched_classes"], [])
        self.assertEqual(report["horizon"], 26)

        # The applied pattern must actually reach the output workbook.
        import zipfile
        def _pattern_sheet(job):
            with job.output_zip.open("rb") as f:
                data = f.read()
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                inner = zf.read("Module2_Allocate_Output.xlsx")
            frame = pd.read_excel(io.BytesIO(inner), sheet_name="Payment Pattern")
            return frame.set_index("RESERVINGCLASS")

        before = _pattern_sheet(plain)
        after = _pattern_sheet(shocked)
        self.assertNotAlmostEqual(
            float(before.loc["ENGINEERING"].iloc[0]),
            float(after.loc["ENGINEERING"].iloc[0]),
            places=6,
        )
        # A class with no override is untouched.
        self.assertAlmostEqual(
            float(before.loc["MARINE"].iloc[0]),
            float(after.loc["MARINE"].iloc[0]),
            places=9,
        )

    def test_an_unusable_pattern_fails_with_per_class_detail(self):
        ds = Dataset.objects.create(
            organization=self.org, kind=Dataset.Kind.PAYMENT_PATTERN, name="Zero"
        )
        PaymentPatternRow.objects.create(
            dataset=ds, reserving_class="ENGINEERING", dev_period=0, weight=0, row_index=0
        )
        ds.refresh_row_count()
        job = self._allocate_job(pattern_ds=ds)
        self.assertEqual(job.status, Module1Job.Status.FAILED)
        # Must name the class, not fall back to "check workbook formats".
        self.assertIn("ENGINEERING", job.error_message)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class PatternInheritanceAndPreviewScopeTests(TestCase):
    """Gaps found in the completeness audit — the disclosure must not silently
    disagree with the process output it is based on."""

    def setUp(self):
        self.org = Organization.objects.create(name="Inh", slug="inh")
        self.user = User.objects.create_user("inh", "inh@example.com", "pw")
        _give_role(self.user, "ActuaryInh", ["module2.run", "datasets.view"], self.org)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def _pattern_snapshot(self, job):
        from datasets.services.snapshots import create_snapshot

        ds = Dataset.objects.create(
            organization=self.org, kind=Dataset.Kind.PAYMENT_PATTERN, name="P"
        )
        PaymentPatternRow.objects.bulk_create([
            PaymentPatternRow(dataset=ds, reserving_class="ENGINEERING",
                              dev_period=p, weight=w, row_index=p)
            for p, w in enumerate([0.4, 0.3, 0.2, 0.1])
        ])
        ds.refresh_row_count()
        return create_snapshot(dataset=ds, consumer_job=job)

    def test_movement_inherits_the_pattern_its_process_job_used(self):
        """Without this, a movement disclosure re-runs the pipeline WITHOUT the
        pattern and publishes figures that disagree with the process output."""
        from processing.tasks import _load_pattern_override

        process = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_PROCESS,
            status=Module1Job.Status.SUCCESS,
        )
        snap = self._pattern_snapshot(process)
        process.input_meta = {
            "dataset_snapshots": {"payment_pattern": [str(snap.id)]},
            "pattern_mode": "shape_only",
        }
        process.save(update_fields=["input_meta"])

        movement = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_MOVEMENT,
            input_meta={},
        )
        # No pattern of its own → must inherit.
        self.assertIsNone(_load_pattern_override(movement))
        inherited = _load_pattern_override(movement, inherit_from=process)
        self.assertIsNotNone(inherited)
        self.assertTrue(inherited.has("ENGINEERING"))

    def test_an_explicit_pattern_on_the_movement_job_wins_over_inheritance(self):
        from processing.tasks import _load_pattern_override

        process = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_PROCESS,
            status=Module1Job.Status.SUCCESS,
        )
        proc_snap = self._pattern_snapshot(process)
        process.input_meta = {
            "dataset_snapshots": {"payment_pattern": [str(proc_snap.id)]}
        }
        process.save(update_fields=["input_meta"])

        movement = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_MOVEMENT,
        )
        own = Dataset.objects.create(
            organization=self.org, kind=Dataset.Kind.PAYMENT_PATTERN, name="Own"
        )
        PaymentPatternRow.objects.bulk_create([
            PaymentPatternRow(dataset=own, reserving_class="MARINE",
                              dev_period=p, weight=w, row_index=p)
            for p, w in enumerate([0.7, 0.3])
        ])
        own.refresh_row_count()
        from datasets.services.snapshots import create_snapshot
        own_snap = create_snapshot(dataset=own, consumer_job=movement)
        movement.input_meta = {
            "dataset_snapshots": {"payment_pattern": [str(own_snap.id)]}
        }
        movement.save(update_fields=["input_meta"])

        resolved = _load_pattern_override(movement, inherit_from=process)
        self.assertTrue(resolved.has("MARINE"))
        self.assertFalse(resolved.has("ENGINEERING"))

    def test_pattern_snapshots_are_not_staged_as_engine_input_sheets(self):
        """They are consumed as rows; staging them writes workbooks nothing reads."""
        from processing.tasks import _materialize_job_snapshots
        from processing.utils import job_root

        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_ALLOCATE,
        )
        job.work_dir = f"module1_jobs/{job.id}"
        job.save(update_fields=["work_dir"])
        snap = self._pattern_snapshot(job)
        job.input_meta = {"dataset_snapshots": {"payment_pattern": [str(snap.id)]}}
        job.save(update_fields=["input_meta"])

        _materialize_job_snapshots(job)
        self.assertFalse((job_root(job) / "in" / "payment_pattern").exists())

    def test_preview_on_a_process_job_reads_through_its_allocate_ancestor(self):
        """A process job's output has no Combined_Summary.xlsx — only its ancestor does."""
        process = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_PROCESS,
            status=Module1Job.Status.SUCCESS,
            source_job=None,
        )
        resp = self.client.get(f"/api/module2/jobs/{process.id}/payment-pattern/")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("allocate ancestor", str(resp.data))


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class PatternStrictModeTests(TestCase):
    """Strict mode is the opt-in for actuaries who want their weights taken literally."""

    FIXTURE = (
        __import__("pathlib").Path(__file__).resolve().parents[2]
        / "benchmarks" / "fixtures" / "m2_allocate_ref" / "Combined_Summary.xlsx"
    )

    def setUp(self):
        if not self.FIXTURE.is_file():
            self.skipTest("reference fixture not available")
        self.org = Organization.objects.create(name="Strict", slug="strict")
        self.user = User.objects.create_user("strict", "s@example.com", "pw")
        _give_role(self.user, "ActuaryStrict", ["module2.run", "datasets.view"], self.org)

    def _run(self, weights, mode):
        from datasets.services.snapshots import create_snapshot
        from processing.tasks import run_module2_allocate_task
        from processing.utils import init_module2_allocate_job_dirs

        ds = Dataset.objects.create(
            organization=self.org, kind=Dataset.Kind.PAYMENT_PATTERN, name=f"P-{mode}"
        )
        PaymentPatternRow.objects.bulk_create([
            PaymentPatternRow(dataset=ds, reserving_class="ENGINEERING",
                              dev_period=p, weight=w, row_index=p)
            for p, w in enumerate(weights)
        ])
        ds.refresh_row_count()

        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.MODULE2_ALLOCATE,
        )
        job.work_dir = f"module1_jobs/{job.id}"
        job.save(update_fields=["work_dir"])
        combined_dir = init_module2_allocate_job_dirs(job)
        (combined_dir / "Combined_Summary.xlsx").write_bytes(self.FIXTURE.read_bytes())
        snap = create_snapshot(dataset=ds, consumer_job=job)
        job.input_meta = {
            "files": {},
            "dataset_snapshots": {"payment_pattern": [str(snap.id)]},
            "pattern_mode": mode,
        }
        job.save(update_fields=["input_meta"])
        run_module2_allocate_task(str(job.id))
        job.refresh_from_db()
        return job

    def test_shape_only_renormalises_weights_that_do_not_sum_to_one(self):
        job = self._run([20, 30, 30, 20], "shape_only")
        self.assertEqual(job.status, Module1Job.Status.SUCCESS, job.error_message)
        self.assertAlmostEqual(
            job.input_meta["override_report"]["rescaled"]["ENGINEERING"], 100.0
        )

    def test_strict_rejects_weights_that_do_not_sum_to_one(self):
        job = self._run([20, 30, 30, 20], "strict")
        self.assertEqual(job.status, Module1Job.Status.FAILED)
        self.assertIn("ENGINEERING", job.error_message)
        self.assertIn("not 1", job.error_message)

    def test_strict_accepts_weights_that_already_sum_to_one(self):
        job = self._run([0.4, 0.3, 0.2, 0.1], "strict")
        self.assertEqual(job.status, Module1Job.Status.SUCCESS, job.error_message)
        self.assertEqual(job.input_meta["override_report"]["rescaled"], {})
