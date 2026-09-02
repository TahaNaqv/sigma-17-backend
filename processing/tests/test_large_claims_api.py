"""WP5 — large-claims endpoint, exclusion persistence, and the end-to-end engine effect."""

import io
import json
import shutil
import tempfile
import zipfile

import pandas as pd
from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from rest_framework.test import APIClient

from accounts.models import Permission, Role
from processing.models import Module1Job
from tenants.models import Membership, Organization

TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="sigma17-test-media-lc-")
FIXTURES = (
    __import__("pathlib").Path(__file__).resolve().parents[2]
    / "benchmarks" / "fixtures" / "summary_ref"
)


def _give_role(user, role_name, perm_keys, org):
    user.profile.active_organization = org
    user.profile.save(update_fields=["active_organization"])
    role = Role.objects.create(name=role_name)
    for key in perm_keys:
        perm, _ = Permission.objects.get_or_create(
            key=key, defaults={"name": key, "module": "Processing", "description": ""}
        )
        role.permissions.add(perm)
    membership = Membership.objects.create(user=user, organization=org, status="active")
    membership.roles.add(role)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class LargeClaimsApiTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)
        super().tearDownClass()

    def setUp(self):
        if not (FIXTURES / "claims_paid").is_dir():
            self.skipTest("reference fixture not available")
        self.org = Organization.objects.create(name="LC", slug="lc")
        self.user = User.objects.create_user("lc", "lc@example.com", "pw")
        _give_role(self.user, "ActuaryLC", ["module1.run"], self.org)
        self.client = APIClient()
        self.client.force_authenticate(self.user)
        self.job = self._job()

    def _job(self):
        from processing.utils import init_job_work_dir, job_input_subdir

        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.SUMMARY,
            status=Module1Job.Status.SUCCESS,
            input_meta={"exp_start": "01-01-2016", "exp_end": "31-12-2017"},
        )
        job.work_dir = f"module1_jobs/{job.id}"
        job.save(update_fields=["work_dir"])
        init_job_work_dir(job)
        for kind in ("claims_paid", "claims_os"):
            dest = job_input_subdir(job, kind)
            for f in (FIXTURES / kind).glob("*.xlsx"):
                shutil.copy(f, dest / f.name)
        return job

    def _get(self, **params):
        query = "&".join(f"{k}={v}" for k, v in params.items())
        return self.client.get(f"/api/module1/jobs/{self.job.id}/large-claims/?{query}")

    def test_returns_ranked_claims_with_concentration(self):
        resp = self._get(per_class=0, top_n=10, rank_on="paid")
        self.assertEqual(resp.status_code, 200, resp.data)
        report = resp.data["report"]
        self.assertEqual(len(report["claims"]), 10)
        # The measured paid-ranked concentration.
        self.assertAlmostEqual(report["concentration"], 0.223, places=2)

    def test_the_slice_is_reported_so_the_ranking_basis_is_never_implicit(self):
        report = self._get(per_class=0).data["report"]
        self.assertEqual(report["slice"], {"treaty": "GROSS", "head_of_damage": "Payment"})
        self.assertEqual(report["selection"]["rank_on"], "incurred")

    def test_ranking_on_incurred_surfaces_claims_with_no_payment(self):
        incurred = self._get(per_class=0, top_n=10, rank_on="incurred").data["report"]
        paid_only = self._get(per_class=0, top_n=10, rank_on="paid").data["report"]
        self.assertTrue(any(c["paid"] == 0 for c in incurred["claims"]))
        self.assertTrue(all(c["paid"] > 0 for c in paid_only["claims"]))

    def test_per_class_ranking_covers_more_classes(self):
        book = self._get(per_class=0, top_n=3).data["report"]
        per_class = self._get(per_class=1, top_n=3).data["report"]
        self.assertGreater(
            len({c["reserving_class"] for c in per_class["claims"]}),
            len({c["reserving_class"] for c in book["claims"]}),
        )

    def test_the_modes_are_served_with_their_measured_consequences(self):
        modes = {m["key"]: m for m in self._get().data["modes"]}
        self.assertEqual(set(modes), {
            "exclude_and_add_back", "exclude_from_ldf_only", "exclude_entirely",
        })
        # The dangerous mode must state its direction and magnitude.
        self.assertIn("RAISE", modes["exclude_from_ldf_only"]["note"])
        self.assertIn("13.3%", modes["exclude_from_ldf_only"]["note"])

    def test_threshold_selection_requires_a_threshold(self):
        self.assertEqual(self._get(kind="threshold").status_code, 400)

    def test_invalid_parameters_are_rejected(self):
        self.assertEqual(self._get(kind="vibes").status_code, 400)
        self.assertEqual(self._get(rank_on="mood").status_code, 400)
        self.assertEqual(self._get(top_n=0).status_code, 400)
        self.assertEqual(self._get(top_n=99999).status_code, 400)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class ExclusionPersistenceTests(TestCase):
    def setUp(self):
        self.org = Organization.objects.create(name="Ex", slug="ex")
        self.user = User.objects.create_user("ex", "ex@example.com", "pw")
        _give_role(self.user, "ActuaryEx", ["module1.run"], self.org)

    def test_the_loader_rebuilds_the_plan_from_the_job(self):
        from processing.tasks import _load_exclusion

        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.SUMMARY,
            input_meta={"large_claims": {
                "mode": "exclude_from_ldf_only",
                "claim_numbers": ["A", "B"],
            }},
        )
        plan = _load_exclusion(job)
        self.assertEqual(plan.mode, "exclude_from_ldf_only")
        self.assertTrue(plan.filters_triangles)
        self.assertFalse(plan.filters_base)

    def test_a_job_without_an_exclusion_returns_none(self):
        from processing.tasks import _load_exclusion

        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.SUMMARY, input_meta={},
        )
        self.assertIsNone(_load_exclusion(job))

    def test_the_default_mode_is_add_back(self):
        from processing.tasks import _load_exclusion

        job = Module1Job.objects.create(
            user=self.user, organization=self.org,
            job_type=Module1Job.JobType.SUMMARY,
            input_meta={"large_claims": {"claim_numbers": ["A"]}},
        )
        plan = _load_exclusion(job)
        self.assertEqual(plan.mode, "exclude_and_add_back")
        self.assertTrue(plan.adds_back)


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class ExclusionReachesTheWorkbookTests(TestCase):
    """The gap that unit tests could not see.

    `adds_back` was a correct property that nothing consumed: the default mode filtered the
    triangles, never added the cost back, and so behaved exactly like `exclude_entirely`.
    Every ExclusionPlan unit test passed the whole time. These run the real engine and read
    the produced workbook.
    """

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.out = {}

    @staticmethod
    def _run(mode, claims):
        import tempfile as _tf
        from pathlib import Path

        from module1_engine.engine import run_generate_summary, run_update_reserve_summary
        from module1_engine.large_claims import ExclusionPlan

        out_dir = _tf.mkdtemp(prefix=f"lc-{mode}-")
        report: dict = {}
        run_generate_summary(
            "01-01-2016", "31-12-2017", "01-01-2016", "31-12-2017",
            str(FIXTURES / "premium"), str(FIXTURES / "claims_paid"),
            str(FIXTURES / "claims_os"), out_dir,
            exclusion=ExclusionPlan.build(claims, mode) if claims else None,
            run_report=report,
        )
        run_update_reserve_summary(out_dir)
        return Path(out_dir), report

    @staticmethod
    def _headers(path):
        from openpyxl import load_workbook

        for f in sorted(path.glob("*.xlsx")):
            if f.name.startswith("Combined"):
                continue
            wb = load_workbook(f)
            if "Reserve Summary" not in wb.sheetnames:
                continue
            ws = wb["Reserve Summary"]
            return [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        return []

    def test_add_back_writes_the_large_claim_columns_into_every_reserve_workbook(self):
        claims = ["SIL/D/C003/0000001213/0317/001", "SIL/D/C003/0000000821/1016/001"]
        path, _ = self._run("exclude_and_add_back", claims)
        headers = self._headers(path)
        self.assertIn("Large Paid", headers)
        self.assertIn("Large OS", headers)
        # And the appended block still sits immediately after them, in order.
        self.assertEqual(headers[headers.index("Large OS") + 1], "Implied LR")
        shutil.rmtree(path, ignore_errors=True)

    def test_a_run_without_an_exclusion_keeps_the_six_historic_base_columns(self):
        path, report = self._run("exclude_and_add_back", [])
        headers = self._headers(path)
        self.assertNotIn("Large Paid", headers)
        self.assertEqual(headers[:6], [
            "Accident_Period", "EP", "Paid Claims", "OS Claims", "Reported Claims",
            "Reported LR",
        ])
        self.assertEqual(report, {})
        shutil.rmtree(path, ignore_errors=True)

    def test_a_selection_that_matches_nothing_is_reported_not_silently_ignored(self):
        """The one failure mode of this feature that is invisible in the output."""
        path, report = self._run("exclude_and_add_back", ["NOT-A-REAL-CLAIM"])
        match = report["large_claims_match"]
        self.assertEqual(match["requested"], 1)
        self.assertEqual(match["matched"], 0)
        self.assertEqual(match["unmatched"], ["NOT-A-REAL-CLAIM"])
        shutil.rmtree(path, ignore_errors=True)

    def test_the_match_report_counts_real_claims(self):
        path, report = self._run(
            "exclude_and_add_back",
            ["SIL/D/C003/0000001213/0317/001", "NOT-A-REAL-CLAIM"],
        )
        match = report["large_claims_match"]
        self.assertEqual((match["requested"], match["matched"]), (2, 1))
        self.assertEqual(match["unmatched"], ["NOT-A-REAL-CLAIM"])
        self.assertEqual(match["mode"], "exclude_and_add_back")
        shutil.rmtree(path, ignore_errors=True)

    def test_exclude_entirely_filters_the_base_instead_of_adding_columns(self):
        claims = ["SIL/D/C003/0000001213/0317/001"]
        path, _ = self._run("exclude_entirely", claims)
        self.assertNotIn("Large Paid", self._headers(path))
        shutil.rmtree(path, ignore_errors=True)
