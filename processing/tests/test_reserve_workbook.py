"""Tests for the reserve-workbook service (Excel-free Update Reserve).

These tests build synthetic reserve workbooks that mirror the structure
the Summary engine produces:
  - sheets `Paid Claims Triangle` and `Reported Triangle`
  - row 1 = column headers
  - somewhere below, a row whose column 1 is the literal "Selected CDF"

We then exercise the three public service functions: list, read, and
apply-overrides + write. The Celery task is integration-tested via the
existing chaining tests; this file is the unit contract for the
workbook adapter itself.
"""

import io
import shutil
import tempfile
import zipfile
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.test import TestCase, override_settings
from openpyxl import Workbook, load_workbook

from processing.models import Module1Job
from processing.services.reserve_workbook import (
    list_reserve_workbooks,
    read_reserve_summary_rows,
    read_workbook_cdfs,
    write_workbooks_with_overrides,
)
from tenants.models import Organization

# Reserve Summary base columns, in the order the Summary engine writes them.
_SUMMARY_HEADERS = [
    "Accident_Period", "EP", "Paid Claims", "OS Claims", "Reported Claims", "Reported LR",
]

User = get_user_model()

TEST_MEDIA_ROOT = tempfile.mkdtemp(prefix="sigma17-reserve-test-")


def _build_reserve_workbook(
    *,
    paid_cdf_values: list[float],
    reported_cdf_values: list[float],
    paid_headers: list[str] | None = None,
) -> bytes:
    """Build a minimal reserve workbook with the structure the engine
    produces. We only need the two triangle sheets and the Selected CDF
    row; the reserve service ignores everything else."""
    wb = Workbook()
    # The default sheet is "Sheet" — rename to Paid Claims Triangle
    ws_paid = wb.active
    ws_paid.title = "Paid Claims Triangle"
    headers = paid_headers or ["Accident Period"] + [str(i) for i in range(len(paid_cdf_values))]
    ws_paid.append(headers)
    # A few placeholder data rows so max_row > the cdf row
    ws_paid.append(["2022"] + [0] * len(paid_cdf_values))
    ws_paid.append(["2023"] + [0] * len(paid_cdf_values))
    ws_paid.append(["Selected CDF"] + paid_cdf_values)

    ws_rep = wb.create_sheet("Reported Triangle")
    ws_rep.append(["Accident Period"] + [str(i) for i in range(len(reported_cdf_values))])
    ws_rep.append(["2022"] + [0] * len(reported_cdf_values))
    ws_rep.append(["Selected CDF"] + reported_cdf_values)

    # Add a third sheet for realism — the service should ignore it.
    wb.create_sheet("Reserve Summary")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_reserve_workbook_with_summary(
    *,
    paid_cdf_values: list[float],
    reported_cdf_values: list[float],
    summary_rows: list[list],
) -> bytes:
    """Reserve workbook with a populated Reserve Summary sheet (A–F), matching
    what the Summary engine emits, plus the two triangle sheets whose Selected
    CDF rows carry literal values. `summary_rows` are newest-accident-period
    first, each = [Accident_Period, EP, Paid, OS, Reported, Reported LR]."""
    wb = Workbook()
    ws_paid = wb.active
    ws_paid.title = "Paid Claims Triangle"
    ws_paid.append(["Accident Period"] + [str(i) for i in range(len(paid_cdf_values))])
    ws_paid.append(["2022"] + [0] * len(paid_cdf_values))
    ws_paid.append(["Selected CDF"] + paid_cdf_values)

    ws_rep = wb.create_sheet("Reported Triangle")
    ws_rep.append(["Accident Period"] + [str(i) for i in range(len(reported_cdf_values))])
    ws_rep.append(["2022"] + [0] * len(reported_cdf_values))
    ws_rep.append(["Selected CDF"] + reported_cdf_values)

    ws_sum = wb.create_sheet("Reserve Summary")
    ws_sum.append(_SUMMARY_HEADERS)
    for r in summary_rows:
        ws_sum.append(r)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_triangle_workbook_with_ldf(*, n_cols: int = 3, ldf_cell: str = "=1") -> bytes:
    """Reserve workbook whose triangle sheets carry a Selected LDF row (engine
    placeholder `=1` by default) and a Selected CDF row of PRODUCT formulas —
    the exact structure the Summary engine emits."""
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    for i, title in enumerate(("Paid Claims Triangle", "Reported Triangle")):
        ws = wb.active if i == 0 else wb.create_sheet(title)
        if i == 0:
            ws.title = title
        ws.append(["Accident Period"] + [str(c) for c in range(n_cols)])
        ws.append(["2022"] + [0] * n_cols)
        ldf_row = ws.max_row + 1
        ws.append(["Selected LDF"] + [ldf_cell] * n_cols)
        last = get_column_letter(1 + n_cols)
        cdf_cells = ["Selected CDF"] + [
            f"=PRODUCT({get_column_letter(2 + c)}{ldf_row}:{last}{ldf_row})"
            for c in range(n_cols)
        ]
        ws.append(cdf_cells)
    wb.create_sheet("Reserve Summary")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _zip_with(files: dict[str, bytes]) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, raw in files.items():
            zf.writestr(name, raw)
    return buf.getvalue()


def _make_source_job(*, user, org, zip_bytes: bytes) -> Module1Job:
    job = Module1Job.objects.create(
        user=user,
        organization=org,
        job_type=Module1Job.JobType.SUMMARY,
        status=Module1Job.Status.SUCCESS,
        work_dir=f"module1_jobs/test-{org.id}",
    )
    job.output_zip.save(f"{job.id}.zip", ContentFile(zip_bytes), save=False)
    job.save()
    return job


@override_settings(MEDIA_ROOT=TEST_MEDIA_ROOT, SECURE_SSL_REDIRECT=False)
class ReserveWorkbookServiceTests(TestCase):
    @classmethod
    def tearDownClass(cls):
        super().tearDownClass()
        shutil.rmtree(TEST_MEDIA_ROOT, ignore_errors=True)

    def setUp(self):
        self.org = Organization.objects.create(name="ResOrg")
        self.user = User.objects.create_user(username="ru", password="testpass123")

    # ---- list ----

    def test_list_workbooks_parses_engine_naming(self):
        wb = _build_reserve_workbook(
            paid_cdf_values=[2.0, 1.5, 1.1],
            reported_cdf_values=[2.1, 1.6, 1.0],
        )
        zb = _zip_with({
            "Combined_Summary.xlsx": b"placeholder",  # should be excluded
            "Motor TP GROSS 2024-12.xlsx": wb,
            "Multi Word Class TP GROSS 2024-12.xlsx": wb,
            "junk.txt": b"x",  # not xlsx
            "Bad.xlsx": b"x",  # not enough name parts
        })
        job = _make_source_job(user=self.user, org=self.org, zip_bytes=zb)

        listed = list_reserve_workbooks(job)
        names = sorted(w["filename"] for w in listed)
        self.assertEqual(
            names,
            sorted([
                "Motor TP GROSS 2024-12.xlsx",
                "Multi Word Class TP GROSS 2024-12.xlsx",
            ]),
        )
        single = next(w for w in listed if w["filename"] == "Motor TP GROSS 2024-12.xlsx")
        self.assertEqual(single["reserving_class"], "Motor")
        self.assertEqual(single["head_of_damage"], "TP")
        self.assertEqual(single["ri_type"], "GROSS")
        self.assertEqual(single["eop_label"], "2024-12")
        multi = next(w for w in listed if w["reserving_class"] == "Multi Word Class")
        self.assertEqual(multi["head_of_damage"], "TP")

    # ---- read CDFs ----

    def test_read_cdfs_returns_values_and_headers(self):
        wb = _build_reserve_workbook(
            paid_cdf_values=[2.0, 1.5, 1.1],
            reported_cdf_values=[2.1, 1.6, 1.0],
        )
        job = _make_source_job(
            user=self.user,
            org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": wb}),
        )

        payload = read_workbook_cdfs(job, "Motor TP GROSS 2024-12.xlsx")
        self.assertEqual(payload["reserving_class"], "Motor")
        triangles = {t["sheet"]: t for t in payload["triangles"]}
        self.assertIn("Paid Claims Triangle", triangles)
        self.assertIn("Reported Triangle", triangles)
        self.assertEqual(triangles["Paid Claims Triangle"]["values"], [2.0, 1.5, 1.1])
        self.assertEqual(triangles["Reported Triangle"]["values"], [2.1, 1.6, 1.0])
        # Headers come from row 1 of each sheet.
        self.assertEqual(
            triangles["Paid Claims Triangle"]["column_labels"],
            ["0", "1", "2"],
        )

    # ---- apply overrides ----

    def test_overrides_replace_selected_cdf_values(self):
        wb = _build_reserve_workbook(
            paid_cdf_values=[2.0, 1.5, 1.1],
            reported_cdf_values=[2.1, 1.6, 1.0],
        )
        job = _make_source_job(
            user=self.user,
            org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": wb}),
        )

        dest = Path(tempfile.mkdtemp(prefix="reserve-override-out-"))
        try:
            written = write_workbooks_with_overrides(
                source_job=job,
                overrides_by_filename={
                    "Motor TP GROSS 2024-12.xlsx": {
                        "Paid Claims Triangle": [3.0, 2.5, 1.2],
                        "Reported Triangle": [3.5, 2.0, 1.1],
                    }
                },
                dest_folder=dest,
            )
            self.assertIn("Motor TP GROSS 2024-12.xlsx", written)

            # Open the modified workbook and confirm the Selected CDF
            # row got rewritten to the user's values.
            wbcheck = load_workbook(dest / "Motor TP GROSS 2024-12.xlsx", data_only=True)
            paid_ws = wbcheck["Paid Claims Triangle"]
            # Find the Selected CDF row
            sel_row = None
            for row in paid_ws.iter_rows(min_row=1, max_row=paid_ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value == "Selected CDF":
                        sel_row = cell.row
                        break
                if sel_row:
                    break
            self.assertIsNotNone(sel_row)
            values = [paid_ws.cell(row=sel_row, column=c).value for c in range(2, 5)]
            self.assertEqual(values, [3.0, 2.5, 1.2])
        finally:
            shutil.rmtree(dest, ignore_errors=True)

    # ---- reserve summary rows (Implied LR / Selected Method editor) ----

    def _summary_job(self):
        wb = _build_reserve_workbook_with_summary(
            paid_cdf_values=[1.1, 1.5],       # reversed → [1.5, 1.1]
            reported_cdf_values=[1.2, 1.6],   # reversed → [1.6, 1.2]
            summary_rows=[
                ["2024Q1", 1000, 300, 50, 350, 0.35],
                ["2023Q1", 2000, 800, 100, 900, 0.45],
            ],
        )
        return _make_source_job(
            user=self.user, org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": wb}),
        )

    def test_read_reserve_summary_rows_derives_cdf_and_columns(self):
        job = self._summary_job()
        payload = read_reserve_summary_rows(job, "Motor TP GROSS 2024-12.xlsx")
        self.assertEqual(payload["reserving_class"], "Motor")
        rows = payload["rows"]
        self.assertEqual(len(rows), 2)
        r0 = rows[0]
        self.assertEqual(r0["accident_period"], "2024Q1")
        self.assertEqual(r0["ep"], 1000)
        self.assertEqual(r0["paid_claims"], 300)
        self.assertEqual(r0["os_claims"], 50)
        self.assertEqual(r0["reported_claims"], 350)
        self.assertAlmostEqual(r0["reported_lr"], 0.35)
        # newest-first row gets the reversed series' first element
        self.assertEqual(r0["paid_cdf"], 1.5)
        self.assertEqual(r0["reported_cdf"], 1.6)
        self.assertEqual(rows[1]["paid_cdf"], 1.1)
        self.assertEqual(rows[1]["reported_cdf"], 1.2)

    def test_read_reserve_summary_rows_reflects_cdf_overrides(self):
        job = self._summary_job()
        payload = read_reserve_summary_rows(
            job, "Motor TP GROSS 2024-12.xlsx",
            cdf_overrides={"Paid Claims Triangle": [2.0, 3.0]},  # reversed → [3.0, 2.0]
        )
        rows = payload["rows"]
        self.assertEqual(rows[0]["paid_cdf"], 3.0)
        self.assertEqual(rows[1]["paid_cdf"], 2.0)
        # Reported CDF untouched by a Paid-only override.
        self.assertEqual(rows[0]["reported_cdf"], 1.6)

    def test_engine_applies_method_overrides_and_reader_matches(self):
        """run_update_reserve_summary writes the chosen Implied LR + Selected
        Method into cells G/O (defaults elsewhere), and the CDF the engine
        writes (H) equals what read_reserve_summary_rows previews."""
        from openpyxl import load_workbook as _load

        from module1_engine import run_update_reserve_summary

        job = self._summary_job()
        preview = read_reserve_summary_rows(job, "Motor TP GROSS 2024-12.xlsx")["rows"]

        staging = Path(tempfile.mkdtemp(prefix="reserve-method-out-"))
        try:
            # Stage the source workbook (verbatim) into the engine folder.
            written = write_workbooks_with_overrides(
                source_job=job, overrides_by_filename={}, dest_folder=staging
            )
            self.assertIn("Motor TP GROSS 2024-12.xlsx", written)

            run_update_reserve_summary(
                str(staging),
                method_overrides={
                    "Motor TP GROSS 2024-12.xlsx": {
                        "2024Q1": {"implied_lr": 0.6, "selected_method": "ELR"},
                    }
                },
            )

            wb = _load(staging / "Motor TP GROSS 2024-12.xlsx", data_only=False)
            ws = wb["Reserve Summary"]
            hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
            g, o, h = hdr["Implied LR"], hdr["Selected Method"], hdr["Paid CDF"]
            # Row 2 = 2024Q1 (overridden); Row 3 = 2023Q1 (defaults).
            self.assertEqual(ws.cell(row=2, column=g).value, 0.6)
            self.assertEqual(ws.cell(row=2, column=o).value, "ELR")
            self.assertIsNone(ws.cell(row=3, column=g).value)
            self.assertEqual(ws.cell(row=3, column=o).value, "Paid CL")
            # Parity: engine's Paid CDF (H) == the reader's preview.
            self.assertEqual(ws.cell(row=2, column=h).value, preview[0]["paid_cdf"])
            self.assertEqual(ws.cell(row=3, column=h).value, preview[1]["paid_cdf"])
        finally:
            shutil.rmtree(staging, ignore_errors=True)

    def test_engine_ignores_unknown_accident_period_and_bad_method(self):
        from openpyxl import load_workbook as _load

        from module1_engine import run_update_reserve_summary

        job = self._summary_job()
        staging = Path(tempfile.mkdtemp(prefix="reserve-method-bad-"))
        try:
            write_workbooks_with_overrides(
                source_job=job, overrides_by_filename={}, dest_folder=staging
            )
            run_update_reserve_summary(
                str(staging),
                method_overrides={
                    "Motor TP GROSS 2024-12.xlsx": {
                        "9999Q9": {"implied_lr": 0.9, "selected_method": "ELR"},  # no such row
                        "2024Q1": {"selected_method": "NOT A METHOD"},            # invalid → ignored
                    }
                },
            )
            wb = _load(staging / "Motor TP GROSS 2024-12.xlsx", data_only=False)
            ws = wb["Reserve Summary"]
            hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
            # 2024Q1 keeps the default method; unknown period changed nothing.
            self.assertEqual(ws.cell(row=2, column=hdr["Selected Method"]).value, "Paid CL")
            self.assertIsNone(ws.cell(row=2, column=hdr["Implied LR"]).value)
        finally:
            shutil.rmtree(staging, ignore_errors=True)

    # ---- Selected LDF editing (derived CDF) ----

    def test_selected_cdf_from_ldf_suffix_product(self):
        from module1_engine import selected_cdf_from_ldf

        # cdf[i] = product(ldf[i:])
        self.assertEqual(
            [round(x, 6) for x in selected_cdf_from_ldf([1.05, 1.02, 1.0])],
            [1.071, 1.02, 1.0],
        )
        # blanks are treated as 1.0 (no development)
        self.assertEqual(selected_cdf_from_ldf([2.0, None, 1.0]), [2.0, 1.0, 1.0])
        self.assertEqual(selected_cdf_from_ldf([]), [])

    def test_read_cdfs_includes_selected_ldf_placeholder(self):
        wb = _build_triangle_workbook_with_ldf(n_cols=3)  # LDF cells = "=1"
        job = _make_source_job(
            user=self.user, org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": wb}),
        )
        payload = read_workbook_cdfs(job, "Motor TP GROSS 2024-12.xlsx")
        paid = next(t for t in payload["triangles"] if t["sheet"] == "Paid Claims Triangle")
        # placeholder `=1` (uncached) surfaces as 1.0, not blank
        self.assertEqual(paid["selected_ldf"], [1.0, 1.0, 1.0])
        self.assertIsNotNone(paid["ldf_row"])
        # PRODUCT CDF formulas are uncached → None (unchanged behaviour)
        self.assertEqual(paid["values"], [None, None, None])

    def test_ldf_override_writes_ldf_row_and_derived_cdf(self):
        from openpyxl import load_workbook as _load

        wb = _build_triangle_workbook_with_ldf(n_cols=3)
        job = _make_source_job(
            user=self.user, org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": wb}),
        )
        dest = Path(tempfile.mkdtemp(prefix="reserve-ldf-out-"))
        try:
            write_workbooks_with_overrides(
                source_job=job,
                overrides_by_filename={},
                ldf_overrides_by_filename={
                    "Motor TP GROSS 2024-12.xlsx": {
                        "Paid Claims Triangle": [1.05, 1.02, 1.0],
                    }
                },
                dest_folder=dest,
            )
            wbc = _load(dest / "Motor TP GROSS 2024-12.xlsx", data_only=True)
            ws = wbc["Paid Claims Triangle"]

            def _row(label):
                for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                    if r[0].value == label:
                        return r[0].row
                return None

            ldf_row, cdf_row = _row("Selected LDF"), _row("Selected CDF")
            self.assertEqual([ws.cell(row=ldf_row, column=c).value for c in range(2, 5)], [1.05, 1.02, 1.0])
            derived = [ws.cell(row=cdf_row, column=c).value for c in range(2, 5)]
            self.assertEqual([round(x, 6) for x in derived], [1.071, 1.02, 1.0])
        finally:
            shutil.rmtree(dest, ignore_errors=True)

    def test_ldf_override_flows_into_reserve_summary_cdf(self):
        """End-to-end: an LDF override's derived CDF is what the engine reads for
        the per-accident-period Paid CDF."""
        from module1_engine import run_update_reserve_summary
        from processing.services.reserve_workbook import read_reserve_summary_rows

        # Build a workbook that also has a Reserve Summary with 1 row so the
        # engine has something to compute; 3 dev cols → 3-length CDF series.
        wb = _build_reserve_workbook_with_summary(
            paid_cdf_values=[1.0, 1.0, 1.0],       # ignored once LDF override applies
            reported_cdf_values=[1.0, 1.0, 1.0],
            summary_rows=[["2024Q1", 1000, 300, 50, 350, 0.35]],
        )
        # Give it a real Selected LDF row so the override has somewhere to write.
        job = _make_source_job(
            user=self.user, org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": wb}),
        )
        staging = Path(tempfile.mkdtemp(prefix="reserve-ldf-e2e-"))
        try:
            write_workbooks_with_overrides(
                source_job=job,
                overrides_by_filename={},
                ldf_overrides_by_filename={
                    "Motor TP GROSS 2024-12.xlsx": {"Paid Claims Triangle": [1.05, 1.02, 1.0]},
                },
                dest_folder=staging,
            )
            run_update_reserve_summary(str(staging))
            wbc = load_workbook(staging / "Motor TP GROSS 2024-12.xlsx", data_only=True)
            ws = wbc["Reserve Summary"]
            hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
            # 1 summary row (idx 0). Its Paid CDF = reversed(derived CDF)[0].
            # derived CDF = [1.071, 1.02, 1.0]; reversed → [1.0, 1.02, 1.071]; idx0 = 1.0.
            self.assertAlmostEqual(ws.cell(row=2, column=hdr["Paid CDF"]).value, 1.0, places=4)
        finally:
            shutil.rmtree(staging, ignore_errors=True)

    # ---- triangle grid (LDF history context) + LDF→methods linkage ----

    def test_read_cdfs_returns_full_triangle_grid_with_history(self):
        """The grid carries the LDF history (age-to-age + benchmark rows) so the
        UI can show the selection in its triangle context; ldf_row/cdf_row index
        into it (1-based)."""
        from openpyxl import Workbook as _WB
        from openpyxl.utils import get_column_letter

        # Build a triangle that mirrors the engine's block stack, including an
        # age-to-age history block and the Simple/Weighted Avg benchmark rows.
        wb = _WB()
        ws = wb.active
        ws.title = "Paid Claims Triangle"
        ws.append(["Accident Period", "0", "1", "2"])       # row 1 header
        ws.append(["2017-Q1", 8.15, 1.01, 1.05])            # row 2  history
        ws.append(["2017-Q2", 1.85, 1.14, 0.0])             # row 3  history
        ws.append(["Simple Avg LDF", 6.78, 0.54, 0.26])     # row 4  benchmark
        ws.append(["Simple Avg CDF", 0.96, 0.14, 0.26])     # row 5  benchmark
        ws.append(["Weighted Avg LDF", None, 4.79, 1.06])   # row 6  benchmark
        ws.append(["Weighted Avg CDF", None, 5.36, 1.11])   # row 7  benchmark
        ldf_row = 8
        ws.append(["Selected LDF", "=1", "=1", "=1"])       # row 8
        last = get_column_letter(4)
        ws.append(["Selected CDF"] + [
            f"=PRODUCT({get_column_letter(2 + c)}{ldf_row}:{last}{ldf_row})" for c in range(3)
        ])                                                   # row 9
        wb.create_sheet("Reported Triangle").append(["Accident Period", "0"])
        wb.create_sheet("Reserve Summary")
        buf = io.BytesIO()
        wb.save(buf)

        job = _make_source_job(
            user=self.user, org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": buf.getvalue()}),
        )
        payload = read_workbook_cdfs(job, "Motor TP GROSS 2024-12.xlsx")
        paid = next(t for t in payload["triangles"] if t["sheet"] == "Paid Claims Triangle")

        self.assertFalse(paid["grid_truncated"])
        grid = paid["grid"]
        self.assertEqual(len(grid), 9)          # rows 1..9
        self.assertEqual(len(grid[0]), 4)       # cols A..D
        # The LDF history is present, exactly as Excel holds it.
        self.assertEqual(grid[1][0], "2017-Q1")
        self.assertEqual(grid[1][1], 8.15)
        # The benchmark rows the actuary selects against are present.
        self.assertEqual(grid[3][0], "Simple Avg LDF")
        self.assertEqual(grid[5][0], "Weighted Avg LDF")
        # ldf_row / cdf_row are 1-based indices into the grid.
        self.assertEqual(paid["ldf_row"], 8)
        self.assertEqual(paid["cdf_row"], 9)
        self.assertEqual(grid[paid["ldf_row"] - 1][0], "Selected LDF")
        self.assertEqual(grid[paid["cdf_row"] - 1][0], "Selected CDF")
        # Formula rows are blank in the grid (data_only) — same as the generic
        # preview — but selected_ldf resolves the `=1` placeholder.
        self.assertIsNone(grid[paid["ldf_row"] - 1][1])
        self.assertEqual(paid["selected_ldf"], [1.0, 1.0, 1.0])

    def test_grid_truncated_when_over_cell_guard(self):
        from django.test import override_settings as _override

        wb = _build_triangle_workbook_with_ldf(n_cols=3)
        job = _make_source_job(
            user=self.user, org=self.org,
            zip_bytes=_zip_with({"Motor TP GROSS 2024-12.xlsx": wb}),
        )
        with _override(MODULE1_OUTPUT_PREVIEW_MAX_CELLS=1):
            payload = read_workbook_cdfs(job, "Motor TP GROSS 2024-12.xlsx")
        paid = next(t for t in payload["triangles"] if t["sheet"] == "Paid Claims Triangle")
        self.assertTrue(paid["grid_truncated"])
        self.assertEqual(paid["grid"], [])
        # The editable rows still work even when the grid is withheld.
        self.assertEqual(paid["selected_ldf"], [1.0, 1.0, 1.0])

    def test_reserve_summary_rows_reflect_ldf_overrides(self):
        """The Select Methods tab's CDFs (→ ultimates) follow the LDFs selected
        in the triangle view."""
        from module1_engine import selected_cdf_from_ldf

        job = self._summary_job()  # paid cdf [1.1, 1.5] → reversed [1.5, 1.1]
        payload = read_reserve_summary_rows(
            job, "Motor TP GROSS 2024-12.xlsx",
            ldf_overrides={"Paid Claims Triangle": [2.0, 1.5]},
        )
        # derived CDF = suffix product = [3.0, 1.5]; engine reverses → [1.5, 3.0]
        self.assertEqual(selected_cdf_from_ldf([2.0, 1.5]), [3.0, 1.5])
        rows = payload["rows"]
        self.assertEqual(rows[0]["paid_cdf"], 1.5)
        self.assertEqual(rows[1]["paid_cdf"], 3.0)
        # Reported side untouched by a Paid-only LDF override.
        self.assertEqual(rows[0]["reported_cdf"], 1.6)

    def test_unreferenced_workbooks_are_copied_verbatim(self):
        wb = _build_reserve_workbook(
            paid_cdf_values=[2.0, 1.5],
            reported_cdf_values=[2.0, 1.5],
        )
        zb = _zip_with({
            "Motor TP GROSS 2024-12.xlsx": wb,
            "Property Fire RI 2024-12.xlsx": wb,
        })
        job = _make_source_job(user=self.user, org=self.org, zip_bytes=zb)

        dest = Path(tempfile.mkdtemp(prefix="reserve-noov-out-"))
        try:
            written = write_workbooks_with_overrides(
                source_job=job,
                overrides_by_filename={
                    # Override one; the other should still be copied so
                    # the engine sees the full batch.
                    "Motor TP GROSS 2024-12.xlsx": {
                        "Paid Claims Triangle": [9.0, 9.0],
                    },
                },
                dest_folder=dest,
            )
            self.assertEqual(
                sorted(written),
                sorted([
                    "Motor TP GROSS 2024-12.xlsx",
                    "Property Fire RI 2024-12.xlsx",
                ]),
            )
            # Original (non-overridden) values intact in the second file.
            wbcheck = load_workbook(
                dest / "Property Fire RI 2024-12.xlsx", data_only=True
            )
            paid_ws = wbcheck["Paid Claims Triangle"]
            sel_row = None
            for row in paid_ws.iter_rows(min_row=1, max_row=paid_ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    if cell.value == "Selected CDF":
                        sel_row = cell.row
                        break
                if sel_row:
                    break
            self.assertEqual(
                [paid_ws.cell(row=sel_row, column=c).value for c in range(2, 4)],
                [2.0, 1.5],
            )
        finally:
            shutil.rmtree(dest, ignore_errors=True)
