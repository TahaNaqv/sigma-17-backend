"""Reserve-workbook helpers for the Excel-free Update Reserve flow.

A Summary job's output ZIP contains:
- `Combined_Summary.xlsx`
- One reserve workbook per (reserving_class, head_of_damage, ri_type, eop),
  named `{reserving_class} {head_of_damage} {ri_type} {YYYY-MM}.xlsx`. Each
  workbook has three sheets: `Paid Claims Triangle`, `Reported Triangle`,
  and `Reserve Summary`. The triangle sheets contain a "Selected LDF" row
  followed by a "Selected CDF" row whose cells are PRODUCT formulas.

The Update Reserve job consumes the **Selected CDF** values (the engine
reads them with `data_only=True`). So the user-facing edit surface in
the UI is the Selected CDF row — we read it, let the user override
values, and write the overrides back as plain numbers (replacing the
PRODUCT formulas). The engine then sees the user's values directly.

This module is the single place that knows how to:
- list reserve workbooks in a source Summary job's ZIP
- read the current Selected CDF row for each triangle sheet
- apply a set of user-provided overrides + write the modified workbooks
  into the task's staging directory
"""

from __future__ import annotations

import io
import logging
import math
import re
import zipfile
from pathlib import Path

from django.conf import settings
from openpyxl import load_workbook

from module1_engine import (
    cdf_for_row,
    normalize_accident_period,
    selected_cdf_from_ldf,
    selected_cdf_row_to_series,
)
from processing.models import Module1Job
from processing.services.source_resolver import (
    ARTIFACT_COMBINED_SUMMARY,
    read_artifact_bytes,
)

logger = logging.getLogger(__name__)


TRIANGLE_SHEET_NAMES = ("Paid Claims Triangle", "Reported Triangle")
SELECTED_CDF_LABEL = "Selected CDF"
SELECTED_LDF_LABEL = "Selected LDF"
RESERVE_SUMMARY_SHEET = "Reserve Summary"
# Reserve Summary base columns (written by the Summary engine), by header name.
# Matching by name (not letter) keeps the reader robust if the sheet already
# carries the appended G–S columns (e.g. an Update-Reserve output re-used as
# a source).
_ACCIDENT_PERIOD_HEADER = "Accident_Period"
_SUMMARY_NUMERIC_HEADERS = {
    "ep": "EP",
    "paid_claims": "Paid Claims",
    "os_claims": "OS Claims",
    "reported_claims": "Reported Claims",
    "reported_lr": "Reported LR",
}


# ---------------------------------------------------------------------------
# Filename parsing
# ---------------------------------------------------------------------------


def _parse_workbook_name(basename: str) -> dict | None:
    """Return {reserving_class, head_of_damage, ri_type, eop_label} or None.

    Mirrors the engine's parsing at module1_engine/engine.py: the last
    three whitespace-separated parts are eop, ri_type, head_of_damage;
    everything before them is the reserving_class (which may itself
    contain spaces).
    """
    if not basename.lower().endswith(".xlsx"):
        return None
    stem = basename[:-len(".xlsx")]
    parts = re.split(r"\s+", stem.strip())
    if len(parts) < 4:
        return None
    eop_label = parts[-1]
    ri_type = parts[-2].upper()
    head_of_damage = parts[-3]
    reserving_class = " ".join(parts[:-3])
    return {
        "filename": basename,
        "reserving_class": reserving_class,
        "head_of_damage": head_of_damage,
        "ri_type": ri_type,
        "eop_label": eop_label,
    }


def list_reserve_workbooks(source_job: Module1Job) -> list[dict]:
    """List the parseable reserve workbooks in `source_job.output_zip`.

    Excludes `Combined_Summary.xlsx` and any entries whose names don't
    match the engine's `{class} {hod} {ri} {eop}.xlsx` shape.
    """
    if source_job is None:
        raise ValueError("Source job is missing.")
    if not source_job.output_zip:
        return []
    try:
        with source_job.output_zip.open("rb") as f:
            raw = f.read()
    except FileNotFoundError as exc:
        raise ValueError("Source job output is missing on disk.") from exc

    out: list[dict] = []
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        for name in zf.namelist():
            if name == ARTIFACT_COMBINED_SUMMARY:
                continue
            if not name.lower().endswith(".xlsx"):
                continue
            parsed = _parse_workbook_name(name)
            if parsed is not None:
                out.append(parsed)
    return out


# ---------------------------------------------------------------------------
# CDF reading
# ---------------------------------------------------------------------------


def _find_row_by_label(ws, label: str) -> int | None:
    """First row whose column-1 cell equals `label` (same primitive the engine
    uses to locate the Selected LDF / Selected CDF rows)."""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value == label:
                return cell.row
    return None


def _find_selected_cdf_row(ws) -> int | None:
    return _find_row_by_label(ws, SELECTED_CDF_LABEL)


def _json_cell(v):
    """Coerce an openpyxl cell value to something JSON-safe for the grid.
    Blank/uncached-formula cells are None (that's why the Selected LDF/CDF rows
    render empty in the generic preview too)."""
    if v is None or isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        # pandas-written gaps can surface as NaN/inf — not JSON-safe.
        return v if math.isfinite(v) else None
    return str(v)


def _read_grid(ws_data, max_row: int, max_col: int) -> list[list]:
    """The whole sheet as a value grid (row 1..max_row × col 1..max_col), read
    with `data_only` — identical to what the generic output preview shows, so the
    triangle view is Excel-faithful (history + Simple/Weighted Avg benchmarks
    included) without re-deriving or re-interpreting any block."""
    return [
        [_json_cell(ws_data.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        for r in range(1, max_row + 1)
    ]


def _cell_number(ws_data, ws_formulas, row: int, col: int):
    """Read a numeric cell, preferring the cached value; fall back to a constant
    formula (the engine seeds Selected LDF cells as the placeholder `=1`, which
    has no cached value — surface it as 1.0 rather than blank)."""
    v = ws_data.cell(row=row, column=col).value
    if v is None:
        f = ws_formulas.cell(row=row, column=col).value
        if isinstance(f, str) and f.startswith("="):
            try:
                return float(f[1:])
            except ValueError:
                return None
        v = f
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _read_triangle_cdf(wb_data_only, wb_formulas, sheet_name: str) -> dict:
    """Pull the column labels (row 1) plus the current Selected LDF and Selected
    CDF rows for one triangle sheet.

    Returns:
      {
        "sheet": sheet_name,
        "cdf_row": <int|null>, "ldf_row": <int|null>,   # 1-based sheet rows
        "column_labels": [str, ...],          # headers above each value
        "values": [float|None, ...],          # current Selected CDF values
        "selected_ldf": [float|None, ...],    # current Selected LDF values (=1 → 1.0)
        "grid": [[cell, ...], ...],           # whole sheet, 1..max_row × 1..max_col
        "grid_truncated": bool,               # True when the cell guard tripped
      }

    The `grid` carries the LDF *history* the actuary selects against — the
    age-to-age factor block plus the Simple/Weighted Avg LDF & CDF benchmarks —
    so the UI can show the selection in its triangle context. `ldf_row`/`cdf_row`
    index into it (1-based), letting the client overlay the editable Selected LDF
    row and the derived Selected CDF row in place.
    """
    empty = {
        "sheet": sheet_name, "cdf_row": None, "ldf_row": None,
        "column_labels": [], "values": [], "selected_ldf": [],
        "grid": [], "grid_truncated": False,
    }
    if sheet_name not in wb_formulas.sheetnames:
        return empty
    ws = wb_formulas[sheet_name]
    ws_data = wb_data_only[sheet_name]
    cdf_row = _find_selected_cdf_row(ws)
    ldf_row = _find_row_by_label(ws, SELECTED_LDF_LABEL)
    if cdf_row is None:
        return empty

    # The engine iterates columns 2..max_column inclusive (column 1 is the label).
    max_col = ws.max_column
    max_row = ws.max_row
    column_labels: list[str] = []
    values: list[float | None] = []
    selected_ldf: list[float | None] = []
    for col_idx in range(2, max_col + 1):
        header_value = ws.cell(row=1, column=col_idx).value
        column_labels.append("" if header_value is None else str(header_value))
        values.append(_cell_number(ws_data, ws, cdf_row, col_idx))
        selected_ldf.append(
            _cell_number(ws_data, ws, ldf_row, col_idx) if ldf_row else None
        )

    max_cells = getattr(settings, "MODULE1_OUTPUT_PREVIEW_MAX_CELLS", 20000)
    truncated = (max_row * max_col) > max_cells
    grid = [] if truncated else _read_grid(ws_data, max_row, max_col)

    return {
        "sheet": sheet_name,
        "cdf_row": cdf_row,
        "ldf_row": ldf_row,
        "column_labels": column_labels,
        "values": values,
        "selected_ldf": selected_ldf,
        "grid": grid,
        "grid_truncated": truncated,
    }


def read_workbook_cdfs(source_job: Module1Job, filename: str) -> dict:
    """Read the Selected CDF row for both triangle sheets of one workbook.

    Returns:
      {
        "filename": str,
        "reserving_class", "head_of_damage", "ri_type", "eop_label": ...,
        "triangles": [ {sheet, cdf_row, column_labels, values}, ... ],
      }
    Raises ValueError if the workbook isn't found or can't be parsed.
    """
    parsed = _parse_workbook_name(filename)
    if parsed is None:
        raise ValueError(f"Workbook name '{filename}' is not parseable.")

    raw = read_artifact_bytes(source_job=source_job, artifact=filename)

    # Two loads: one for formulas (sheet structure / cell label lookup),
    # one for values (the cached formula results the engine consumes).
    wb_formulas = load_workbook(io.BytesIO(raw), data_only=False)
    wb_data_only = load_workbook(io.BytesIO(raw), data_only=True)

    triangles = [
        _read_triangle_cdf(wb_data_only, wb_formulas, name)
        for name in TRIANGLE_SHEET_NAMES
    ]

    return {**parsed, "triangles": triangles}


# ---------------------------------------------------------------------------
# Reserve Summary rows — powers the Implied LR / Selected Method editor
# ---------------------------------------------------------------------------


def _num(value) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _selected_cdf_series(wb_data_only, wb_formulas, sheet_name: str) -> list[float]:
    """Per-accident-period CDF series for one triangle sheet, derived exactly as
    the engine does (blank→2.0 + reverse). Returns [] when the sheet or its
    Selected CDF row is absent."""
    if sheet_name not in wb_formulas.sheetnames:
        return []
    ws = wb_formulas[sheet_name]
    ws_data = wb_data_only[sheet_name]
    cdf_row = _find_selected_cdf_row(ws)
    if cdf_row is None:
        return []
    raw = [
        ws_data.cell(row=cdf_row, column=col).value
        for col in range(2, ws.max_column + 1)
    ]
    return selected_cdf_row_to_series(raw)


def read_reserve_summary_rows(
    source_job: Module1Job,
    filename: str,
    *,
    cdf_overrides: dict[str, list[float | None]] | None = None,
    ldf_overrides: dict[str, list[float | None]] | None = None,
) -> dict:
    """Per-accident-period Reserve Summary rows for the Method-Selection UI.

    Returns the base columns (EP, Paid, OS, Reported, Reported LR) plus the
    Paid/Reported CDF the engine will assign each row — so the client can
    preview the five method ultimates and choose Implied LR / Selected Method
    against exactly what the Update Reserve job will compute.

    `ldf_overrides` (per-sheet Selected-LDF arrays for THIS workbook) and
    `cdf_overrides` (direct Selected-CDF arrays) are applied first — the same
    write path the job uses — so the previewed Paid/Reported CDF, and therefore
    the five ultimates, reflect the LDFs the actuary just selected in the
    triangle view. LDF wins over CDF per sheet.

    Row order matches the engine's `enumerate(...)` over Reserve Summary rows,
    so `accident_period` is a safe join key back into `method_overrides`.
    """
    parsed = _parse_workbook_name(filename)
    if parsed is None:
        raise ValueError(f"Workbook name '{filename}' is not parseable.")

    raw = read_artifact_bytes(source_job=source_job, artifact=filename)
    if cdf_overrides or ldf_overrides:
        raw = _apply_overrides_to_bytes(
            raw, cdf_overrides=cdf_overrides, ldf_overrides=ldf_overrides
        )

    wb_formulas = load_workbook(io.BytesIO(raw), data_only=False)
    wb_data = load_workbook(io.BytesIO(raw), data_only=True)
    if RESERVE_SUMMARY_SHEET not in wb_formulas.sheetnames:
        raise ValueError(f"'{RESERVE_SUMMARY_SHEET}' sheet not found in {filename}.")

    paid_series = _selected_cdf_series(wb_data, wb_formulas, "Paid Claims Triangle")
    reported_series = _selected_cdf_series(wb_data, wb_formulas, "Reported Triangle")

    ws = wb_data[RESERVE_SUMMARY_SHEET]
    header_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h is not None and str(h) not in header_col:
            header_col[str(h)] = col
    if _ACCIDENT_PERIOD_HEADER not in header_col:
        raise ValueError(f"'{_ACCIDENT_PERIOD_HEADER}' column not found in {filename}.")

    ap_col = header_col[_ACCIDENT_PERIOD_HEADER]
    rows: list[dict] = []
    # enumerate identically to the engine so `idx` (and thus the CDF a row gets)
    # aligns row-for-row with what the Update Reserve job will write.
    for idx, excel_row in enumerate(range(2, ws.max_row + 1)):
        ap = ws.cell(row=excel_row, column=ap_col).value
        row = {"accident_period": normalize_accident_period(ap)}
        for key, header in _SUMMARY_NUMERIC_HEADERS.items():
            col = header_col.get(header)
            row[key] = _num(ws.cell(row=excel_row, column=col).value) if col else 0.0
        row["paid_cdf"] = cdf_for_row(paid_series, idx)
        row["reported_cdf"] = cdf_for_row(reported_series, idx)
        rows.append(row)

    return {**parsed, "rows": rows}


# ---------------------------------------------------------------------------
# Override application
# ---------------------------------------------------------------------------


def _apply_overrides_to_bytes(
    raw_bytes: bytes,
    cdf_overrides: dict[str, list[float | None]] | None = None,
    ldf_overrides: dict[str, list[float | None]] | None = None,
) -> bytes:
    """Apply per-sheet overrides to a triangle workbook and return the bytes.

    - `ldf_overrides` (primary): write the Selected LDF row literals AND the
      derived Selected CDF row literals (= `selected_cdf_from_ldf`, the Excel
      `=PRODUCT(...)` suffix product). This is the Excel-faithful path — the
      actuary selects LDFs and the CDF derives.
    - `cdf_overrides` (legacy/direct): write the Selected CDF row literals for
      sheets NOT already covered by an LDF override.

    Both are keyed by sheet name; values are positional from column 2. Missing
    sheets / shorter arrays leave the untouched cells as they were. Writing a
    numeric value drops any formula in that cell (the engine reads the literal).
    """
    cdf_overrides = cdf_overrides or {}
    ldf_overrides = ldf_overrides or {}
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=False)

    for sheet_name, ldf_values in ldf_overrides.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ldf_row = _find_row_by_label(ws, SELECTED_LDF_LABEL)
        cdf_row = _find_selected_cdf_row(ws)
        if ldf_row is not None:
            for offset, value in enumerate(ldf_values):
                ws.cell(row=ldf_row, column=2 + offset, value=value)
        if cdf_row is not None:
            for offset, value in enumerate(selected_cdf_from_ldf(ldf_values)):
                ws.cell(row=cdf_row, column=2 + offset, value=value)

    for sheet_name, new_values in cdf_overrides.items():
        if sheet_name in ldf_overrides:
            continue  # LDF override already set this sheet's CDF
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cdf_row = _find_selected_cdf_row(ws)
        if cdf_row is None:
            continue
        for offset, value in enumerate(new_values):
            ws.cell(row=cdf_row, column=2 + offset, value=value)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def write_workbooks_with_overrides(
    *,
    source_job: Module1Job,
    overrides_by_filename: dict[str, dict[str, list[float | None]]],
    dest_folder: Path,
    ldf_overrides_by_filename: dict[str, dict[str, list[float | None]]] | None = None,
) -> list[str]:
    """For each reserve workbook in `source_job`'s output ZIP, copy it into
    `dest_folder`, applying any per-workbook LDF and/or CDF overrides first.

    `overrides_by_filename` = direct Selected-CDF overrides (legacy).
    `ldf_overrides_by_filename` = Selected-LDF overrides (writes LDF + derived
    CDF). Returns the basenames written. Workbooks with no override are copied
    verbatim so the engine still sees the full batch.
    """
    if source_job is None:
        raise ValueError("Source job is missing.")
    if not source_job.output_zip:
        raise ValueError("Source job has no downloadable output.")

    ldf_overrides_by_filename = ldf_overrides_by_filename or {}
    dest_folder.mkdir(parents=True, exist_ok=True)

    try:
        with source_job.output_zip.open("rb") as f:
            raw_zip = f.read()
    except FileNotFoundError as exc:
        raise ValueError("Source job output is missing on disk.") from exc

    written: list[str] = []
    with zipfile.ZipFile(io.BytesIO(raw_zip), "r") as zf:
        for name in zf.namelist():
            if name == ARTIFACT_COMBINED_SUMMARY:
                # The Update Reserve task handles Combined_Summary separately
                # via the existing chained-source code path; skip it here.
                continue
            if not name.lower().endswith(".xlsx"):
                continue
            if _parse_workbook_name(name) is None:
                continue
            file_bytes = zf.read(name)
            cdf_ov = overrides_by_filename.get(name)
            ldf_ov = ldf_overrides_by_filename.get(name)
            if cdf_ov or ldf_ov:
                file_bytes = _apply_overrides_to_bytes(
                    file_bytes, cdf_overrides=cdf_ov, ldf_overrides=ldf_ov
                )
            (dest_folder / name).write_bytes(file_bytes)
            written.append(name)
    return written
