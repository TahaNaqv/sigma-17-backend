"""Tests for the Excel import service + endpoint.

Covers the round trip: an .xlsx with the engine's mixed-case headers →
Dataset + typed snake_case row records.
"""

import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from openpyxl import Workbook
from rest_framework.test import APIClient

from accounts.models import Permission, Role
from datasets.models import Dataset, PremiumRow
from datasets.services.excel_import import (
    ExcelImportError,
    import_excel_to_dataset,
)
from tenants.models import Membership, Organization

User = get_user_model()


def _make_org(name: str) -> Organization:
    return Organization.objects.create(name=name)


def _make_user(*, username: str, perm_keys: list[str], org: Organization) -> User:
    user = User.objects.create_user(username=username, password="testpass123")
    user.profile.active_organization = org
    user.profile.save(update_fields=["active_organization"])
    role, _ = Role.objects.get_or_create(name=f"role-{username}")
    for key in perm_keys:
        perm, _ = Permission.objects.get_or_create(
            key=key,
            defaults={"name": key, "module": "Datasets", "description": ""},
        )
        role.permissions.add(perm)
    m = Membership.objects.create(user=user, organization=org, status="active")
    m.roles.add(role)
    return user


def _premium_xlsx_bytes(rows: list[dict] | None = None) -> bytes:
    """Build an xlsx workbook with the engine's mixed-case Premium headers."""
    wb = Workbook()
    ws = wb.active
    headers = [
        "POLICYNUMBER",
        "POLICYSTARTDATE",
        "POLICYENDDATE",
        "RiskStartDate",
        "RiskEndDate",
        "ISSUEDATE",
        "RESERVINGCLASS",
        "POLICYCLASS",
        "PRODUCTTYPE",
        "RI_TREATY_TYPE",
        "PREMIUMAMOUNT",
        "COMMISSIONAMOUNT",
    ]
    ws.append(headers)
    if rows is None:
        rows = [
            {
                "POLICYNUMBER": "POL-1",
                "POLICYSTARTDATE": "2024-01-01",
                "POLICYENDDATE": "2024-12-31",
                "RiskStartDate": "2024-01-01",
                "RiskEndDate": "2024-12-31",
                "ISSUEDATE": "2023-12-15",
                "RESERVINGCLASS": "Motor",
                "POLICYCLASS": "Motor",
                "PRODUCTTYPE": "Comprehensive",
                "RI_TREATY_TYPE": "GROSS",
                "PREMIUMAMOUNT": 1000.50,
                "COMMISSIONAMOUNT": 100.00,
            },
            {
                "POLICYNUMBER": "POL-2",
                "POLICYSTARTDATE": "2024-02-01",
                "POLICYENDDATE": "2025-01-31",
                "RiskStartDate": "2024-02-01",
                "RiskEndDate": "2025-01-31",
                "ISSUEDATE": "2024-01-15",
                "RESERVINGCLASS": "Property",
                "POLICYCLASS": "Property",
                "PRODUCTTYPE": "Fire",
                "RI_TREATY_TYPE": "RI",
                "PREMIUMAMOUNT": 2500,
                "COMMISSIONAMOUNT": 250,
            },
        ]
    for row in rows:
        ws.append([row.get(h) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@override_settings(SECURE_SSL_REDIRECT=False)
class ExcelImportServiceTests(TestCase):
    """Direct calls to `import_excel_to_dataset` (no HTTP layer)."""

    def setUp(self):
        self.org = _make_org("Imp Org")
        self.user = _make_user(
            username="importer",
            perm_keys=["datasets.view", "datasets.edit"],
            org=self.org,
        )

    def test_happy_path_creates_rows_with_snake_case_columns(self):
        ds = import_excel_to_dataset(
            organization=self.org,
            kind=Dataset.Kind.PREMIUM,
            name="2024 Q1 Premium",
            file_bytes=_premium_xlsx_bytes(),
            created_by=self.user,
        )
        self.assertEqual(ds.source, Dataset.Source.EXCEL_IMPORT)
        self.assertEqual(ds.row_count, 2)
        rows = list(PremiumRow.objects.filter(dataset=ds).order_by("row_index"))
        self.assertEqual(rows[0].policy_number, "POL-1")
        self.assertEqual(rows[0].reserving_class, "Motor")
        self.assertEqual(rows[0].ri_treaty_type, "GROSS")
        self.assertEqual(rows[0].premium_amount, Decimal("1000.50"))
        self.assertEqual(rows[1].ri_treaty_type, "RI")
        # Date coercion
        self.assertEqual(str(rows[0].policy_start_date), "2024-01-01")

    def test_unknown_columns_are_ignored(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["RESERVINGCLASS", "RI_TREATY_TYPE", "RANDOM_EXTRA"])
        ws.append(["Motor", "GROSS", "ignored"])
        buf = io.BytesIO()
        wb.save(buf)

        ds = import_excel_to_dataset(
            organization=self.org,
            kind=Dataset.Kind.PREMIUM,
            name="Sparse",
            file_bytes=buf.getvalue(),
            created_by=self.user,
        )
        self.assertEqual(ds.row_count, 1)
        row = PremiumRow.objects.get(dataset=ds)
        self.assertEqual(row.reserving_class, "Motor")
        # No exception on the unknown column.

    def test_missing_required_column_raises(self):
        # Build a sheet without RESERVINGCLASS
        wb = Workbook()
        ws = wb.active
        ws.append(["RI_TREATY_TYPE"])
        ws.append(["GROSS"])
        buf = io.BytesIO()
        wb.save(buf)

        with self.assertRaises(ExcelImportError) as ctx:
            import_excel_to_dataset(
                organization=self.org,
                kind=Dataset.Kind.PREMIUM,
                name="Bad",
                file_bytes=buf.getvalue(),
                created_by=self.user,
            )
        self.assertIn("RESERVINGCLASS", str(ctx.exception))

    def test_row_validation_failure_rolls_back_dataset(self):
        # ri_treaty_type=INVALID is not in the choices ["GROSS","RI"]
        bad_rows = [
            {
                "RESERVINGCLASS": "Motor",
                "RI_TREATY_TYPE": "NOT_A_CHOICE",
            }
        ]
        wb = Workbook()
        ws = wb.active
        ws.append(["RESERVINGCLASS", "RI_TREATY_TYPE"])
        for r in bad_rows:
            ws.append([r["RESERVINGCLASS"], r["RI_TREATY_TYPE"]])
        buf = io.BytesIO()
        wb.save(buf)

        with self.assertRaises(ExcelImportError):
            import_excel_to_dataset(
                organization=self.org,
                kind=Dataset.Kind.PREMIUM,
                name="Bad rows",
                file_bytes=buf.getvalue(),
                created_by=self.user,
            )
        # Atomicity: no dataset row should exist after a failed import.
        self.assertFalse(Dataset.objects.filter(name="Bad rows").exists())


@override_settings(SECURE_SSL_REDIRECT=False)
class ExcelImportEndpointTests(TestCase):
    """POST /api/datasets/import-excel/ end-to-end."""

    def setUp(self):
        self.org = _make_org("End Org")
        self.user = _make_user(
            username="ender",
            perm_keys=["datasets.view", "datasets.edit"],
            org=self.org,
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def test_endpoint_creates_dataset_from_uploaded_xlsx(self):
        xlsx = _premium_xlsx_bytes()
        upload = io.BytesIO(xlsx)
        upload.name = "premium.xlsx"
        resp = self.client.post(
            "/api/datasets/import-excel/",
            data={
                "kind": "premium",
                "name": "Imported Q1",
                "file": upload,
            },
            format="multipart",
        )
        self.assertEqual(resp.status_code, 201, resp.data)
        self.assertEqual(resp.data["source"], "excel_import")
        self.assertEqual(resp.data["row_count"], 2)
        self.assertEqual(resp.data["kind"], "premium")

    def test_endpoint_rejects_non_xlsx(self):
        upload = io.BytesIO(b"not really xlsx")
        upload.name = "premium.csv"
        resp = self.client.post(
            "/api/datasets/import-excel/",
            data={"kind": "premium", "name": "Bad ext", "file": upload},
            format="multipart",
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("file", resp.data)

    def test_endpoint_rejects_missing_required_column(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["RANDOM_COLUMN"])
        ws.append(["x"])
        buf = io.BytesIO()
        wb.save(buf)
        upload = io.BytesIO(buf.getvalue())
        upload.name = "bad.xlsx"
        resp = self.client.post(
            "/api/datasets/import-excel/",
            data={"kind": "premium", "name": "Bad cols", "file": upload},
            format="multipart",
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("detail", resp.data)

    def test_endpoint_requires_edit_permission(self):
        # Create a viewer in the same org and try to import
        viewer = _make_user(
            username="viewer",
            perm_keys=["datasets.view"],
            org=self.org,
        )
        c = APIClient()
        c.force_authenticate(user=viewer)
        upload = io.BytesIO(_premium_xlsx_bytes())
        upload.name = "premium.xlsx"
        resp = c.post(
            "/api/datasets/import-excel/",
            data={"kind": "premium", "name": "Should fail", "file": upload},
            format="multipart",
        )
        self.assertEqual(resp.status_code, 403)
