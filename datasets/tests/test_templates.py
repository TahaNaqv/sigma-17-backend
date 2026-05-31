"""Tests for the .xlsx upload-template generator.

The critical guarantee: a generated template's headers/sheet-names are
byte-identical to what the import service / engine expect, so a template
can never drift from the parser. We assert that directly, then round-trip
the flat-table templates back through the importer.
"""

from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook

from datasets.models import Dataset
from datasets.services.columns import DB_TO_EXCEL_FOR_KIND
from datasets.services.engine_adapter import KIND_RECIPE
from datasets.services.excel_import import import_excel_to_dataset
from datasets.services.templates import TEMPLATES, render_template
from tenants.models import Organization


def _load(key: str):
    return load_workbook(BytesIO(render_template(key)))


class TemplateRenderTests(TestCase):
    def test_every_template_has_instructions_last_and_data_first(self):
        # Importer/engine read the FIRST sheet, so data must be first and
        # Instructions must NOT be first.
        for key in TEMPLATES:
            wb = _load(key)
            self.assertIn("Instructions", wb.sheetnames, f"{key} missing Instructions")
            self.assertNotEqual(
                wb.sheetnames[0], "Instructions", f"{key} has Instructions first"
            )

    def test_dataset_kind_headers_match_source_of_truth(self):
        """Drift guard — header row must equal the parser's expected headers."""
        for key, tpl in TEMPLATES.items():
            if tpl.reserve:
                continue
            wb = _load(key)
            for kind in tpl.kinds:
                sheet_name = KIND_RECIPE[kind]["sheet_name"]
                ws = wb[sheet_name]
                header_row = [c.value for c in ws[1]]
                expected = list(DB_TO_EXCEL_FOR_KIND[kind].values())
                self.assertEqual(
                    header_row, expected, f"{key}/{sheet_name} header drift"
                )

    def test_expense_cf_sheet_is_named_expense_cf(self):
        wb = _load(Dataset.Kind.EXPENSE_CF)
        self.assertIn("Expense-CF", wb.sheetnames)

    def test_previous_period_composite_has_both_sheets(self):
        wb = _load("previous_period")
        self.assertIn("LIC_BOP", wb.sheetnames)
        self.assertIn("UPR-DAC_BOP", wb.sheetnames)

    def test_reserve_has_both_triangle_sheets(self):
        wb = _load("reserve")
        self.assertIn("Paid Claims Triangle", wb.sheetnames)
        self.assertIn("Reported Triangle", wb.sheetnames)
        # Triangle skeleton: first header cell + a Selected CDF row label.
        ws = wb["Paid Claims Triangle"]
        self.assertEqual(ws["A1"].value, "Accident Period")
        labels = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]
        self.assertIn("Selected CDF", labels)


class TemplateRoundTripTests(TestCase):
    """An empty (header-only) flat template must import cleanly: column
    mapping validates, zero rows ingested, no 'missing column' error."""

    FLAT_KINDS = (
        Dataset.Kind.PREMIUM,
        Dataset.Kind.CLAIMS_PAID,
        Dataset.Kind.CLAIMS_OS,
        Dataset.Kind.EXPENSE_CF,
        Dataset.Kind.PREVIOUS_PERIOD_LIC,
        Dataset.Kind.PREVIOUS_PERIOD_UPR,
    )

    def setUp(self):
        self.org = Organization.objects.create(name="acme")

    def test_blank_templates_import_with_zero_rows(self):
        for kind in self.FLAT_KINDS:
            dataset = import_excel_to_dataset(
                organization=self.org,
                kind=kind,
                name=f"tpl-{kind}",
                file_bytes=render_template(kind),
            )
            self.assertEqual(dataset.kind, kind)
            self.assertEqual(dataset.row_count, 0, f"{kind} should import 0 rows")
