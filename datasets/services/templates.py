"""Generate empty, correctly-structured .xlsx upload templates.

Users discover the exact shape an upload must have by downloading a
template: the right sheet name(s), the right headers (byte-identical to
what the importer/engine read), data-validation dropdowns for enum
columns, sensible number/date formats, and a leading Instructions sheet.

Single source of truth — this module *reuses* the same maps the import
service and engine adapter use, so a template can never drift from what
the parser accepts:
  - headers / db→excel:  columns.DB_TO_EXCEL_FOR_KIND
  - required fields:      columns.REQUIRED_FIELDS_FOR_KIND
  - sheet names per kind: engine_adapter.KIND_RECIPE

The only genuinely new metadata is per-column type/enum/description
(COLUMN_META below) used for dropdowns, formatting and documentation.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..models import Dataset
from .columns import DB_TO_EXCEL_FOR_KIND, REQUIRED_FIELDS_FOR_KIND
from .engine_adapter import KIND_RECIPE


# How many blank rows get pre-formatted (dropdowns / number formats) below
# the header. Generous enough for real uploads without bloating the file.
TEMPLATE_DATA_ROWS = 200


# --------------------------------------------------------------------------
# Per-column metadata, keyed by DB field name (same keys as the *_DB_TO_EXCEL
# maps in columns.py). `dtype` drives number format + dropdowns; `description`
# feeds the Instructions sheet. Required-ness is NOT stored here — it is
# per-kind and read from REQUIRED_FIELDS_FOR_KIND so the two never disagree.
# --------------------------------------------------------------------------
@dataclass(frozen=True)
class ColumnSpec:
    dtype: str  # one of: text | date | decimal | int | enum
    description: str
    enum_values: tuple[str, ...] = ()


_GROSS_RI = ("GROSS", "RI")

COLUMN_META: dict[str, ColumnSpec] = {
    # shared
    "reserving_class": ColumnSpec("text", "Reserving class categorisation"),
    "policy_class": ColumnSpec("text", "Policy class categorisation"),
    "product_type": ColumnSpec("text", "Type of insurance product"),
    "ri_treaty_type": ColumnSpec("enum", "Reinsurance treaty type", _GROSS_RI),
    "issue_date": ColumnSpec("date", "Date the policy was issued"),
    "uwy": ColumnSpec("int", "Underwriting year (e.g. 2024)"),
    "head_of_damage": ColumnSpec("text", "Category / head of damage"),
    # premium
    "policy_number": ColumnSpec("text", "Unique identifier for the policy"),
    "policy_start_date": ColumnSpec("date", "Policy inception date"),
    "policy_end_date": ColumnSpec("date", "Policy expiry date"),
    "risk_start_date": ColumnSpec("date", "Start date of risk coverage"),
    "risk_end_date": ColumnSpec("date", "End date of risk coverage"),
    "premium_amount": ColumnSpec("decimal", "Total premium amount"),
    "commission_amount": ColumnSpec("decimal", "Commission amount paid"),
    # claims paid
    "amount_paid": ColumnSpec("decimal", "Amount paid on the claim"),
    "amount_recovered": ColumnSpec("decimal", "Amount recovered on the claim"),
    "loss_date": ColumnSpec("date", "Date of the loss event"),
    "payment_date": ColumnSpec("date", "Date the payment was made"),
    # claims os
    "amount_outstanding": ColumnSpec("decimal", "Outstanding claim amount"),
    "as_at": ColumnSpec("date", "Valuation / reporting date"),
    # expense cf
    "comm_payable_prev": ColumnSpec("decimal", "Commission payable — prior period"),
    "comm_payable_curr": ColumnSpec("decimal", "Commission payable — current period"),
    "rec_gop_prev": ColumnSpec("decimal", "Recoverable GOP — prior period"),
    "rec_gop_curr": ColumnSpec("decimal", "Recoverable GOP — current period"),
    "rec_provision_prev": ColumnSpec("decimal", "Recoverable provision — prior period"),
    "rec_provision_curr": ColumnSpec("decimal", "Recoverable provision — current period"),
    "premium_received": ColumnSpec("decimal", "Premium received (cash flow)"),
    "claims_paid": ColumnSpec("decimal", "Claims paid (cash flow)"),
    "insurance_acquisition_cash_flows": ColumnSpec("decimal", "Insurance acquisition cash flows"),
    "other_cash_flows": ColumnSpec("decimal", "Other cash flows"),
    "ri_premium_paid": ColumnSpec("decimal", "RI premium paid (cash flow)"),
    "ri_claims_received": ColumnSpec("decimal", "RI claims received (cash flow)"),
    "ri_fixed_commission_received": ColumnSpec("decimal", "RI fixed commission received"),
    "directly_attributable_expenses": ColumnSpec("decimal", "Directly attributable expenses, excl. acquisition cash flows"),
    "other_acquisition_cash_flows": ColumnSpec("decimal", "Other acquisition cash flows"),
    "ri_rec_gop_prev": ColumnSpec("decimal", "RI recoverable GOP — prior period"),
    "ri_rec_gop_curr": ColumnSpec("decimal", "RI recoverable GOP — current period"),
    "claim_pay_prev": ColumnSpec("decimal", "Claims payable (pipeline) — prior period"),
    "claim_pay_curr": ColumnSpec("decimal", "Claims payable (pipeline) — current period"),
    "ri_payable_prev": ColumnSpec("decimal", "RI payable — prior period"),
    "ri_payable_curr": ColumnSpec("decimal", "RI payable — current period"),
    "ri_rec_provision_prev": ColumnSpec("decimal", "RI recoverable provision — prior period"),
    "ri_rec_provision_curr": ColumnSpec("decimal", "RI recoverable provision — current period"),
    # previous period LIC
    "accident_period": ColumnSpec("text", "Accident period, e.g. 2023-Q1"),
    "gross_ri": ColumnSpec("enum", "Gross or reinsurance", _GROSS_RI),
    "outstanding": ColumnSpec("decimal", "Outstanding amount"),
    "ss": ColumnSpec("decimal", "Salvage & subrogation (SS)"),
    "payment": ColumnSpec("decimal", "Payment amount"),
    "s_and_s": ColumnSpec("decimal", "S&S amount"),
    "ulae": ColumnSpec("decimal", "Unallocated loss adjustment expense"),
    "ra_os": ColumnSpec("decimal", "Risk adjustment on outstanding"),
    "ra_ibnr": ColumnSpec("decimal", "Risk adjustment on IBNR"),
    "discounting_impact": ColumnSpec("decimal", "Discounting impact"),
    # previous period UPR
    "gross_upr": ColumnSpec("decimal", "Gross unearned premium reserve"),
    "dac": ColumnSpec("decimal", "Deferred acquisition cost"),
    "ri_upr": ColumnSpec("decimal", "Reinsurance unearned premium reserve"),
    "ucr": ColumnSpec("decimal", "Unexpired claims reserve"),
}

_NUMBER_FORMAT = {
    "decimal": "#,##0.00",
    "int": "0",
    "date": "yyyy-mm-dd",
}


# --------------------------------------------------------------------------
# Template registry
# --------------------------------------------------------------------------
@dataclass(frozen=True)
class TemplateDef:
    filename: str
    title: str
    kinds: tuple[str, ...] = ()  # dataset kinds, one sheet each (sheet name via KIND_RECIPE)
    reserve: bool = False  # special triangle-skeleton builder instead of `kinds`


_K = Dataset.Kind

TEMPLATES: dict[str, TemplateDef] = {
    _K.PREMIUM: TemplateDef("premium_template.xlsx", "Premium Data", (_K.PREMIUM,)),
    _K.CLAIMS_PAID: TemplateDef("claims_paid_template.xlsx", "Claims Paid Data", (_K.CLAIMS_PAID,)),
    _K.CLAIMS_OS: TemplateDef("claims_os_template.xlsx", "Claims Outstanding Data", (_K.CLAIMS_OS,)),
    _K.EXPENSE_CF: TemplateDef("expense_cf_template.xlsx", "Expense Cash Flow", (_K.EXPENSE_CF,)),
    _K.PREVIOUS_PERIOD_LIC: TemplateDef(
        "previous_period_lic_template.xlsx", "Previous Period — LIC_BOP", (_K.PREVIOUS_PERIOD_LIC,)
    ),
    _K.PREVIOUS_PERIOD_UPR: TemplateDef(
        "previous_period_upr_template.xlsx", "Previous Period — UPR-DAC_BOP", (_K.PREVIOUS_PERIOD_UPR,)
    ),
    # Composite workbook for the Module 2 Process `previous_period` upload —
    # both sheets in one file.
    "previous_period": TemplateDef(
        "previous_period_template.xlsx",
        "Previous Period (LIC_BOP + UPR-DAC_BOP)",
        (_K.PREVIOUS_PERIOD_LIC, _K.PREVIOUS_PERIOD_UPR),
    ),
    # Reserve triangle skeleton for the Update Reserve upload.
    "reserve": TemplateDef("reserve_template.xlsx", "Reserve Triangles", reserve=True),
}


# --------------------------------------------------------------------------
# Styling helpers
# --------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # dark blue (optional cols)
_REQUIRED_FILL = PatternFill("solid", fgColor="C00000")  # red (required cols)
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14)
_THIN = Side(style="thin", color="D9D9D9")
_HEADER_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header_cell(cell, *, required: bool) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _REQUIRED_FILL if required else _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _HEADER_BORDER


def _build_kind_sheet(ws, kind: str) -> None:
    """Write one data sheet for a dataset kind: styled header row plus
    pre-formatted blank data rows (dropdowns + number formats)."""
    db_to_excel = DB_TO_EXCEL_FOR_KIND[kind]
    required = set(REQUIRED_FIELDS_FOR_KIND.get(kind, ()))

    for col_idx, (db_field, header) in enumerate(db_to_excel.items(), start=1):
        spec = COLUMN_META.get(db_field, ColumnSpec("text", ""))
        is_required = db_field in required
        letter = get_column_letter(col_idx)

        # Header — text stays byte-identical to the parser's expected header.
        cell = ws.cell(row=1, column=col_idx, value=header)
        _style_header_cell(cell, required=is_required)
        note = "Required" if is_required else "Optional"
        if spec.description:
            note = f"{note} — {spec.description}"
        cell.comment = Comment(note, "Sigma 17")

        ws.column_dimensions[letter].width = max(14, min(len(header) + 4, 40))

        # Pre-format the blank data rows.
        rng = f"{letter}2:{letter}{1 + TEMPLATE_DATA_ROWS}"
        if spec.dtype == "enum" and spec.enum_values:
            dv = DataValidation(
                type="list",
                formula1='"%s"' % ",".join(spec.enum_values),
                allow_blank=True,
                showDropDown=False,
            )
            ws.add_data_validation(dv)
            dv.add(rng)
        fmt = _NUMBER_FORMAT.get(spec.dtype)
        if fmt:
            for r in range(2, 2 + TEMPLATE_DATA_ROWS):
                ws.cell(row=r, column=col_idx).number_format = fmt

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(db_to_excel))
    ws.auto_filter.ref = f"A1:{last_col}1"


def _build_reserve_sheet(ws, n_dev: int = 12, n_examples: int = 8) -> None:
    """Blank claims-development triangle skeleton.

    Row labels are accident periods (first column); columns are development
    periods 0..n. A trailing `Selected CDF` row is where the engine reads
    closure factors from. Example labels are illustrative only.
    """
    headers = ["Accident Period"] + list(range(n_dev))
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=value)
        _style_header_cell(cell, required=col_idx == 1)
    ws.column_dimensions["A"].width = 18

    # Illustrative accident-period labels (user replaces these).
    example_fill = PatternFill("solid", fgColor="F2F2F2")
    base_year = 2021
    for i in range(n_examples):
        year = base_year + i // 4
        quarter = (i % 4) + 1
        c = ws.cell(row=2 + i, column=1, value=f"{year}-Q{quarter}")
        c.fill = example_fill
        c.font = Font(italic=True, color="808080")

    cdf_row = 2 + n_examples + 1
    label = ws.cell(row=cdf_row, column=1, value="Selected CDF")
    label.font = Font(bold=True)
    ws.freeze_panes = "B2"


def _build_instructions_sheet(ws, tpl: TemplateDef) -> None:
    """Leading sheet documenting how to fill the template."""
    ws.cell(row=1, column=1, value=f"{tpl.title} — upload template").font = _TITLE_FONT
    intro = [
        "1. Fill your data into the data sheet(s) in this workbook.",
        "2. Do NOT rename the sheets or change the header row — names and",
        "   spelling/casing must stay exactly as provided.",
        "3. Red headers are required; blue headers are optional.",
        "4. Leave cells blank for unknown optional values (do not type 0).",
        "5. Save as .xlsx and upload it back into Sigma 17.",
    ]
    for i, line in enumerate(intro, start=3):
        ws.cell(row=i, column=1, value=line)

    if tpl.reserve:
        ws.cell(
            row=11, column=1,
            value="Reserve triangles: rows are accident periods (e.g. 2023-Q1), "
            "columns are development periods (0,1,2,...). Enter incremental "
            "claim amounts. The 'Selected CDF' row holds chosen closure factors.",
        )
        ws.column_dimensions["A"].width = 90
        return

    # Column reference table for each data sheet.
    row = 11
    header_font = Font(bold=True)
    for kind in tpl.kinds:
        sheet_name = KIND_RECIPE[kind]["sheet_name"]
        ws.cell(row=row, column=1, value=f"Sheet: {sheet_name}").font = _TITLE_FONT
        row += 1
        for col, label in enumerate(("Column", "Required", "Type", "Description"), start=1):
            ws.cell(row=row, column=col, value=label).font = header_font
        row += 1
        required = set(REQUIRED_FIELDS_FOR_KIND.get(kind, ()))
        for db_field, header in DB_TO_EXCEL_FOR_KIND[kind].items():
            spec = COLUMN_META.get(db_field, ColumnSpec("text", ""))
            type_label = spec.dtype
            if spec.dtype == "enum":
                type_label = "one of: " + " / ".join(spec.enum_values)
            ws.cell(row=row, column=1, value=header)
            ws.cell(row=row, column=2, value="Yes" if db_field in required else "No")
            ws.cell(row=row, column=3, value=type_label)
            ws.cell(row=row, column=4, value=spec.description)
            row += 1
        row += 1

    for letter, width in (("A", 28), ("B", 10), ("C", 22), ("D", 48)):
        ws.column_dimensions[letter].width = width


# --------------------------------------------------------------------------
# Public API
# --------------------------------------------------------------------------
def render_template(key: str) -> bytes:
    """Render the template `key` to .xlsx bytes. Raises KeyError if unknown.

    The data sheet(s) come FIRST: both the importer and the engine read the
    first sheet of an uploaded workbook, so a filled-in single-sheet template
    must have its data sheet as sheet #1. Instructions are appended last.
    """
    tpl = TEMPLATES[key]
    wb = Workbook()
    default = wb.active  # unnamed default sheet; removed once real sheets exist

    if tpl.reserve:
        for sheet_name in ("Paid Claims Triangle", "Reported Triangle"):
            _build_reserve_sheet(wb.create_sheet(sheet_name))
    else:
        for kind in tpl.kinds:
            sheet_name = KIND_RECIPE[kind]["sheet_name"]
            _build_kind_sheet(wb.create_sheet(sheet_name), kind)

    _build_instructions_sheet(wb.create_sheet("Instructions"), tpl)
    wb.remove(default)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
