"""Adapter that converts a DatasetSnapshot into a .xlsx file the engine
can read from disk.

The actuarial engine (~100k LOC) reads .xlsx files from staging
directories and we don't want to touch it. Instead, before each
dataset-driven Celery task runs the engine, we snapshot the referenced
datasets and write a single .xlsx per snapshot into the same staging
directory shape the engine already expects (`in/premium/*.xlsx`,
`in/claims_paid/*.xlsx`, `in/claims_os/*.xlsx`).

This preserves reproducibility (snapshots are immutable) while letting
the engine stay completely unaware that datasets exist.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from ..models import Dataset, DatasetSnapshot
from ..services.snapshots import create_snapshot
from .columns import DB_TO_EXCEL_FOR_KIND


# Per-kind rendering recipe. Three modes:
#   - folder:      one xlsx per snapshot dropped into a folder (single
#                  generic "Sheet1"); the engine globs *.xlsx
#   - named_file:  one xlsx at a specific path with a specific sheet name
#                  (Expense_CF.xlsx / "Expense-CF")
#   - named_sheet: one xlsx at a specific path, but the kind contributes
#                  ONE sheet within that workbook. Other kinds with the
#                  same filename contribute additional sheets.
KIND_RECIPE = {
    Dataset.Kind.PREMIUM: {"mode": "folder", "sheet_name": "Sheet1"},
    Dataset.Kind.CLAIMS_PAID: {"mode": "folder", "sheet_name": "Sheet1"},
    Dataset.Kind.CLAIMS_OS: {"mode": "folder", "sheet_name": "Sheet1"},
    Dataset.Kind.EXPENSE_CF: {
        "mode": "named_file",
        "filename": "Expense_CF.xlsx",
        "sheet_name": "Expense-CF",
    },
    Dataset.Kind.PREVIOUS_PERIOD_LIC: {
        "mode": "named_sheet",
        "filename": "Previous_Period.xlsx",
        "sheet_name": "LIC_BOP",
    },
    Dataset.Kind.PREVIOUS_PERIOD_UPR: {
        "mode": "named_sheet",
        "filename": "Previous_Period.xlsx",
        "sheet_name": "UPR-DAC_BOP",
    },
}


def _snapshot_to_dataframe(snapshot: DatasetSnapshot) -> pd.DataFrame:
    """Convert a snapshot's stored rows into a DataFrame with the
    engine-expected mixed-case column headers."""
    excel_columns = DB_TO_EXCEL_FOR_KIND[snapshot.kind]
    if not snapshot.rows_payload:
        return pd.DataFrame(columns=list(excel_columns.values()))
    df = pd.DataFrame(snapshot.rows_payload)
    df = df.drop(columns=[c for c in ("id", "row_index") if c in df.columns])
    return df.rename(columns=excel_columns)


def dataset_to_dataframe(dataset) -> pd.DataFrame:
    """Live dataset rows as a DataFrame with engine-expected headers.

    The snapshot path (`_snapshot_to_dataframe`) covers jobs, which must read frozen rows.
    This reads the CURRENT rows and exists for previews that run *before* a job — notably
    the UPR impact preview, which must answer "what would this policy do to my book?"
    without creating one.
    """
    from ..models import ROW_MODEL_FOR_KIND

    excel_columns = DB_TO_EXCEL_FOR_KIND[dataset.kind]
    model = ROW_MODEL_FOR_KIND.get(dataset.kind)
    if model is None:
        return pd.DataFrame(columns=list(excel_columns.values()))
    rows = list(
        model.objects.filter(dataset=dataset).values(*excel_columns.keys())
    )
    if not rows:
        return pd.DataFrame(columns=list(excel_columns.values()))
    return pd.DataFrame(rows).rename(columns=excel_columns)


def _snapshot_to_xlsx_bytes(
    snapshot: DatasetSnapshot,
    *,
    sheet_name: str = "Sheet1",
) -> bytes:
    """Render a snapshot's rows as a single-sheet xlsx."""
    df = _snapshot_to_dataframe(snapshot)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buf.getvalue()


def write_snapshot_to_folder(snapshot: DatasetSnapshot, folder: Path) -> Path:
    """Render an existing snapshot as a single .xlsx into folder.

    The file name embeds the snapshot id so multiple snapshots can sit
    side-by-side in the same folder without colliding; the engine globs
    *.xlsx and concatenates them.
    """
    folder.mkdir(parents=True, exist_ok=True)
    xlsx = _snapshot_to_xlsx_bytes(snapshot)
    out = folder / f"snapshot-{snapshot.id}.xlsx"
    out.write_bytes(xlsx)
    return out


def write_snapshot_to_named_file(
    snapshot: DatasetSnapshot,
    path: Path,
    *,
    sheet_name: str,
) -> Path:
    """Render a snapshot as a single-sheet xlsx at `path`. Replaces any
    existing file. Used by `named_file` kinds (e.g. Expense_CF.xlsx)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    xlsx = _snapshot_to_xlsx_bytes(snapshot, sheet_name=sheet_name)
    path.write_bytes(xlsx)
    return path


def write_snapshot_as_sheet(
    snapshot: DatasetSnapshot,
    path: Path,
    *,
    sheet_name: str,
) -> Path:
    """Add or replace `sheet_name` in the workbook at `path` with this
    snapshot's rows. Other sheets in the file are preserved.

    Used by `named_sheet` kinds where multiple datasets contribute
    sheets to the same engine-expected workbook (e.g. LIC_BOP and
    UPR-DAC_BOP in Previous_Period.xlsx).
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    df = _snapshot_to_dataframe(snapshot)
    if not path.exists():
        # First contributor — create the workbook with just this sheet.
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        path.write_bytes(buf.getvalue())
        return path

    # Subsequent contributor — load, drop any existing sheet with the
    # same name, append the new sheet, save. We round-trip via pandas
    # to keep the column-rename / dtype handling consistent with the
    # initial-creation path.
    wb = load_workbook(path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    # openpyxl needs at least one sheet, so we use pandas to write the
    # new sheet into a temp workbook then copy rows over. Simpler: dump
    # the existing workbook back via pandas using its sheets, plus the
    # new one.
    existing_sheets = {
        name: pd.read_excel(path, sheet_name=name, engine="openpyxl")
        for name in wb.sheetnames
    }
    existing_sheets[sheet_name] = df
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, frame in existing_sheets.items():
            frame.to_excel(writer, index=False, sheet_name=name)
    path.write_bytes(buf.getvalue())
    return path


def materialize_datasets_to_folder(
    *,
    datasets: Iterable[Dataset],
    folder: Path,
    consumer_job=None,
) -> list[DatasetSnapshot]:
    """Snapshot each dataset *and* write it to folder in one call.

    Convenient for tests and ad-hoc one-shots; the production job path
    splits the two steps (snapshot at submit time, render at task time)
    via `create_snapshot` + `write_snapshot_to_folder` so the snapshot's
    `created_at` reflects when the user submitted the job.
    """
    snapshots: list[DatasetSnapshot] = []
    for ds in datasets:
        snap = create_snapshot(dataset=ds, consumer_job=consumer_job)
        write_snapshot_to_folder(snap, folder)
        snapshots.append(snap)
    return snapshots
