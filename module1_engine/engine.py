"""
Headless Module 1 actuarial engine (ported from desktop module1.py).
Callers pass directory paths where .xlsx files have been staged.
"""

import pandas as pd
import glob
import os
import re
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime


def select_experience_period(start_period_str, end_period_str):
    start_period = pd.to_datetime(start_period_str, format='%d-%m-%Y')
    end_period = pd.to_datetime(end_period_str, format='%d-%m-%Y')
    return start_period, end_period

def import_data(folder_path, amount_column_name, is_os=False):
    excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
    if not excel_files:
        print("Warning: No Excel files found.")
        return None
    
    recovery_categories = [
        "TP - Morror Recovery", "TP - Insurance Recovery(RASEED)", "TP - Right of Recovery","Right of Recovery","Salvage","Recovery",
        "OD - Salvage / Scrap", "OD – Right of Recovery", "OD - Salvage / Client", "OD - Reversal of Total Loss","Subrogation",
        "Rental OD - Right of Recovery","OD - Others"
    ]

    data_list = []
    for file in excel_files:
        if file.startswith('~$'):  # Skip temporary files
            continue
        if is_os:
            needed_columns = ['AMOUNTOUTSTANDING','ISSUEDATE','LOSSDATE','As at','RESERVINGCLASS','POLICYCLASS','RI_TREATY_TYPE','HEADOFDAMAGE']
            try:
                existing_columns = pd.read_excel(file, nrows=0, engine='openpyxl').columns
            except Exception as e:
                print(f"Failed to read file {file}: {str(e)}")
                continue
            read_columns = [col for col in needed_columns if col in existing_columns]
            
            try:
                df = pd.read_excel(file, usecols=read_columns, parse_dates=['ISSUEDATE','LOSSDATE','As at'], engine='openpyxl')
            except Exception as e:
                print(f"Failed to read data from file {file}: {str(e)}")
                continue
            df['UWY'] = df['ISSUEDATE'].dt.year.apply(lambda x: max(2018, x))
            if 'POLICYCLASS' in df.columns and 'HEADOFDAMAGE' in df.columns:
                df['Amount'] = df.apply(lambda x: x['AMOUNTOUTSTANDING'] if x['POLICYCLASS'] == 'Motor' and x['HEADOFDAMAGE'] in recovery_categories else x['AMOUNTOUTSTANDING'], axis=1)
            else:
                df.rename(columns={'AMOUNTOUTSTANDING': 'Amount'}, inplace=True)
            df['Type'] = 'OS'
        else:
            needed_columns = ['AMOUNTPAID','AMOUNTRECOVERED','ISSUEDATE','LOSSDATE','PAYMENTDATE','RESERVINGCLASS','POLICYCLASS','RI_TREATY_TYPE','HEADOFDAMAGE']
            try:
                existing_columns = pd.read_excel(file, nrows=0, engine='openpyxl').columns
            except Exception as e:
                print(f"Failed to read file {file}: {str(e)}")
                continue
            read_columns = [col for col in needed_columns if col in existing_columns]

            try:
                df = pd.read_excel(file, usecols=read_columns, parse_dates=['ISSUEDATE', 'LOSSDATE', 'PAYMENTDATE'], engine='openpyxl')
            except Exception as e:
                print(f"Failed to read data from file {file}: {str(e)}")
                continue
            df['UWY'] = df['ISSUEDATE'].dt.year.apply(lambda x: max(2018, x))

            if 'POLICYCLASS' in df.columns and 'HEADOFDAMAGE' in df.columns:
                df['Amount'] = df.apply(lambda x: x['AMOUNTRECOVERED'] if x['POLICYCLASS'] == 'Motor' and x['HEADOFDAMAGE'] in recovery_categories else x['AMOUNTPAID'], axis=1)
            else:
                df.rename(columns={'AMOUNTPAID': 'Amount'}, inplace=True)
            df['Type'] = 'Paid'

        data_list.append(df)

    return pd.concat(data_list, ignore_index=True) if data_list else None

def preprocess_data(premium_data_folder):
    dfs = []
    for file_name in os.listdir(premium_data_folder):
        if file_name.endswith('.xlsx') and not file_name.startswith('~$'):  # Skip temporary files
            file_path = os.path.join(premium_data_folder, file_name)
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                df['RiskStartDate'] = pd.to_datetime(df['RiskStartDate'], errors='coerce')
                df['RiskEndDate'] = pd.to_datetime(df['RiskEndDate'], errors='coerce')
                dfs.append(df)
            except Exception as e:
                print("An error occurred while reading a file.")
    return pd.concat(dfs, ignore_index=True) if dfs else None

def preprocess_dates(df):
    df['Underwriting Quarter'] = df['POLICYSTARTDATE'].dt.to_period('Q')
    df['Underwriting Year'] = df['POLICYSTARTDATE'].dt.year
    df['UWY'] = df['Underwriting Year'].apply(lambda x: max(2018, x))
    
    df['ISSUEDATE'] = pd.to_datetime(df['ISSUEDATE'], errors='coerce')
    df['RiskStartDate'] = pd.to_datetime(df['RiskStartDate'], errors='coerce')
    df['RiskEndDate'] = pd.to_datetime(df['RiskEndDate'], errors='coerce')
    df['POLICYSTARTDATE'] = pd.to_datetime(df['POLICYSTARTDATE'], errors='coerce')
    df['POLICYENDDATE'] = pd.to_datetime(df['POLICYENDDATE'], errors='coerce')
    
    return df

def calculate_quarterly_premium(df, start_period, end_period):
    quarters = pd.date_range(start_period, end_period, freq='QE')
    result = []

    for current_quarter in quarters:
        quarter_start = current_quarter.to_period('Q').start_time
        quarter_end = quarter_start + pd.offsets.QuarterEnd(1)
        reserving_class_earned_premium = {}

        for reserving_class in df['RESERVINGCLASS'].unique():
            mask = (df['RESERVINGCLASS'] == reserving_class) & (df['RiskEndDate'] >= quarter_start) & (df['RiskStartDate'] <= quarter_end)
            temp_df = df[mask].copy()
            temp_df['Adjusted Start'] = temp_df['RiskStartDate'].clip(lower=quarter_start)
            temp_df['Adjusted End'] = temp_df['RiskEndDate'].clip(upper=quarter_end)
            temp_df['Days Active'] = (temp_df['Adjusted End'] - temp_df['Adjusted Start']).dt.days + 1
            temp_df['Total Policy Days'] = (temp_df['RiskEndDate'] - temp_df['RiskStartDate']).dt.days + 1
            temp_df['Total Policy Days'] = temp_df['Total Policy Days'].replace(0, np.nan)
            temp_df['PREMIUMAMOUNT'] = pd.to_numeric(temp_df['PREMIUMAMOUNT'], errors='coerce')
            temp_df['Quarterly Premium'] = (temp_df['Days Active'] / temp_df['Total Policy Days']).fillna(0) * temp_df['PREMIUMAMOUNT']
            total_earned_premium = temp_df['Quarterly Premium'].sum()
            reserving_class_earned_premium[reserving_class] = total_earned_premium

        result.append({'Accident_Period': current_quarter.to_period('Q').strftime('%Y-Q%q'), **reserving_class_earned_premium})

    return pd.DataFrame(result)

def get_quarter_end_dates(bop, eop):
    return pd.date_range(start=bop, end=eop, freq='QE')

def calculate_upr(df, bop, eop):
    year_end = eop
    previous_quarter = eop - pd.DateOffset(months=3)
    
    # Convert date columns to datetime if they are of type 'object'
    date_columns = ['ISSUEDATE', 'RiskStartDate', 'RiskEndDate', 'POLICYSTARTDATE', 'POLICYENDDATE']
    for col in date_columns:
        if col in df.columns and df[col].dtype == 'object':
            df[col] = pd.to_datetime(df[col])
    
    # Convert amount columns to numeric
    df['PREMIUMAMOUNT'] = pd.to_numeric(df['PREMIUMAMOUNT'], errors='coerce')
    df['COMMISSIONAMOUNT'] = pd.to_numeric(df['COMMISSIONAMOUNT'], errors='coerce')
    
    # Preprocess other dates if necessary
    preprocess_dates(df)
    
    # Calculate duration in days
    df['Duration'] = (df['RiskEndDate'] - df['RiskStartDate']).dt.days + 1
    df['Duration'] = pd.to_numeric(df['Duration'], errors='coerce')
    # Safe denominator to avoid division by zero (Duration can be 0 for same-day policies)
    duration_safe = np.maximum(df['Duration'], 1)

    # Define conditions and choices for UPR calculation
    conditions = [
        (df['POLICYCLASS'] != "Marine cargo") & (df['ISSUEDATE'] <= year_end) & (df['PRODUCTTYPE'].isin(["Contractors All Risks", "Erection All Risks"])),
        (df['POLICYCLASS'] != "Marine cargo") & (df['ISSUEDATE'] <= year_end) & (~df['PRODUCTTYPE'].isin(["Contractors All Risks", "Erection All Risks"])),
        (df['POLICYCLASS'] == "Marine cargo") & (df['ISSUEDATE'] <= year_end) & (df['ISSUEDATE'] > previous_quarter)
    ]
    choices = [
        1 - (((np.minimum(np.maximum((year_end - df['RiskStartDate']).dt.days + 1, 0), df['Duration']) ** 2) / (duration_safe ** 2))),
        (np.maximum(0, np.minimum(df['Duration'], (df['RiskEndDate'] - year_end).dt.days)) / duration_safe),
        1
    ]
    df['UPR'] = np.select(conditions, choices, default=0) * df['PREMIUMAMOUNT']
    df['Commission Expense'] = df['COMMISSIONAMOUNT']
    df['DAC'] = np.select(conditions, choices, default=0) * df['COMMISSIONAMOUNT']
    
    quarter_dates = get_quarter_end_dates(bop, eop)
    
    for date in quarter_dates:
        date_str = date.strftime('%Y-%m-%d')
        previous_quarter = date - pd.DateOffset(months=3)
        
        conditions = [
            (df['POLICYCLASS'] != "Marine Cargo") & (df['ISSUEDATE'] <= date) & (df['PRODUCTTYPE'].isin(["Contractors All Risks", "Erection All Risks"])),
            (df['POLICYCLASS'] != "Marine Cargo") & (df['ISSUEDATE'] <= date) & (~df['PRODUCTTYPE'].isin(["Contractors All Risks", "Erection All Risks"])),
            (df['POLICYCLASS'] == "Marine Cargo") & (df['ISSUEDATE'] <= date) & (df['ISSUEDATE'] > previous_quarter)
        ]
        choices = [
            1 - (((np.minimum(np.maximum((date - df['RiskStartDate']).dt.days + 1, 0), df['Duration']) ** 2) / (duration_safe ** 2))),
            (np.maximum(0, np.minimum(df['Duration'], (df['RiskEndDate'] - date).dt.days)) / duration_safe),
            1  # Full premium amount for marine cargo within the quarter
        ]
        df[f'UPR_{date_str}'] = np.select(conditions, choices, default=0) * df['PREMIUMAMOUNT']
        
    return df

def summarize_upr_by_reserving_class(df, bop=None, eop=None):
    df['GWP'] = df.apply(lambda x: x['PREMIUMAMOUNT'] if x['RI_TREATY_TYPE'] == 'GROSS' else None, axis=1)
    df['Gross UPR'] = df.apply(lambda x: x['UPR'] if x['RI_TREATY_TYPE'] == 'GROSS' else None, axis=1)
    df['Commission Expense'] = df.apply(lambda x: x['COMMISSIONAMOUNT'] if x['RI_TREATY_TYPE'] == 'GROSS' else None, axis=1)
    df['DAC'] = df.apply(lambda x: x['DAC'] if x['RI_TREATY_TYPE'] == 'GROSS' else None, axis=1)

    df['RI GWP'] = df.apply(lambda x: x['PREMIUMAMOUNT'] if x['RI_TREATY_TYPE'] == 'RI' else None, axis=1)
    df['RI UPR'] = df.apply(lambda x: x['UPR'] if x['RI_TREATY_TYPE'] == 'RI' else None, axis=1)
    df['RI Commission'] = df.apply(lambda x: x['COMMISSIONAMOUNT'] if x['RI_TREATY_TYPE'] == 'RI' else None, axis=1)
    df['UCR'] = df.apply(lambda x: x['DAC'] if x['RI_TREATY_TYPE'] == 'RI' else None, axis=1)

    # Apply date filter
    if bop and eop:
        date_filter = (df['ISSUEDATE'] >= bop) & (df['ISSUEDATE'] <= eop)
        df['GWP'] = df['GWP'].where(date_filter)
        df['RI GWP'] = df['RI GWP'].where(date_filter)
        df['Commission Expense'] = df['Commission Expense'].where(date_filter)
        df['RI Commission'] = df['RI Commission'].where(date_filter)

    # Summarize by RESERVINGCLASS and UWY
    upr_summary = df.groupby(['RESERVINGCLASS', 'UWY']).agg({
        'GWP': 'sum',
        'Gross UPR': 'sum',
        'Commission Expense': 'sum',
        'DAC': 'sum',
        'RI GWP': 'sum',
        'RI UPR': 'sum',
        'RI Commission': 'sum',
        'UCR': 'sum'
    }).reset_index()

    additional_summaries = pd.DataFrame()
    allocation_ep = pd.DataFrame()
    
    last_period_closing = bop - pd.Timedelta(days=1)
    additional_dates = pd.date_range(bop, eop, freq='QE')
    all_dates = [last_period_closing] + list(additional_dates)
    prev_quarter_gross_upr = None
    prev_quarter_ri_upr = None
    
    # Ensure the UPR for last_period_closing is calculated
    quarter_dates = get_quarter_end_dates(last_period_closing, eop)
    df = calculate_upr(df, last_period_closing, eop)
    
    for date in all_dates:
        date_str = date.strftime('%Y-%m-%d')
        quarter_year_str = date.to_period('Q').strftime('%Y-Q%q')
        
        # Ensure the UPR columns exist before proceeding
        upr_col = f'UPR_{date_str}'
        if upr_col not in df.columns:
            df[upr_col] = 0
        
        df[f'Gross_UPR_{date_str}'] = df.apply(lambda x: x[upr_col] if x['RI_TREATY_TYPE'] == 'GROSS' else None, axis=1)
        df[f'RI_UPR_{date_str}'] = df.apply(lambda x: x[upr_col] if x['RI_TREATY_TYPE'] == 'RI' else None, axis=1)
        df[f'GWP_{quarter_year_str}'] = df.apply(lambda x: x['PREMIUMAMOUNT'] if x['RI_TREATY_TYPE'] == 'GROSS' and x['ISSUEDATE'].to_period('Q').strftime('%Y-Q%q') == quarter_year_str else None, axis=1)
        df[f'RI_GWP_{quarter_year_str}'] = df.apply(lambda x: x['PREMIUMAMOUNT'] if x['RI_TREATY_TYPE'] == 'RI' and x['ISSUEDATE'].to_period('Q').strftime('%Y-Q%q') == quarter_year_str else None, axis=1)

        additional_summary = df.groupby(['RESERVINGCLASS', 'UWY']).agg({
            f'Gross_UPR_{date_str}': 'sum',
            f'RI_UPR_{date_str}': 'sum',
            f'GWP_{quarter_year_str}': 'sum',
            f'RI_GWP_{quarter_year_str}': 'sum'
        }).reset_index()

        if prev_quarter_gross_upr is not None and prev_quarter_ri_upr is not None:
            additional_summary[f'GEP_{quarter_year_str}'] = (
                additional_summary[f'GWP_{quarter_year_str}'] 
                - additional_summary[f'Gross_UPR_{date_str}']  # EOP UPR
                + prev_quarter_gross_upr  # BOP UPR
            )
            additional_summary[f'RI_EP_{quarter_year_str}'] = (
                additional_summary[f'RI_GWP_{quarter_year_str}'] 
                - additional_summary[f'RI_UPR_{date_str}']  # EOP UPR
                + prev_quarter_ri_upr  # BOP UPR
            )

            # Calculate total EP for each Accident_Period and ReservingClass
            allocation_ep_quarter = additional_summary[['RESERVINGCLASS', 'UWY', f'GEP_{quarter_year_str}', f'RI_EP_{quarter_year_str}']]
            allocation_ep_quarter = allocation_ep_quarter.rename(columns={
                f'GEP_{quarter_year_str}': 'GEP',
                f'RI_EP_{quarter_year_str}': 'RI_EP'
            })
            allocation_ep_quarter['Accident_Period'] = quarter_year_str

            # Calculate total GEP and RI_EP for each Accident_Period and ReservingClass
            total_ep = allocation_ep_quarter.groupby(['RESERVINGCLASS', 'Accident_Period'])[['GEP', 'RI_EP']].sum().reset_index()
            total_ep = total_ep.rename(columns={'GEP': 'Total_GEP', 'RI_EP': 'Total_RI_EP'})

            # Merge totals back to the original DataFrame
            allocation_ep_quarter = allocation_ep_quarter.merge(total_ep, on=['RESERVINGCLASS', 'Accident_Period'])

            # Calculate percentages (avoid division by zero; np.where would still evaluate division for all rows)
            allocation_ep_quarter['GEP_Percent'] = (allocation_ep_quarter['GEP'] / allocation_ep_quarter['Total_GEP'].replace(0, np.nan)).fillna(0)
            allocation_ep_quarter['RI_EP_Percent'] = (allocation_ep_quarter['RI_EP'] / allocation_ep_quarter['Total_RI_EP'].replace(0, np.nan)).fillna(0)

            # Format percentages
            #allocation_ep_quarter['GEP_Percent'] = allocation_ep_quarter['GEP_Percent'].apply(lambda x: "{:.2%}".format(x))
            #allocation_ep_quarter['RI_EP_Percent'] = allocation_ep_quarter['RI_EP_Percent'].apply(lambda x: "{:.2%}".format(x))

            # Transform the DataFrame to the desired format
            gross_data = allocation_ep_quarter[['RESERVINGCLASS', 'UWY', 'Accident_Period', 'GEP_Percent']]
            gross_data = gross_data.rename(columns={'GEP_Percent': 'EP_Percent'})
            gross_data['GROSS/RI'] = 'GROSS'

            ri_data = allocation_ep_quarter[['RESERVINGCLASS', 'UWY', 'Accident_Period', 'RI_EP_Percent']]
            ri_data = ri_data.rename(columns={'RI_EP_Percent': 'EP_Percent'})
            ri_data['GROSS/RI'] = 'RI'

            # Concatenate the gross and ri data
            allocation_ep_quarter = pd.concat([gross_data, ri_data], ignore_index=True)

            # Select relevant columns and reset index
            allocation_ep_quarter = allocation_ep_quarter[['RESERVINGCLASS', 'UWY', 'Accident_Period', 'GROSS/RI', 'EP_Percent']]

            # Concatenate with the main DataFrame
            allocation_ep = pd.concat([allocation_ep, allocation_ep_quarter], ignore_index=True)

        if additional_summaries.empty:
            additional_summaries = additional_summary
        else:
            additional_summaries = pd.merge(
                additional_summaries,
                additional_summary,
                on=['RESERVINGCLASS', 'UWY'],
                how='outer'
            )

        # Update previous quarter's UPR
        prev_quarter_gross_upr = additional_summary[f'Gross_UPR_{date_str}']
        prev_quarter_ri_upr = additional_summary[f'RI_UPR_{date_str}']

    # Calculate UPR run off
    max_policy_end_date = eop + pd.DateOffset(months=18)
    upr_runoff_dates = pd.date_range(eop, max_policy_end_date, freq='QE')

    upr_runoff = pd.DataFrame()
    prev_quarter_gross_upr = None
    duration_safe = np.maximum(df['Duration'], 1)  # avoid division by zero in runoff loop

    # Ensure calculate_upr handles additional quarterly dates
    for date in upr_runoff_dates:
        date_str = date.strftime('%Y-%m-%d')
        quarter_year_str = date.to_period('Q').strftime('%Y-Q%q')

        conditions = [
            (df['POLICYCLASS'] != "Marine cargo") & (df['ISSUEDATE'] <= date) & (df['PRODUCTTYPE'].isin(["Contractors All Risks", "Erection All Risks"])),
            (df['POLICYCLASS'] != "Marine cargo") & (df['ISSUEDATE'] <= date) & (~df['PRODUCTTYPE'].isin(["Contractors All Risks", "Erection All Risks"])),
            (df['POLICYCLASS'] == "Marine cargo") & (df['ISSUEDATE'] <= date) & (df['ISSUEDATE'] > date - pd.Timedelta(days=91))
        ]
        choices = [
            1 - (((np.minimum(np.maximum((date - df['RiskStartDate']).dt.days + 1, 0), df['Duration']) ** 2) / (duration_safe ** 2))),
            (np.maximum(0, np.minimum(df['Duration'], (df['RiskEndDate'] - date).dt.days)) / duration_safe),
            1
        ]
        df[f'UPR_{date_str}'] = np.select(conditions, choices, default=0) * df['PREMIUMAMOUNT']

        df[f'Gross_UPR_{date_str}'] = df.apply(lambda x: x[f'UPR_{date_str}'] if x['RI_TREATY_TYPE'] == 'GROSS' else None, axis=1)

        runoff_summary = df.groupby(['RESERVINGCLASS', 'UWY']).agg({
            f'Gross_UPR_{date_str}': 'sum'
        }).reset_index()

        runoff_summary = runoff_summary.rename(columns={
            f'Gross_UPR_{date_str}': date.to_period('Q').strftime('%Y-Q%q')
        })

        if prev_quarter_gross_upr is not None:
            # Only calculate GEP if the quarter is after the eop quarter
            if date > eop:
                runoff_summary[f'GEP_{quarter_year_str}'] = (
                    prev_quarter_gross_upr  # BOP UPR
                    - runoff_summary[date.to_period("Q").strftime('%Y-Q%q')]  # EOP UPR
                )

        prev_quarter_gross_upr = runoff_summary[date.to_period('Q').strftime('%Y-Q%q')]

        upr_runoff = pd.merge(upr_runoff, runoff_summary, on=['RESERVINGCLASS', 'UWY'], how='outer') if not upr_runoff.empty else runoff_summary

    # Separate UPR and Earned Premium columns
    #upr_columns = [col for col in upr_runoff.columns if 'GEP' not in col]
    #earned_premium_columns = [col for col in upr_runoff.columns if 'GEP' in col]
    #upr_runoff = pd.concat([upr_runoff[upr_columns], upr_runoff[earned_premium_columns]], axis=1)
    
    # Retain only RESERVINGCLASS, UWY, and earned premium columns
    earned_premium_columns = ['RESERVINGCLASS', 'UWY'] + [col for col in upr_runoff.columns if 'GEP' in col]
    upr_runoff = upr_runoff[earned_premium_columns]

    return upr_summary, additional_summaries, allocation_ep, upr_runoff


def calculate_policy_level_upr(df, bop, eop):
    # Filter out rows where RiskEndDate is less than eop
    df = df[df['RiskEndDate'] >= eop]

    # Ensure the UPR is calculated for the remaining rows
    df = calculate_upr(df, bop, eop)

    # Group by policy number, reserving class, and RI_TREATY_TYPE
    policy_level_upr = df.groupby(['POLICYNUMBER', 'RESERVINGCLASS', 'RI_TREATY_TYPE']).agg({
        'UPR': 'sum'
    }).reset_index()

    # Pivot the data to separate Gross and RI UPR
    policy_level_upr = policy_level_upr.pivot_table(
        index=['POLICYNUMBER', 'RESERVINGCLASS'], 
        columns='RI_TREATY_TYPE', 
        values='UPR', 
        fill_value=0
    ).reset_index()
    
    # Rename columns for clarity
    policy_level_upr.columns.name = None
    policy_level_upr.rename(columns={'GROSS': 'Gross UPR', 'RI': 'RI UPR'}, inplace=True)
    
    # Filter out policies with both Gross UPR and RI UPR as 0
    policy_level_upr = policy_level_upr[(policy_level_upr['Gross UPR'] != 0) | (policy_level_upr['RI UPR'] != 0)]
    
    return policy_level_upr

def run_policy_level_upr(
    bop_str: str,
    eop_str: str,
    premium_data_folder: str,
    output_dir: str,
) -> str:
    """Write policy-level UPR workbook to output_dir. Returns output file path."""
    if not premium_data_folder:
        raise ValueError("Premium data folder is required.")
    if not output_dir:
        raise ValueError("Output directory is required.")

    bop = pd.to_datetime(bop_str, format="%d-%m-%Y")
    eop = pd.to_datetime(eop_str, format="%d-%m-%Y")

    df = preprocess_data(premium_data_folder)
    if df is None:
        raise ValueError("No valid data found in the premium data folder.")

    policy_level_upr_df = calculate_policy_level_upr(df, bop, eop)
    eop_date_str = eop.strftime("%d-%m-%Y")
    output_file_path = os.path.join(output_dir, f"Policy Level UPR {eop_date_str}.xlsx")
    policy_level_upr_df.to_excel(output_file_path, index=False)
    return output_file_path

def calculate_claims_paid_summary(df, bop, eop):
    # Convert bop and eop to datetime if they are not already
    bop = pd.to_datetime(bop, format='%d-%m-%Y')
    eop = pd.to_datetime(eop, format='%d-%m-%Y')
    
    # Filter the DataFrame to only include rows with PAYMENTDATE between bop and eop
    df = df[(df['PAYMENTDATE'] >= bop) & (df['PAYMENTDATE'] <= eop)]
    
    df['Payment Year'] = df['PAYMENTDATE'].dt.year
    df['Loss Year'] = df['LOSSDATE'].dt.year
    
    conditions = [
        (df['RI_TREATY_TYPE'] == 'GROSS') & (df['Payment Year'] == df['Loss Year']),
        (df['RI_TREATY_TYPE'] == 'RI') & (df['Payment Year'] == df['Loss Year']),
        (df['RI_TREATY_TYPE'] == 'GROSS') & (df['Payment Year'] != df['Loss Year']),
        (df['RI_TREATY_TYPE'] == 'RI') & (df['Payment Year'] != df['Loss Year'])
    ]
    choices = ['Gross CY Paid', 'RI CY Paid', 'Gross PY Paid', 'RI PY Paid']
    df['Payment Category'] = np.select(conditions, choices, default='Other')
    claims_paid_summary = df.groupby(['RESERVINGCLASS','UWY','Payment Category'])['Amount'].sum().unstack(fill_value=0).reset_index()
    return claims_paid_summary

def calculate_claims_os_summary(df, end_period):
    df['As at Year'] = df['As at'].dt.year
    df['Loss Year'] = df['LOSSDATE'].dt.year
    df['Accident_Period'] = df['LOSSDATE'].dt.to_period('Q').dt.strftime('%Y-Q%q')
    df['Underwriting Year'] = df['ISSUEDATE'].dt.year
    df['UWY'] = df['Underwriting Year'].apply(lambda x: max(2018, x))
    
    recovery_categories = [
        "TP - Morror Recovery", "TP - Insurance Recovery(RASEED)", "TP - Right of Recovery","Right of Recovery","Salvage","Recovery",
        "OD - Salvage / Scrap", "OD – Right of Recovery", "OD - Salvage / Client", "OD - Reversal of Total Loss","Subrogation",
        "Rental OD - Right of Recovery","OD - Others"
    ]
    conditions = [
        (df['RI_TREATY_TYPE'] == 'GROSS') & (df['As at Year'] == df['Loss Year']) & (~df['HEADOFDAMAGE'].isin(recovery_categories)),
        (df['RI_TREATY_TYPE'] == 'RI') & (df['As at Year'] == df['Loss Year']) & (~df['HEADOFDAMAGE'].isin(recovery_categories)),
        (df['RI_TREATY_TYPE'] == 'GROSS') & (df['As at Year'] != df['Loss Year']) & (~df['HEADOFDAMAGE'].isin(recovery_categories)),
        (df['RI_TREATY_TYPE'] == 'RI') & (df['As at Year'] != df['Loss Year']) & (~df['HEADOFDAMAGE'].isin(recovery_categories)),
        (df['RI_TREATY_TYPE'] == 'GROSS') & (df['As at Year'] == df['Loss Year']) & (df['HEADOFDAMAGE'].isin(recovery_categories)),
        (df['RI_TREATY_TYPE'] == 'RI') & (df['As at Year'] == df['Loss Year']) & (df['HEADOFDAMAGE'].isin(recovery_categories)),
        (df['RI_TREATY_TYPE'] == 'GROSS') & (df['As at Year'] != df['Loss Year']) & (df['HEADOFDAMAGE'].isin(recovery_categories)),
        (df['RI_TREATY_TYPE'] == 'RI') & (df['As at Year'] != df['Loss Year']) & (df['HEADOFDAMAGE'].isin(recovery_categories))
    ]
    choices = ['Gross CY OS', 'RI CY OS', 'Gross PY OS', 'RI PY OS', 'Gross CY SS', 'RI CY SS', 'Gross PY SS', 'RI PY SS']
    df['OS Category'] = np.select(conditions, choices, default='Other')
    claims_os_summary = df.groupby(['RESERVINGCLASS','UWY','OS Category'])['Amount'].sum().unstack(fill_value=0).reset_index()
    
    # Filtering LIC (OS) Summary by end_period
    lic_os_summary = df[df['As at'] == end_period].groupby(['RESERVINGCLASS', 'Accident_Period', 'UWY', 'RI_TREATY_TYPE', 'As at']).agg(
        Outstanding=('Amount', lambda x: x[df['OS Category'].str.contains('OS')].sum()),
        SS=('Amount', lambda x: x[df['OS Category'].str.contains('SS')].sum())
    ).reset_index()
    
    # Adding the "Gross/RI" column
    lic_os_summary['GROSS/RI'] = np.where(lic_os_summary['RI_TREATY_TYPE'] == 'GROSS', 'GROSS', 'RI')
    
    # Formatting the As at date
    lic_os_summary['As at'] = lic_os_summary['As at'].dt.strftime('%d/%m/%Y')

    # Reordering columns to match the required format
    lic_os_summary = lic_os_summary[['RESERVINGCLASS', 'UWY', 'Accident_Period', 'GROSS/RI', 'Outstanding', 'SS', 'As at']]
    
    return claims_os_summary, lic_os_summary

def export_upr_summary_to_excel(summary_df, file_path):
    summary_df.to_excel(file_path, index=False, sheet_name='UPR Summary')

def export_claims_paid_summary_to_excel(summary_claims_paid_df, file_path):
    summary_claims_paid_df.to_excel(file_path, index=False, sheet_name='Claims Paid Summary')    

def export_claims_os_summary_to_excel(summary_claims_os_df, file_path):
    summary_claims_os_df.to_excel(file_path, index=False, sheet_name='Claims OS Summary') 
    
def calculate_incremental_triangle(df, start_period, end_period, is_os=False):
    # Generate a list of quarterly periods from start_period to end_period
    accident_periods = pd.date_range(start_period, end_period, freq='QE').to_period('Q')
    max_development_period = len(accident_periods)
    
    # Create a DataFrame with zeroes, indexed by the accident periods and columns for each development period
    triangle = pd.DataFrame(0, index=accident_periods.strftime('%Y-Q%q'), columns=range(max_development_period))
    
    for index, row in df.iterrows():
        if is_os and pd.notna(row['As at']):
            accident_quarter = row['LOSSDATE'].to_period('Q')
            as_at_quarter = pd.to_datetime(row['As at']).to_period('Q')
            development_quarter = (as_at_quarter - accident_quarter).n
            
            if accident_quarter in accident_periods and 0 <= development_quarter < max_development_period:
                triangle.at[accident_quarter.strftime('%Y-Q%q'), development_quarter] += row['Amount']
        elif 'PAYMENTDATE' in df.columns and pd.notna(row['PAYMENTDATE']):
            accident_quarter = row['LOSSDATE'].to_period('Q')
            payment_quarter = row['PAYMENTDATE'].to_period('Q')
            development_quarter = (payment_quarter - accident_quarter).n
            
            if accident_quarter in accident_periods and 0 <= development_quarter < max_development_period:
                triangle.at[accident_quarter.strftime('%Y-Q%q'), development_quarter] += row['Amount']
    
    # Reset the index to keep it clean for further operations
    triangle.reset_index(inplace=True)
    triangle.rename(columns={'index': 'Accident Period'}, inplace=True)
    return triangle

def calculate_cumulative_triangle(incremental_triangle):
    cumulative_triangle = incremental_triangle.copy()
    for i in range(len(cumulative_triangle)):
        max_col_index = len(cumulative_triangle.columns) - i - 1
        cumulative_triangle.iloc[i, 1:max_col_index+1] = incremental_triangle.iloc[i, 1:max_col_index+1].cumsum()
        cumulative_triangle.iloc[i, max_col_index+1:] = np.nan
    return cumulative_triangle

def calculate_age_to_age_factors(cumulative_triangle):
    age_to_age_factors = pd.DataFrame(index=cumulative_triangle.index, columns=cumulative_triangle.columns)
    for i in range(1, len(cumulative_triangle.columns) - 1):
        current_column = cumulative_triangle.iloc[:, i]
        next_column = cumulative_triangle.iloc[:, i + 1]
        # Safe division: avoid evaluating division when current_column is 0
        age_to_age_factors.iloc[:, i] = (next_column / current_column.replace(0, np.nan)).fillna(0)
        age_to_age_factors.iloc[:, 0] = cumulative_triangle.iloc[:, 0]
    return age_to_age_factors

def run_update_reserve_summary(folder_path: str) -> None:
    """Update reserve workbooks in folder_path. Raises on error."""
    if not folder_path:
        raise ValueError("Folder path is required.")

    combined_summary_file = "Combined_Summary.xlsx"
    ibnr_summaries = []
    output_dir = folder_path

    header_style = {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="4F81BD"),
    }

    for file in os.listdir(folder_path):
        if file.endswith(".xlsx") and file != combined_summary_file:
            output_file_path = os.path.join(folder_path, file)

            file_base = os.path.splitext(file)[0]
            parts = re.split(r"\s+", file_base)

            if len(parts) < 3:
                continue

            ri_type = parts[-2].upper()
            head_of_damage = parts[-3]
            reserving_class = " ".join(parts[:-3])

            try:
                # Load workbook with formulas
                wb = load_workbook(output_file_path, data_only=False)
                ws_paid_triangle = wb['Paid Claims Triangle']
                ws_reported_triangle = wb['Reported Triangle']

                # Find the row index for "Selected CDF"
                def find_selected_cdf_row(sheet):
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
                        for cell in row:
                            if cell.value == 'Selected CDF':
                                return cell.row
                    return None

                selected_cdf_row_paid = find_selected_cdf_row(ws_paid_triangle)
                selected_cdf_row_reported = find_selected_cdf_row(ws_reported_triangle)

                if not selected_cdf_row_paid or not selected_cdf_row_reported:
                    continue

                num_cols = ws_paid_triangle.max_column - 1
                selected_cdf_values_paid_data = []
                selected_cdf_values_reported_data = []

                # Load workbook with data only to get actual values
                wb_data_only = load_workbook(output_file_path, data_only=True)
                ws_paid_triangle_data = wb_data_only['Paid Claims Triangle']
                ws_reported_triangle_data = wb_data_only['Reported Triangle']

                for col in range(2, num_cols + 2):
                    paid_value = ws_paid_triangle_data.cell(row=selected_cdf_row_paid, column=col).value
                    reported_value = ws_reported_triangle_data.cell(row=selected_cdf_row_reported, column=col).value
                    selected_cdf_values_paid_data.append(float(paid_value) if paid_value is not None else 2.0)
                    selected_cdf_values_reported_data.append(float(reported_value) if reported_value is not None else 2.0)

                # Reverse CDF values
                selected_cdf_values_paid_data.reverse()
                selected_cdf_values_reported_data.reverse()

                # Update Reserve Summary sheet
                ws = wb['Reserve Summary']
                new_headers = [
                    'Implied LR', 'Paid CDF', 'Reported CDF', 'Paid CL Ultimate',
                    'Reported CL Ultimate', 'ELR Ultimate', 'Paid BF Ultimate', 'Reported BF Ultimate',
                    'Selected Method', 'Ultimate Claims', 'IBNR', 'ULR', 'CDF'
                ]
                existing_headers = [
                    ws.cell(row=1, column=col).value
                    for col in range(1, ws.max_column + 1)
                    if ws.cell(row=1, column=col).value
                ]
                new_headers_start_col = len(existing_headers) + 1
                headers = existing_headers + new_headers

                reserving_class_data = []

                n_paid = len(selected_cdf_values_paid_data)
                n_reported = len(selected_cdf_values_reported_data)
                for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(existing_headers))):
                    data = {existing_headers[col_idx]: cell.value for col_idx, cell in enumerate(row)}
                    data['Implied LR'] = None
                    i_paid = min(idx, n_paid - 1) if n_paid else 0
                    i_reported = min(idx, n_reported - 1) if n_reported else 0
                    data['Paid CDF'] = selected_cdf_values_paid_data[i_paid] if n_paid else 2.0
                    data['Reported CDF'] = selected_cdf_values_reported_data[i_reported] if n_reported else 2.0
                    data['Paid CL Ultimate'] = data['Paid Claims'] * data['Paid CDF']
                    data['Reported CL Ultimate'] = data['Reported Claims'] * data['Reported CDF']
                    data['ELR Ultimate'] = f'=IFERROR(G{idx + 2} * B{idx + 2},0)'
                    data['Paid BF Ultimate'] = f'=IFERROR((1-1/H{idx + 2}) * B{idx + 2} * G{idx + 2} + D{idx + 2}, 0)'
                    data['Reported BF Ultimate'] = f'=IFERROR((1-1/I{idx + 2}) * B{idx + 2} * G{idx + 2} + E{idx + 2}, 0)'
                    data['Selected Method'] = 'Paid CL'
                    data['Ultimate Claims'] = f'=IF(O{idx + 2}="Paid CL", J{idx + 2}, IF(O{idx + 2}="Reported CL", K{idx + 2}, IF(O{idx + 2}="ELR", L{idx + 2}, IF(O{idx + 2}="Reported BF", N{idx + 2}, M{idx + 2}))))'
                    data['IBNR'] = f'=IFERROR(P{idx + 2} - E{idx + 2},0)'
                    data['ULR'] = f'=IFERROR(P{idx + 2}/B{idx + 2},0)'
                    data['CDF'] = f'=IFERROR(P{idx + 2}/C{idx + 2},0)'
                    reserving_class_data.append(data)

                    # Write updated values back to worksheet
                    for col_idx, key in enumerate(headers, start=1):
                        ws.cell(row=idx + 2, column=col_idx, value=data.get(key, None))

                # Write new headers
                for col_idx, header in enumerate(new_headers, start=new_headers_start_col):
                    ws.cell(row=1, column=col_idx, value=header)

                # Apply header style
                style_headers(ws, header_style)

                # Add data validation
                max_row = ws.max_row
                dv = DataValidation(type="list", formula1='"Paid CL,Reported CL,ELR,Paid BF,Reported BF"', allow_blank=False)
                dv.add(f'O2:O{max_row}')
                ws.add_data_validation(dv)

                # Format columns
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 15.22
                    for cell in col:
                        cell.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'

                # Save workbook
                wb.save(output_file_path)

                # Append metadata
                df = pd.DataFrame(reserving_class_data)
                df['RESERVINGCLASS'] = reserving_class
                df['Payment/Recovery'] = head_of_damage
                df['GROSS/RI'] = ri_type
                ibnr_summaries.append(df)

            except Exception:
                continue

    # Combine IBNR summaries and save to Combined_Summary.xlsx
    if ibnr_summaries:
        combined_ibnr_summary = pd.concat(ibnr_summaries, ignore_index=True)
        combined_summary_path = os.path.join(output_dir, combined_summary_file)

        if os.path.exists(combined_summary_path):
            with pd.ExcelWriter(
                combined_summary_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            ) as writer:
                combined_ibnr_summary.to_excel(writer, sheet_name="IBNR Summary", index=False)
                combined_ws = writer.book["IBNR Summary"]
                max_row = combined_ws.max_row
                dv = DataValidation(
                    type="list",
                    formula1='"Paid CL,Reported CL,ELR,Paid BF,Reported BF"',
                    allow_blank=False,
                )
                dv.add(f"O2:O{max_row}")
                combined_ws.add_data_validation(dv)
        else:
            with pd.ExcelWriter(combined_summary_path, engine="openpyxl") as writer:
                combined_ibnr_summary.to_excel(writer, sheet_name="IBNR Summary", index=False)
                combined_ws = writer.book["IBNR Summary"]
                combined_ws.sheet_state = "visible"
                max_row = combined_ws.max_row
                dv = DataValidation(
                    type="list",
                    formula1='"Paid CL,Reported CL,ELR,Paid BF,Reported BF"',
                    allow_blank=False,
                )
                dv.add(f"O2:O{max_row}")
                combined_ws.add_data_validation(dv)


def style_headers(ws, header_style):
    for cell in ws[1]:  # Assuming headers are in the first row
        cell.font = header_style['font']
        cell.fill = header_style['fill']

def append_data_below(writer, df, sheet_name, header_style):
    if sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]
        startrow = ws.max_row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=startrow + 1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        if startrow == 1:  # If no headers yet, apply header style
            style_headers(ws, header_style)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.book[sheet_name]
        style_headers(ws, header_style)
        
def append_data_right(writer, df, sheet_name, header_style):
    if sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]
        max_col = ws.max_column
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=max_col + 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        style_headers(ws, header_style)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.book[sheet_name]
        style_headers(ws, header_style)
        
def overwrite_sheet(writer, df, sheet_name, header_style):
    if sheet_name in writer.book.sheetnames:
        del writer.book[sheet_name]
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.book[sheet_name]
    style_headers(ws, header_style)

def generate_ul_ae_ra_df(dataframe, value):
    df = dataframe[['RESERVINGCLASS']].drop_duplicates().reset_index(drop=True)
    df.columns = ['RESERVINGCLASS']
    df['GROSS/RI'] = value
    df['ULAE %'] = ''
    df['RA %'] = ''
    return df[['RESERVINGCLASS', 'GROSS/RI', 'ULAE %', 'RA %']]
    
def run_generate_summary(
    start_period_str: str,
    end_period_str: str,
    bop_str: str,
    eop_str: str,
    premium_data_folder: str,
    claims_paid_folder: str,
    claims_os_folder: str,
    output_dir: str,
) -> None:
    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    header_style = {"font": header_font, "fill": header_fill}

    start_period, end_period = select_experience_period(start_period_str, end_period_str)

    bop = pd.to_datetime(bop_str, format="%d-%m-%Y")
    eop = pd.to_datetime(eop_str, format="%d-%m-%Y")

    df = preprocess_data(premium_data_folder)
    if df is None:
        raise ValueError("No valid data found in the premium data folder.")

    summary_df = calculate_quarterly_premium(df, bop, eop)
    df = calculate_upr(df, bop=bop, eop=eop)
    upr_summary_df, additional_summaries, allocation_ep, upr_runoff = summarize_upr_by_reserving_class(
        df, bop=bop, eop=eop
    )

    paid_data = import_data(claims_paid_folder, "AMOUNTPAID", is_os=False)
    if paid_data is None or "PAYMENTDATE" not in paid_data.columns:
        raise ValueError("No valid 'PAYMENTDATE' found in the data.")
    claims_paid_summary_df = calculate_claims_paid_summary(paid_data, bop, eop)

    os_data = import_data(claims_os_folder, "AMOUNTOUTSTANDING", is_os=True)
    if os_data is None or "As at" not in os_data.columns:
        raise ValueError("No valid 'As at' found in the data.")
    claims_os_summary_df, lic_os_summary = calculate_claims_os_summary(os_data, eop)

    filtered_os_data = os_data[os_data["As at"] == eop]
    filtered_os_summary_df, _ = calculate_claims_os_summary(filtered_os_data, eop)

    combined_summary = pd.merge(upr_summary_df, claims_paid_summary_df, on=["RESERVINGCLASS", "UWY"], how="outer")
    combined_summary = pd.merge(combined_summary, filtered_os_summary_df, on=["RESERVINGCLASS", "UWY"], how="outer")
    combined_summary.fillna(0, inplace=True)
    combined_summary.sort_values(by=["RESERVINGCLASS", "UWY"], inplace=True)

    combined_data = pd.concat([paid_data, os_data], ignore_index=True)
    combined_data["LOSSDATE"] = pd.to_datetime(combined_data["LOSSDATE"])
    combined_data["Accident_Period"] = combined_data["LOSSDATE"].dt.to_period("Q").apply(
        lambda x: x.strftime("%Y-Q%q")
    )
    combined_data = combined_data[
        (combined_data["LOSSDATE"] >= start_period) & (combined_data["LOSSDATE"] <= end_period)
    ]

    if not output_dir:
        raise ValueError("No output directory selected.")

    output_file_path = os.path.join(output_dir, "Combined_Summary.xlsx")

    # UW Summary calculations
    df_premium_gross = df[df['RI_TREATY_TYPE'] == 'GROSS']
    df_premium_gross = calculate_upr(df_premium_gross, bop=bop, eop=eop)

    paid_data_gross = paid_data[paid_data['RI_TREATY_TYPE'] == 'GROSS']
    paid_data_summary_gross = paid_data_gross.groupby(['RESERVINGCLASS', 'UWY'])[['AMOUNTPAID', 'AMOUNTRECOVERED']].sum().reset_index()

    os_data_gross = os_data[(os_data['RI_TREATY_TYPE'] == 'GROSS') & (os_data['As at'] == eop)]
    os_data_summary_gross = os_data_gross.groupby(['RESERVINGCLASS', 'UWY'])['AMOUNTOUTSTANDING'].sum().reset_index()

    premium_summary_gross = df_premium_gross.groupby(['RESERVINGCLASS', 'UWY']).agg({
        'PREMIUMAMOUNT': 'sum',
        'COMMISSIONAMOUNT': 'sum',
        'UPR': 'sum'
    }).reset_index()

    combined_summary_gross = premium_summary_gross.merge(
        paid_data_summary_gross, on=['RESERVINGCLASS', 'UWY'], how='left'
    ).merge(
        os_data_summary_gross, on=['RESERVINGCLASS', 'UWY'], how='left'
    ).fillna(0)

    combined_summary_gross['GWP'] = combined_summary_gross['PREMIUMAMOUNT']
    combined_summary_gross['Commission Expense'] = combined_summary_gross['COMMISSIONAMOUNT']
    combined_summary_gross['Paid Claims'] = combined_summary_gross['AMOUNTPAID'] + combined_summary_gross['AMOUNTRECOVERED']
    combined_summary_gross['OS'] = combined_summary_gross['AMOUNTOUTSTANDING']
    combined_summary_gross['GEP'] = combined_summary_gross['GWP'] - combined_summary_gross['UPR']
    combined_summary_gross['Incurred Claims'] = combined_summary_gross['Paid Claims'] + combined_summary_gross['OS']
    # Safe division: replace 0 with nan so division is skipped, then fillna(0)
    combined_summary_gross['Paid LR'] = (combined_summary_gross['Paid Claims'] / combined_summary_gross['GEP'].replace(0, np.nan)).fillna(0)
    combined_summary_gross['Inc LR'] = (combined_summary_gross['Incurred Claims'] / combined_summary_gross['GEP'].replace(0, np.nan)).fillna(0)
    combined_summary_gross['Comm Ratio'] = (combined_summary_gross['Commission Expense'] / combined_summary_gross['GWP'].replace(0, np.nan)).fillna(0)
    
    uw_summary = combined_summary_gross[['RESERVINGCLASS', 'UWY', 'GWP', 'UPR', 'GEP', 'Commission Expense', 'Paid Claims', 'OS', 'Incurred Claims', 'Paid LR', 'Inc LR', 'Comm Ratio']]
    uw_summary['Exp Ratio'] = ''
    uw_summary['RI %'] = ''
    
    df['FY'] = df['ISSUEDATE'].dt.year
    premium_fy_summary = df.groupby(['POLICYCLASS', 'FY', 'RI_TREATY_TYPE'])['PREMIUMAMOUNT'].sum().reset_index()

    paid_data['FY'] = paid_data['PAYMENTDATE'].dt.year
    paid_fy_summary = paid_data.groupby(['POLICYCLASS', 'FY', 'RI_TREATY_TYPE']).agg({
        'AMOUNTPAID': 'sum',
        'AMOUNTRECOVERED': 'sum'
    }).reset_index()
    paid_fy_summary['Paid Claims'] = paid_fy_summary['AMOUNTPAID'] + paid_fy_summary['AMOUNTRECOVERED']

    os_data['FY'] = os_data['As at'].dt.year
    os_data = os_data[os_data['As at'] == eop]
    os_fy_summary = os_data.groupby(['POLICYCLASS', 'FY', 'RI_TREATY_TYPE'])['AMOUNTOUTSTANDING'].sum().reset_index()

    fy_summary = premium_fy_summary.merge(
        paid_fy_summary[['POLICYCLASS', 'FY', 'RI_TREATY_TYPE', 'Paid Claims']],
        on=['POLICYCLASS', 'FY', 'RI_TREATY_TYPE'],
        how='left'
    ).merge(
        os_fy_summary[['POLICYCLASS', 'FY', 'RI_TREATY_TYPE', 'AMOUNTOUTSTANDING']],
        on=['POLICYCLASS', 'FY', 'RI_TREATY_TYPE'],
        how='left'
    ).fillna(0)

    fy_summary = fy_summary[['POLICYCLASS', 'FY', 'RI_TREATY_TYPE', 'PREMIUMAMOUNT', 'Paid Claims', 'AMOUNTOUTSTANDING']]

    # Save or append data to the Excel file using openpyxl
    
    if os.path.exists(output_file_path):
        # Append to existing Excel file
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            # ---- Combined Summary ----
            overwrite_sheet(writer, combined_summary, 'Combined Summary', header_style)
    
            # ---- Additional Summary ----
            # Check if sheet exists in the file
            existing_sheets = pd.ExcelFile(output_file_path, engine='openpyxl').sheet_names
            if "Additional Summary" in existing_sheets:
                # Read the original Additional Summary
                original_additional_summary = pd.read_excel(output_file_path, sheet_name="Additional Summary", engine='openpyxl')
                # Merge with new data
                additional_summaries = pd.merge(
                    original_additional_summary,
                    additional_summaries,
                    on=['RESERVINGCLASS', 'UWY'],
                    how='outer'
                )
            # Write merged Additional Summary
            overwrite_sheet(writer, additional_summaries, 'Additional Summary', header_style)
    
            # ---- Other sheets ----
            append_data_below(writer, allocation_ep, 'Allocation EP', header_style)
            overwrite_sheet(writer, lic_os_summary, 'LIC (OS) Summary', header_style)
            overwrite_sheet(writer, upr_runoff, 'UPR Run-Off', header_style)
            overwrite_sheet(writer, uw_summary, 'UW Summary', header_style)
            overwrite_sheet(writer, fy_summary, 'FY Summary', header_style)
    
    else:
        # File does not exist → create new
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            combined_summary.to_excel(writer, sheet_name='Combined Summary', index=False)
            style_headers(writer.sheets['Combined Summary'], header_style)
    
            additional_summaries.to_excel(writer, sheet_name='Additional Summary', index=False)
            style_headers(writer.sheets['Additional Summary'], header_style)
    
            append_data_below(writer, allocation_ep, 'Allocation EP', header_style)
            overwrite_sheet(writer, lic_os_summary, 'LIC (OS) Summary', header_style)
            overwrite_sheet(writer, upr_runoff, 'UPR Run-Off', header_style)
            overwrite_sheet(writer, uw_summary, 'UW Summary', header_style)
            overwrite_sheet(writer, fy_summary, 'FY Summary', header_style)

        
    # Reserve Summary logic
    ibnr_summaries = []

    for reserving_class in df['RESERVINGCLASS'].unique():
        for head_of_damage in combined_data['HEADOFDAMAGE'].unique():
            for ri_type in ['GROSS', 'RI']:
                filtered_data = combined_data[(combined_data['RESERVINGCLASS'] == reserving_class) & (combined_data['RI_TREATY_TYPE'] == ri_type) & (combined_data['HEADOFDAMAGE'] == head_of_damage)]

                if "MOTOR" in reserving_class:
                    output_file_name = f"{reserving_class} {head_of_damage} {ri_type} {eop.strftime('%Y-%m')}.xlsx"
                else:
                    output_file_name = f"{reserving_class} {head_of_damage} {ri_type} {eop.strftime('%Y-%m')}.xlsx"

                if ri_type == 'GROSS':
                    ep_column_prefix = 'GEP_'
                else:
                    ep_column_prefix = 'RI_EP_'
                
                accident_periods = pd.period_range(start=start_period, end=end_period, freq='Q')
                accident_period_strs = [f"{p.year}-Q{p.quarter}" for p in accident_periods]
                ep_columns = [f'{ep_column_prefix}{p}' for p in accident_period_strs]
                
                # additional_summaries only has GEP/RI_EP for quarters between BOP and EOP; fill missing with 0
                existing_ep_columns = [c for c in ep_columns if c in additional_summaries.columns]
                missing_ep_columns = [c for c in ep_columns if c not in additional_summaries.columns]
                reserving_class_ep = additional_summaries[additional_summaries['RESERVINGCLASS'] == reserving_class][['RESERVINGCLASS', 'UWY'] + existing_ep_columns].copy()
                for col in missing_ep_columns:
                    reserving_class_ep[col] = 0
                reserving_class_ep = reserving_class_ep[['RESERVINGCLASS', 'UWY'] + ep_columns]
                
                summed_ep = reserving_class_ep.melt(id_vars=['RESERVINGCLASS', 'UWY'],var_name='Accident_Period', value_name='EP')
                summed_ep['Accident_Period'] = summed_ep['Accident_Period'].str.replace(ep_column_prefix, '', regex=False)
                summed_ep = summed_ep.groupby('Accident_Period')['EP'].sum().reset_index()

                reserving_class_claims = filtered_data.groupby(['Accident_Period', 'Type'])['Amount'].sum().unstack().fillna(0)
                reserving_class_claims = reserving_class_claims.reindex(columns=['Paid', 'OS'], fill_value=0)
                reserving_class_claims.columns = ['Paid Claims', 'OS Claims']

                os_data_reserving_class = os_data[(os_data['RESERVINGCLASS'] == reserving_class) & (os_data['RI_TREATY_TYPE'] == ri_type) & (os_data['HEADOFDAMAGE'] == head_of_damage)]
                os_claim_amounts = []
                for accident_period in reserving_class_claims.index:
                    os_claim_amount = os_data_reserving_class[(os_data_reserving_class['LOSSDATE'].dt.to_period('Q') == accident_period) &
                                                              (os_data_reserving_class['As at'].dt.to_period('Q') == eop.to_period('Q'))]['Amount'].sum()
                    os_claim_amounts.append(os_claim_amount)

                reserving_class_claims['OS Claims'] = os_claim_amounts
                reserving_class_claims['Reported Claims'] = reserving_class_claims['Paid Claims'] + reserving_class_claims['OS Claims']

                reserving_class_data = pd.merge(summed_ep, reserving_class_claims, on='Accident_Period', how='left').fillna(0)
                reserving_class_data['Reported LR'] = (reserving_class_data['Reported Claims'] / reserving_class_data['EP'].replace(0, np.nan)).fillna(0)
                #reserving_class_data['Reported LR'] = reserving_class_data.apply(lambda row: '0%' if row['EP'] == 0 else '{:.0f}%'.format(row['Reported LR']), axis=1)
                
                
                end_date_str = eop.strftime('%Y-%m')
                output_file_path = os.path.join(output_dir, output_file_name)
                with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
                    reserving_class_data.to_excel(writer, index=False, sheet_name='Reserve Summary')
                    
                    if paid_data is not None:
                        paid_data_reserving_class = paid_data[(paid_data['RESERVINGCLASS'] == reserving_class) & (paid_data['RI_TREATY_TYPE'] == ri_type) & (paid_data['HEADOFDAMAGE'] == head_of_damage)]
                        triangle_df_paid = calculate_incremental_triangle(paid_data_reserving_class, start_period, end_period)
                        triangle_df_paid.to_excel(writer, sheet_name='Paid Claims Triangle', index=False)

                        cumulative_triangle_df_paid = calculate_cumulative_triangle(triangle_df_paid)
                        cumulative_triangle_df_paid.to_excel(writer, sheet_name='Paid Claims Triangle', startrow=len(triangle_df_paid) + 3, index=False)

                        age_to_age_paid_df = calculate_age_to_age_factors(cumulative_triangle_df_paid)
                        start_row_age_to_age = len(triangle_df_paid) + 3 + len(cumulative_triangle_df_paid) + 3
                        age_to_age_paid_df.to_excel(writer, sheet_name='Paid Claims Triangle', startrow=start_row_age_to_age, index=False)

                        numeric_age_to_age = age_to_age_paid_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
                        simple_paid_averages = numeric_age_to_age.mean(axis=0)
                        simple_paid_cdf = simple_paid_averages[::-1].cumprod()[::-1]
                        avg_paid_df = pd.DataFrame([simple_paid_averages], columns=age_to_age_paid_df.columns)
                        avg_paid_df['Accident Period'] = 'Simple Avg LDF'
                        avg_paid_df.to_excel(writer, sheet_name='Paid Claims Triangle', startrow=len(triangle_df_paid) + 3 + len(cumulative_triangle_df_paid) + 3 + len(age_to_age_paid_df) + 2, index=False)

                        simple_paid_cdf = simple_paid_averages[::-1].cumprod()[::-1]  # reverse cumulative product
                        simple_paid_cdf_df = pd.DataFrame([simple_paid_cdf], columns=cumulative_triangle_df_paid.columns)
                        simple_paid_cdf_df['Accident Period'] = 'Simple Avg CDF'
                        simple_paid_cdf_df.to_excel(writer, sheet_name='Paid Claims Triangle', startrow=len(triangle_df_paid) + 3 + len(cumulative_triangle_df_paid) + 3 + len(age_to_age_paid_df) + 4, index=False, header=False)

                        cumulative_triangle_df_paid_numeric = cumulative_triangle_df_paid.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
                        cumulative_sums = cumulative_triangle_df_paid_numeric.sum(axis=0)
                        weighted_paid_ldfs = []
                        num_accident_periods = cumulative_triangle_df_paid.shape[0]

                        for i in range(1, len(cumulative_sums)):
                            num_periods = num_accident_periods - i
                            if num_periods > 0:
                                numerator = cumulative_triangle_df_paid_numeric.iloc[:num_periods, i].sum()
                                denominator = cumulative_triangle_df_paid_numeric.iloc[:num_periods, i - 1].sum()
                                if denominator != 0:
                                    weighted_paid_ldfs.append(numerator / denominator)
                                else:
                                    weighted_paid_ldfs.append(np.nan)
                            else:
                                weighted_paid_ldfs.append(np.nan)
                        weighted_paid_ldfs.insert(0, np.nan)
                        weighted_avg_ldf_df = pd.DataFrame([weighted_paid_ldfs], columns=cumulative_triangle_df_paid.columns[1:])
                        weighted_avg_ldf_df['Accident Period'] = 'Weighted Avg LDF'
                        weighted_avg_ldf_df = weighted_avg_ldf_df[['Accident Period'] + list(cumulative_triangle_df_paid.columns[1:])]
                        start_row_weighted_ldf = start_row_age_to_age + len(age_to_age_paid_df) + 5
                        weighted_avg_ldf_df.to_excel(writer, sheet_name='Paid Claims Triangle', startrow=start_row_weighted_ldf, index=False, header=False)

                        weighted_paid_ldfs_series = pd.Series(weighted_paid_ldfs)
                        weighted_paid_cdf = weighted_paid_ldfs_series[::-1].cumprod()[::-1]
                        weighted_paid_cdf_df = pd.DataFrame([weighted_paid_cdf], columns=cumulative_triangle_df_paid.columns[1:])
                        weighted_paid_cdf_df['Accident Period'] = 'Weighted Avg CDF'
                        weighted_paid_cdf_df = weighted_paid_cdf_df[['Accident Period'] + list(cumulative_triangle_df_paid.columns[1:])]
                        weighted_paid_cdf_df.to_excel(writer, sheet_name='Paid Claims Triangle', startrow=len(triangle_df_paid) + 3 + len(cumulative_triangle_df_paid) + 3 + len(age_to_age_paid_df) + 6, index=False, header=False)
                        
                        selected_ldf_start_row = start_row_weighted_ldf + 3
                        ws = writer.book['Paid Claims Triangle']
                        ws.cell(row=selected_ldf_start_row, column=1, value='Selected LDF')

                        for col_idx in range(2, len(cumulative_triangle_df_paid.columns) + 1):
                            col_letter = get_column_letter(col_idx)
                            ws.cell(row=selected_ldf_start_row, column=col_idx, value=f'=1')  # Placeholder formula

                        # Write Selected CDF formulas
                        selected_cdf_start_row = selected_ldf_start_row + 1
                        ws.cell(row=selected_cdf_start_row, column=1, value='Selected CDF')

                        last_col_letter = get_column_letter(len(cumulative_triangle_df_paid.columns))
                        last_col_ref = f'${last_col_letter}${selected_ldf_start_row}'

                        for col_idx in range(2, len(cumulative_triangle_df_paid.columns) + 1):
                            col_letter = get_column_letter(col_idx)
                            formula = f'=PRODUCT({col_letter}{selected_ldf_start_row}:{last_col_ref})'
                            ws.cell(row=selected_cdf_start_row, column=col_idx, value=formula)
                        

                    if os_data is not None:
                        os_data_reserving_class = os_data[(os_data['RESERVINGCLASS'] == reserving_class) & (os_data['RI_TREATY_TYPE'] == ri_type) & (os_data['HEADOFDAMAGE'] == head_of_damage)]
                        os_triangle_df = calculate_incremental_triangle(os_data_reserving_class, start_period, end_period, is_os=True)
                        #os_triangle_df.to_excel(writer, sheet_name='OS Claims Triangle', index=False)

                        cumulative_os_triangle_df = calculate_cumulative_triangle(os_triangle_df)
                        #cumulative_os_triangle_df.to_excel(writer, sheet_name='OS Claims Triangle', startrow=len(os_triangle_df) + 3, index=False)

                    reported_triangle_df = cumulative_triangle_df_paid.copy()
                    reported_triangle_df.iloc[:, 1:] += os_triangle_df.iloc[:, 1:] if os_data is not None else 0
                    reported_triangle_df.to_excel(writer, sheet_name='Reported Triangle', index=False)

                    age_to_age_reported_df = calculate_age_to_age_factors(reported_triangle_df)
                    age_to_age_reported_df.to_excel(writer, sheet_name='Reported Triangle', startrow=len(reported_triangle_df) + 3, index=False)

                    numeric_age_to_age_reported = age_to_age_reported_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')

                    simple_reported_averages = numeric_age_to_age_reported.mean(axis=0)
                    simple_reported_cdf = simple_reported_averages[::-1].cumprod()[::-1]
                    
                    avg_reported_df = pd.DataFrame([simple_reported_averages], columns=age_to_age_reported_df.columns)
                    avg_reported_df['Accident Period'] = 'Simple Avg LDF'
                    avg_reported_df.to_excel(writer, sheet_name='Reported Triangle', startrow=len(reported_triangle_df) + 3 + len(age_to_age_reported_df) + 2, index=False)

                    simple_reported_cdf = simple_reported_averages[::-1].cumprod()[::-1]
                    simple_reported_cdf_df = pd.DataFrame([simple_reported_cdf], columns=reported_triangle_df.columns)
                    simple_reported_cdf_df['Accident Period'] = 'Simple Avg CDF'
                    simple_reported_cdf_df.to_excel(writer, sheet_name='Reported Triangle', startrow=len(reported_triangle_df) + 3 + len(age_to_age_reported_df) + 4, index=False, header=False)

                    reported_triangle_df_numeric = reported_triangle_df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
                    cumulative_sums_reported = reported_triangle_df_numeric.sum(axis=0)
                    weighted_reported_ldfs = []
                    num_accident_periods = reported_triangle_df.shape[0]
                    for i in range(1, len(cumulative_sums_reported)):
                        num_periods = num_accident_periods - i
                        if num_periods > 0:
                            numerator = reported_triangle_df_numeric.iloc[:num_periods, i].sum()
                            denominator = reported_triangle_df_numeric.iloc[:num_periods, i - 1].sum()
                            if denominator != 0:
                                weighted_reported_ldfs.append(numerator / denominator)
                            else:
                                weighted_reported_ldfs.append(np.nan)
                        else:
                            weighted_reported_ldfs.append(np.nan)
                    weighted_reported_ldfs.insert(0, np.nan)
                    weighted_avg_ldf_df = pd.DataFrame([weighted_reported_ldfs], columns=reported_triangle_df.columns[1:])
                    weighted_avg_ldf_df['Accident Period'] = 'Weighted Avg LDF'
                    weighted_avg_ldf_df = weighted_avg_ldf_df[['Accident Period'] + list(reported_triangle_df.columns[1:])]
                    weighted_avg_ldf_df.to_excel(writer, sheet_name='Reported Triangle', startrow=len(reported_triangle_df) + 3 + len(age_to_age_reported_df) + 5, index=False, header=False)

                    weighted_reported_ldfs_series = pd.Series(weighted_reported_ldfs)
                    weighted_reported_cdf = weighted_reported_ldfs_series[::-1].cumprod()[::-1]
                    weighted_reported_cdf_df = pd.DataFrame([weighted_reported_cdf], columns=reported_triangle_df.columns[1:])
                    weighted_reported_cdf_df['Accident Period'] = 'Weighted Avg CDF'
                    weighted_reported_cdf_df = weighted_reported_cdf_df[['Accident Period'] + list(reported_triangle_df.columns[1:])]
                    weighted_reported_cdf_df.to_excel(writer, sheet_name='Reported Triangle', startrow=len(reported_triangle_df) + 3 + len(age_to_age_reported_df) + 6, index=False, header=False)
                    
                    selected_ldf_start_row_reported = len(reported_triangle_df) + 3 + len(age_to_age_reported_df) + 8
                    ws_reported = writer.book['Reported Triangle']
                    ws_reported.cell(row=selected_ldf_start_row_reported, column=1, value='Selected LDF')

                    for col_idx in range(2, len(reported_triangle_df.columns) + 1):
                        col_letter = get_column_letter(col_idx)
                        ws_reported.cell(row=selected_ldf_start_row_reported, column=col_idx, value=f'=1')  # Placeholder formula

                    # Write Selected CDF formulas for Reported Triangle
                    selected_cdf_start_row_reported = selected_ldf_start_row_reported + 1
                    ws_reported.cell(row=selected_cdf_start_row_reported, column=1, value='Selected CDF')

                    last_col_letter_reported = get_column_letter(len(reported_triangle_df.columns))
                    last_col_ref_reported = f'${last_col_letter_reported}${selected_ldf_start_row_reported}'

                    for col_idx in range(2, len(reported_triangle_df.columns) + 1):
                        col_letter = get_column_letter(col_idx)
                        formula = f'=PRODUCT({col_letter}{selected_ldf_start_row_reported}:{last_col_ref_reported})'
                        ws_reported.cell(row=selected_cdf_start_row_reported, column=col_idx, value=formula)

                #reserve_summary = generate_reserve_summary(reserving_class_data, reserving_class, head_of_damage, ri_type, output_file_path, ibnr_summaries, output_dir)
                #if reserve_summary is not None:
                #    ibnr_summaries.append(reserve_summary)
