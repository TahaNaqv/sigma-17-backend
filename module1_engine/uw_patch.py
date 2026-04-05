"""Apply UW / ULAE / Discount parameters to Combined_Summary.xlsx (openpyxl)."""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any, BinaryIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Sheet names aligned with engine output and dashboard copy
SHEET_UW_SUMMARY = "UW Summary"
SHEET_ULAE_RA = "ULAE-RA"
SHEET_DISCOUNT_RATE = "Discount Rate"

HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_FILL = PatternFill(start_color="008080", end_color="008080", fill_type="solid")

DEFAULT_DISCOUNT_PERIODS = [
    "0-1 Year",
    "1-2 Years",
    "2-3 Years",
    "3-5 Years",
    "5+ Years",
]


def _optional_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    return float(v)


def _norm_key(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    if not t:
        return ""
    try:
        f = float(t)
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return t


def _header_map(ws, row: int = 1) -> dict[str, int]:
    m: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=col).value
        if v is not None and str(v).strip():
            m[str(v).strip()] = col
    return m


def _style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _workbook_bytes_io(path_or_file: str | Path | BinaryIO | bytes) -> io.BytesIO:
    if isinstance(path_or_file, bytes):
        return io.BytesIO(path_or_file)
    if hasattr(path_or_file, "read"):
        data = path_or_file.read()
        if hasattr(path_or_file, "seek"):
            path_or_file.seek(0)
        return io.BytesIO(data)
    return io.BytesIO(Path(path_or_file).read_bytes())


def build_default_payload_template_from_workbook(
    path_or_file: str | Path | BinaryIO | bytes,
) -> dict[str, Any]:
    """
    Read UW Summary (and existing ULAE / Discount sheets if present) to build
    JSON-serializable defaults for the dashboard.
    """
    bio = _workbook_bytes_io(path_or_file)
    xl = pd.ExcelFile(bio, engine="openpyxl")
    if SHEET_UW_SUMMARY not in xl.sheet_names:
        raise ValueError(f"Workbook must contain a '{SHEET_UW_SUMMARY}' sheet.")

    df = pd.read_excel(xl, sheet_name=SHEET_UW_SUMMARY, engine="openpyxl")
    if "RESERVINGCLASS" not in df.columns or "UWY" not in df.columns:
        raise ValueError("UW Summary must have RESERVINGCLASS and UWY columns.")

    exp_ratio: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for _, row in df.iterrows():
        rc = row["RESERVINGCLASS"]
        uwy = row["UWY"]
        k = (_norm_key(rc), _norm_key(uwy))
        if k[0] == "" or k[1] == "" or k in seen:
            continue
        seen.add(k)
        exp_ratio.append(
            {
                "reserving_class": str(rc).strip() if rc is not None else "",
                "uwy": str(uwy).strip() if uwy is not None else "",
                "exp_ratio": None,
                "ri_percent": None,
            }
        )

    classes = sorted({e["reserving_class"] for e in exp_ratio})
    ulae_ra: list[dict[str, Any]] = []
    for c in classes:
        for gross_ri in ("GROSS", "RI"):
            ulae_ra.append(
                {
                    "reserving_class": c,
                    "gross_ri": gross_ri,
                    "ulae_percent": None,
                    "ra_percent": None,
                }
            )

    discount: list[dict[str, Any]] = []
    if SHEET_DISCOUNT_RATE in xl.sheet_names:
        ddf = pd.read_excel(xl, sheet_name=SHEET_DISCOUNT_RATE, engine="openpyxl")
        tp_col = next(
            (c for c in ddf.columns if str(c).strip().lower() in ("time period", "time_period")),
            None,
        )
        cy_col = next(
            (c for c in ddf.columns if "cy" in str(c).lower() and "discount" in str(c).lower()),
            None,
        )
        py_col = next(
            (c for c in ddf.columns if "py" in str(c).lower() and "discount" in str(c).lower()),
            None,
        )
        if tp_col is not None and cy_col is not None and py_col is not None:
            for _, drow in ddf.iterrows():
                tp = drow[tp_col]
                if pd.isna(tp) or str(tp).strip() == "":
                    continue
                discount.append(
                    {
                        "time_period": str(tp).strip(),
                        "cy_discount": float(drow[cy_col]) if pd.notna(drow[cy_col]) else None,
                        "py_discount": float(drow[py_col]) if pd.notna(drow[py_col]) else None,
                    }
                )
    if not discount:
        for p in DEFAULT_DISCOUNT_PERIODS:
            discount.append(
                {"time_period": p, "cy_discount": None, "py_discount": None}
            )

    if SHEET_ULAE_RA in xl.sheet_names:
        udf = pd.read_excel(xl, sheet_name=SHEET_ULAE_RA, engine="openpyxl")
        need = {"RESERVINGCLASS", "GROSS/RI", "ULAE %", "RA %"}
        if need.issubset(set(udf.columns)):
            from_file: list[dict[str, Any]] = []
            for _, urow in udf.iterrows():
                gr = str(urow["GROSS/RI"]).strip().upper()
                if gr not in ("GROSS", "RI"):
                    continue
                rc = urow["RESERVINGCLASS"]
                if pd.isna(rc):
                    continue
                from_file.append(
                    {
                        "reserving_class": str(rc).strip(),
                        "gross_ri": gr,
                        "ulae_percent": float(urow["ULAE %"]) if pd.notna(urow["ULAE %"]) else None,
                        "ra_percent": float(urow["RA %"]) if pd.notna(urow["RA %"]) else None,
                    }
                )
            if from_file:
                ulae_ra = from_file

    return {"exp_ratio": exp_ratio, "ulae_ra": ulae_ra, "discount": discount}


def apply_uw_parameters_to_combined_summary(
    path_or_file: str | Path | BinaryIO | bytes,
    payload: dict[str, Any],
) -> bytes:
    """
    Apply payload to Combined_Summary workbook and return xlsx bytes.

    payload keys (all optional lists):
      - exp_ratio: {reserving_class, uwy, exp_ratio, ri_percent}
      - ulae_ra: {reserving_class, gross_ri, ulae_percent, ra_percent}
      - discount: {time_period, cy_discount, py_discount}
    """
    if isinstance(path_or_file, bytes):
        bio = io.BytesIO(path_or_file)
        wb = load_workbook(bio, data_only=False)
    elif hasattr(path_or_file, "read"):
        raw = path_or_file.read()
        if hasattr(path_or_file, "seek"):
            path_or_file.seek(0)
        wb = load_workbook(io.BytesIO(raw), data_only=False)
    else:
        wb = load_workbook(path_or_file, data_only=False)

    if SHEET_UW_SUMMARY not in wb.sheetnames:
        raise ValueError(f"Workbook must contain '{SHEET_UW_SUMMARY}'.")

    ws = wb[SHEET_UW_SUMMARY]
    h = _header_map(ws, 1)
    rc_col = h.get("RESERVINGCLASS")
    uwy_col = h.get("UWY")
    exp_col = h.get("Exp Ratio")
    ri_col = h.get("RI %")
    if not rc_col or not uwy_col:
        raise ValueError("UW Summary missing RESERVINGCLASS or UWY.")
    if not exp_col or not ri_col:
        raise ValueError("UW Summary missing Exp Ratio or RI % columns.")

    exp_map: dict[tuple[str, str], dict[str, Any]] = {}
    for item in payload.get("exp_ratio") or []:
        rc = _norm_key(item.get("reserving_class"))
        uw = _norm_key(item.get("uwy"))
        if rc and uw:
            exp_map[(rc, uw)] = item

    for r in range(2, ws.max_row + 1):
        rc = _norm_key(ws.cell(row=r, column=rc_col).value)
        uw = _norm_key(ws.cell(row=r, column=uwy_col).value)
        item = exp_map.get((rc, uw))
        if not item:
            continue
        er = item.get("exp_ratio")
        rp = item.get("ri_percent")
        ws.cell(
            row=r,
            column=exp_col,
            value=_optional_float(er) if er is not None and er != "" else None,
        )
        ws.cell(
            row=r,
            column=ri_col,
            value=_optional_float(rp) if rp is not None and rp != "" else None,
        )

    # ULAE-RA sheet: replace or create
    if SHEET_ULAE_RA in wb.sheetnames:
        del wb[SHEET_ULAE_RA]
    uls = wb.create_sheet(SHEET_ULAE_RA)
    uls.append(["RESERVINGCLASS", "GROSS/RI", "ULAE %", "RA %"])
    _style_header_row(uls, 1)
    for item in payload.get("ulae_ra") or []:
        rc = item.get("reserving_class")
        gr = str(item.get("gross_ri", "")).strip().upper()
        if gr not in ("GROSS", "RI"):
            continue
        ul = item.get("ulae_percent")
        ra = item.get("ra_percent")
        uls.append([rc, gr, _optional_float(ul), _optional_float(ra)])

    # Discount Rate sheet
    if SHEET_DISCOUNT_RATE in wb.sheetnames:
        del wb[SHEET_DISCOUNT_RATE]
    dws = wb.create_sheet(SHEET_DISCOUNT_RATE)
    dws.append(["Time Period", "CY Discount", "PY Discount"])
    _style_header_row(dws, 1)
    for item in payload.get("discount") or []:
        tp = item.get("time_period")
        if not tp:
            continue
        cy = item.get("cy_discount")
        py = item.get("py_discount")
        dws.append([tp, _optional_float(cy), _optional_float(py)])

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()


def strip_composite_reserving_class(label: str) -> str:
    """Strip trailing ' GROSS' / ' RI' from UI labels for UW row matching."""
    s = (label or "").strip()
    s = re.sub(r"\s+(GROSS|RI)\s*$", "", s, flags=re.IGNORECASE)
    return s.strip()
