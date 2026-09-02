"""The Reserve Summary's appended formulas.

`run_update_reserve_summary` appends thirteen columns of live Excel formulas to each reserve
workbook. Its append START column was always computed from the header row, but the formulas
themselves hardcoded column letters (G, B, H, ...). Adding any base column would therefore
have shifted the appended block underneath its own formulas and silently corrupted every
workbook — and **no golden covers this code path**, so nothing would have caught it.

The letters are now derived from the header row by name. These tests pin the generated
strings byte-for-byte against the historic output, and prove the derivation still holds when
a base column is added.
"""

import io

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from module1_engine.engine import run_update_reserve_summary

BASE_HEADERS = [
    "Accident_Period", "EP", "Paid Claims", "OS Claims", "Reported Claims", "Reported LR",
]
APPENDED = [
    "Implied LR", "Paid CDF", "Reported CDF", "Paid CL Ultimate", "Reported CL Ultimate",
    "ELR Ultimate", "Paid BF Ultimate", "Reported BF Ultimate", "Selected Method",
    "Ultimate Claims", "IBNR", "ULR", "CDF",
]

#: Exactly what the engine produced before the letters were made positional.
HISTORIC_ROW_2 = {
    "ELR Ultimate": "=IFERROR(G2 * B2,0)",
    "Paid BF Ultimate": "=IFERROR((1-1/H2) * B2 * G2 + D2, 0)",
    "Reported BF Ultimate": "=IFERROR((1-1/I2) * B2 * G2 + E2, 0)",
    "Ultimate Claims": (
        '=IF(O2="Paid CL", J2, IF(O2="Reported CL", K2, '
        '=IF(O2="ELR", L2, IF(O2="Reported BF", N2, M2))))'
    ),
    "IBNR": "=IFERROR(P2 - E2,0)",
    "ULR": "=IFERROR(P2/B2,0)",
    "CDF": "=IFERROR(P2/C2,0)",
}


def _reserve_workbook(tmp_path, extra_base_columns=()):
    """A minimal reserve workbook of the shape run_generate_summary produces."""
    wb = Workbook()
    summary = wb.active
    summary.title = "Reserve Summary"
    summary.append(BASE_HEADERS + list(extra_base_columns))
    summary.append(["2024-Q1", 1000.0, 400.0, 100.0, 500.0, 0.5] + [0.0] * len(extra_base_columns))
    summary.append(["2024-Q2", 2000.0, 300.0, 200.0, 500.0, 0.25] + [0.0] * len(extra_base_columns))

    for name in ("Paid Claims Triangle", "Reported Triangle"):
        ws = wb.create_sheet(name)
        ws.append(["Accident Period", 0, 1])
        ws.append(["2024-Q1", 100, 150])
        ws.append(["Selected LDF", 1.2, 1.0])
        ws.append(["Selected CDF", 1.2, 1.0])

    path = tmp_path / "Motor Payment GROSS 2024-12.xlsx"
    wb.save(path)
    return path


def _formulas(path, row=2):
    wb = load_workbook(path, data_only=False)
    ws = wb["Reserve Summary"]
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=c).value
    ]
    return {
        name: ws.cell(row=row, column=i + 1).value
        for i, name in enumerate(headers)
    }


def test_generated_formulas_match_the_historic_strings(tmp_path):
    """Byte-for-byte, on the six-column sheet the engine produces today."""
    path = _reserve_workbook(tmp_path)
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)
    for name, expected in HISTORIC_ROW_2.items():
        if name == "Ultimate Claims":
            continue  # asserted separately; the historic literal is long
        assert produced[name] == expected, name


def test_the_ultimate_claims_branch_formula_is_unchanged(tmp_path):
    path = _reserve_workbook(tmp_path)
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)["Ultimate Claims"]
    assert produced == (
        '=IF(O2="Paid CL", J2, IF(O2="Reported CL", K2, '
        'IF(O2="ELR", L2, IF(O2="Reported BF", N2, M2))))'
    )


def test_the_appended_block_starts_where_it_always_did(tmp_path):
    path = _reserve_workbook(tmp_path)
    run_update_reserve_summary(str(tmp_path))
    wb = load_workbook(path)
    ws = wb["Reserve Summary"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert headers[: len(BASE_HEADERS)] == BASE_HEADERS
    assert headers[len(BASE_HEADERS):] == APPENDED
    assert get_column_letter(len(BASE_HEADERS) + 1) == "G"


def test_adding_a_base_column_shifts_the_formulas_instead_of_corrupting_them(tmp_path):
    """The regression this fix exists to prevent.

    With a seventh base column the appended block starts at H, so every reference must move
    one letter. Under the old hardcoded letters the formulas would have kept pointing at G —
    which is now the new base column — and produced silently wrong ultimates.
    """
    path = _reserve_workbook(tmp_path, extra_base_columns=["Large Loss"])
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)

    # Implied LR has moved G -> H, EP is still B.
    assert produced["ELR Ultimate"] == "=IFERROR(H2 * B2,0)"
    # Paid CDF H -> I; OS Claims still D.
    assert produced["Paid BF Ultimate"] == "=IFERROR((1-1/I2) * B2 * H2 + D2, 0)"
    # Ultimate Claims P -> Q; Reported Claims still E.
    assert produced["IBNR"] == "=IFERROR(Q2 - E2,0)"
    assert produced["CDF"] == "=IFERROR(Q2/C2,0)"
    # Crucially, nothing still points at G, which is now Large Loss.
    for name in ("ELR Ultimate", "Paid BF Ultimate", "IBNR", "ULR", "CDF"):
        assert "G2" not in str(produced[name]), f"{name} still references the new base column"


# ---------------------------------------------------------------------------
# The large-claim add-back
# ---------------------------------------------------------------------------
#
# `exclude_and_add_back` is the DEFAULT mode, and for a while it was routed no further than
# the triangle filter — the add-back itself was never applied, so the default silently
# behaved as `exclude_entirely` and understated every ultimate. Every unit test on
# ExclusionPlan passed throughout, because none of them reached the workbook. These do.


def _add_back_workbook(tmp_path, large_paid, large_os):
    """A reserve workbook carrying the two columns the add-back mode writes."""
    wb = Workbook()
    summary = wb.active
    summary.title = "Reserve Summary"
    summary.append(BASE_HEADERS + ["Large Paid", "Large OS"])
    #                AP        EP      Paid    OS     Reported  LR   LgPd  LgOS
    summary.append(["2024-Q1", 1000.0, 400.0, 100.0, 500.0, 0.5, large_paid, large_os])

    for name in ("Paid Claims Triangle", "Reported Triangle"):
        ws = wb.create_sheet(name)
        ws.append(["Accident Period", 0, 1])
        ws.append(["2024-Q1", 100, 150])
        ws.append(["Selected LDF", 1.5, 1.0])
        ws.append(["Selected CDF", 1.5, 1.0])

    path = tmp_path / "Motor Payment GROSS 2024-12.xlsx"
    wb.save(path)
    return path


def test_the_add_back_columns_change_the_chain_ladder_ultimate(tmp_path):
    """Attritional base x attritional CDF, then the large claims at known incurred."""
    path = _add_back_workbook(tmp_path, large_paid=150.0, large_os=50.0)
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)

    cdf = produced["Paid CDF"]
    # (400 - 150) * cdf + (150 + 50)
    assert produced["Paid CL Ultimate"] == pytest.approx((400.0 - 150.0) * cdf + 200.0)

    rep_cdf = produced["Reported CDF"]
    # (500 - 200) * rep_cdf + 200
    assert produced["Reported CL Ultimate"] == pytest.approx((500.0 - 200.0) * rep_cdf + 200.0)


def test_a_large_claim_carries_no_ibnr_of_its_own(tmp_path):
    """Its case reserve is taken as its ultimate — the standard treatment, and the reason
    its development must not run through an attritional factor.

    Compare a book where the whole reported amount is one large claim: the reported chain
    ladder must return exactly that amount, developed by nothing."""
    path = _add_back_workbook(tmp_path, large_paid=400.0, large_os=100.0)
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)
    assert produced["Reported CL Ultimate"] == pytest.approx(500.0)
    # IBNR is a live formula: Ultimate - Reported. With Reported CL selected it is zero.
    assert produced["Reported CL Ultimate"] - 500.0 == pytest.approx(0.0)


def test_absent_add_back_columns_reduce_to_the_historic_expression(tmp_path):
    """The bit-identity guard. Six base columns, no large-claim columns, base * CDF."""
    path = _reserve_workbook(tmp_path)
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)
    assert produced["Paid CL Ultimate"] == pytest.approx(400.0 * produced["Paid CDF"])
    assert produced["Reported CL Ultimate"] == pytest.approx(500.0 * produced["Reported CDF"])


def test_blank_add_back_cells_are_treated_as_zero(tmp_path):
    """A slice no excluded claim touches leaves the cells empty, not 0."""
    path = _add_back_workbook(tmp_path, large_paid=None, large_os=None)
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)
    assert produced["Paid CL Ultimate"] == pytest.approx(400.0 * produced["Paid CDF"])


def test_the_add_back_formulas_shift_for_the_two_extra_base_columns(tmp_path):
    """Eight base columns push the appended block to I; nothing may still point at G/H."""
    path = _add_back_workbook(tmp_path, large_paid=10.0, large_os=5.0)
    run_update_reserve_summary(str(tmp_path))
    produced = _formulas(path)
    assert produced["ELR Ultimate"] == "=IFERROR(I2 * B2,0)"
    assert produced["IBNR"] == "=IFERROR(R2 - E2,0)"
    for name in ("ELR Ultimate", "Paid BF Ultimate", "Reported BF Ultimate", "IBNR", "ULR", "CDF"):
        assert "G2" not in str(produced[name]) and "H2" not in str(produced[name]), name
